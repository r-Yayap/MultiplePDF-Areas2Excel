import os
import pandas as pd
from tkinter import Tk, filedialog,messagebox
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def choose_directory(title="Select Directory"):
    root = Tk()
    root.withdraw()  # Hide the main window
    selected_path = filedialog.askdirectory(title=title)
    root.destroy()  # Close the main window
    return selected_path


def choose_file_save_location(title="Save Excel File As"):
    root = Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title=title
    )
    root.destroy()  # Close the main window
    return file_path

def list_files(start_path):
    file_dict = {'PDF': {}, 'DWG': {}}
    for root, dirs, files in os.walk(start_path):
        for file in files:
            file_path = os.path.join(root, file)
            directory = os.path.relpath(root, start_path)
            file_name, file_extension = os.path.splitext(file)
            file_type = 'PDF' if file_extension.lower() == '.pdf' else 'DWG' if file_extension.lower() == '.dwg' else 'Other'
            file_size = os.path.getsize(file_path)  # Get file size in bytes
            file_modified = os.path.getmtime(file_path)  # Get file modified time

            # Exclude 'Other' file types
            if file_type in ['PDF', 'DWG']:
                # Store file name, size, and modification time in the dictionary
                file_dict[file_type][file_path] = {
                    'name': file_name,
                    'size': file_size,
                    'modified': datetime.fromtimestamp(file_modified)  # Convert timestamp to datetime
                }

    return file_dict


def prepare_data_for_export(file_dict, selected_folder):
    pdf_dict = file_dict['PDF']
    dwg_dict = file_dict['DWG']

    # Combine PDF and DWG dictionaries
    combined_dict = {}

    # Helper function to add or update entries in the combined_dict
    def update_combined_dict(file_name, pdf_data=None, dwg_data=None):
        if file_name not in combined_dict:
            combined_dict[file_name] = {
                'File': file_name, 'PDF': None, 'DWG': None,
                'FolderPDF': [], 'FolderDWG': [], 'PDFSize': None, 'DWGSize': None,
                'PDFModified': None, 'DWGModified': None, 'PDFDuplicateCount': None,
                'DWGDuplicateCount': None
            }

        # Add PDF information
        if pdf_data:
            pdf_name = pdf_data['name']
            combined_dict[file_name]['PDF'] = pdf_name
            combined_dict[file_name]['PDFSize'] = pdf_data['size']
            combined_dict[file_name]['PDFModified'] = pdf_data['modified']

            # Append the folder path to FolderPDF
            pdf_folder = os.path.relpath(os.path.dirname(pdf_data['path']), selected_folder)
            combined_dict[file_name]['FolderPDF'].append(pdf_folder)
            if len(combined_dict[file_name]['FolderPDF']) > 1:
                combined_dict[file_name]['PDFDuplicateCount'] = len(combined_dict[file_name]['FolderPDF'])

        # Add DWG information
        if dwg_data:
            dwg_name = dwg_data['name']
            combined_dict[file_name]['DWG'] = dwg_name
            combined_dict[file_name]['DWGSize'] = dwg_data['size']
            combined_dict[file_name]['DWGModified'] = dwg_data['modified']

            # Append the folder path to FolderDWG
            dwg_folder = os.path.relpath(os.path.dirname(dwg_data['path']), selected_folder)
            combined_dict[file_name]['FolderDWG'].append(dwg_folder)
            if len(combined_dict[file_name]['FolderDWG']) > 1:
                combined_dict[file_name]['DWGDuplicateCount'] = len(combined_dict[file_name]['FolderDWG'])

    # Process PDF files
    for pdf_path, pdf_info in pdf_dict.items():
        file_name = os.path.splitext(os.path.basename(pdf_path))[0]
        dwg_info = dwg_dict.get(pdf_path, None)
        update_combined_dict(file_name, {'path': pdf_path, **pdf_info}, dwg_info)

    # Process DWG files
    for dwg_path, dwg_info in dwg_dict.items():
        file_name = os.path.splitext(os.path.basename(dwg_path))[0]
        pdf_info = pdf_dict.get(dwg_path, None)
        update_combined_dict(file_name, pdf_info, {'path': dwg_path, **dwg_info})

    # Create a DataFrame from the combined dictionary
    df = pd.DataFrame.from_dict(combined_dict, orient='index').reset_index(drop=True)

    # Flatten folder lists into a readable format for Excel
    df['FolderPDF'] = df['FolderPDF'].apply(lambda x: ', '.join(x) if x else None)
    df['FolderDWG'] = df['FolderDWG'].apply(lambda x: ', '.join(x) if x else None)

    # Drop the 'File' column
    df = df.drop(columns=['File'])

    return df


def save_to_excel(df, excel_file):
    # Save the DataFrame to Excel
    df.to_excel(excel_file, index=False)

    # Load the workbook and worksheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define red fill for duplicates
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply conditional formatting for PDFDuplicateCount column
    pdf_col = df.columns.get_loc('PDFDuplicateCount') + 1  # Excel column index (1-based)
    dwg_col = df.columns.get_loc('DWGDuplicateCount') + 1  # Excel column index (1-based)

    ws.conditional_formatting.add(
        f"{chr(64 + pdf_col)}2:{chr(64 + pdf_col)}{len(df) + 1}",
        CellIsRule(operator="greaterThan", formula=["1"], stopIfTrue=True, fill=red_fill)
    )

    # Apply conditional formatting for DWGDuplicateCount column
    ws.conditional_formatting.add(
        f"{chr(64 + dwg_col)}2:{chr(64 + dwg_col)}{len(df) + 1}",
        CellIsRule(operator="greaterThan", formula=["1"], stopIfTrue=True, fill=red_fill)
    )

    # Save the workbook
    wb.save(excel_file)
    wb.close()
    print(f"Directory listing exported to {excel_file}")


def pdf_dwg_counter():
    # Choose the directory using a dialog box
    directory_path = choose_directory()

    if directory_path:
        # Get the dictionary of files in the directory
        file_dict = list_files(directory_path)

        # Choose where to save the Excel file
        excel_file_path = choose_file_save_location()

        if excel_file_path:
            # Prepare the data
            df = prepare_data_for_export(file_dict, directory_path)

            # Save the data to the specified Excel file
            save_to_excel(df, excel_file_path)

            # Prompt to open the Excel file
            open_file = messagebox.askyesno("Open Excel File", f"Do you want to open the Excel file now?")
            if open_file:
                os.startfile(excel_file_path)
        else:
            print("No file location selected.")
    else:
        print("No directory selected.")


if __name__ == "__main__":
    pdf_dwg_counter()