#extractor.py

import multiprocessing
import os
import re
import shutil
import secrets
import pymupdf as fitz
import sys
import csv
import glob
import gc
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils import adjust_coordinates_for_rotation, find_tessdata
import psutil #for debug


# Define patterns
REVISION_REGEX = re.compile(r"^[A-Z]{1,2}\d{1,2}[a-zA-Z]?$", re.IGNORECASE)
DATE_REGEX = re.compile(r"""
    (?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}) |                    # e.g. 07/05/25 or 19-03-2025
    (?:\d{1,2}\s*[-]?\s*[A-Za-z]{3,9}\s*[-]?\s*\d{2,4})    # e.g. 7 Apr 25 or 7-Apr-2025
""", re.VERBOSE | re.IGNORECASE)
DESC_KEYWORDS = ["issued for", "issue", "submission", "schematic", "detailed", "concept", "design", "construction", "revised", "resubmission","ifc","tender","addendum"]

def print_ram(): #for debug
    process = psutil.Process(os.getpid())
    mem_mb = process.memory_info().rss / 1024 / 1024
    print(f"📊 Current RAM usage: {mem_mb:.2f} MB")

def process_single_pdf_standalone(pdf_path, areas, revision_area, ocr_settings, pdf_folder, temp_image_folder, unid, revision_regex):
    extractor = TextExtractor(
        pdf_folder=pdf_folder,
        output_excel_path="",
        areas=areas,
        ocr_settings=ocr_settings,
        revision_regex=revision_regex,
        batch_threshold=0
    )
    extractor.revision_area = revision_area
    extractor.temp_image_folder = temp_image_folder

    result_rows = extractor.process_single_pdf(pdf_path)

    csv_path = os.path.join(temp_image_folder, f"temp_{unid}.csv")
    jsonl_path = os.path.join(temp_image_folder, f"temp_{unid}.ndjson")

    # Stream directly to temp CSV/NDJSON (per worker)
    with open(csv_path, "w", newline="", encoding="utf-8") as csv_file:
        jsonl_file = None
        if revision_area:  # ✅ Only open NDJSON if revision_area is defined
            jsonl_file = open(jsonl_path, "w", encoding="utf-8")

        csv_writer = csv.writer(csv_file)

        for row in result_rows:
            # fill missing elements
            if len(row) < 8:
                row = list(row) + [None] * (8 - len(row))
            file_size, mod_date, folder, filename, page_no, areas, revisions, page_rect = row

            # Defensive fix for page_rect (if needed, as before)
            if page_rect is not None and isinstance(page_rect, (list, tuple)) and len(page_rect) == 4:
                page_rect = fitz.Rect(page_rect)
            else:
                page_rect = None

            page_size_str = f"{page_rect.width:.1f} x {page_rect.height:.1f}" if page_rect else ""

            # Defensive check: revisions should be a list, last item should be a dict
            if isinstance(revisions, list) and len(revisions) > 0 and isinstance(revisions[-1], dict):
                latest_rev = revisions[-1].get('rev', '')
                latest_desc = revisions[-1].get('desc', '')
                latest_date = revisions[-1].get('date', '')
            else:
                latest_rev = ""
                latest_desc = ""
                latest_date = ""

            # Defensive: ensure areas is a list of tuples (text, img_path)
            if not isinstance(areas, list):
                areas = []
            areas = [(a if isinstance(a, tuple) and len(a) == 2 else (str(a), "")) for a in areas]

            text_values = [text for text, _ in areas]
            if len(text_values) < len(extractor.areas):
                text_values += [""] * (len(extractor.areas) - len(text_values))

            revisions_flat = [
                f"{r.get('rev', '')} | {r.get('desc', '')} | {r.get('date', '')}" if isinstance(r, dict) else str(r)
                for r in revisions or []
            ]

            csv_row = [unid, file_size, mod_date, folder, filename, page_no, page_size_str] + \
                      text_values + [latest_rev, latest_desc, latest_date, revisions_flat]

            csv_writer.writerow(csv_row)

            # NDJSON Row (optional)
            if jsonl_file:
                json.dump({"unid": unid, "revisions": revisions}, jsonl_file, ensure_ascii=False)
                jsonl_file.write("\n")

        if jsonl_file:
            jsonl_file.close()

    extractor = None
    gc.collect()

    return True

def _unwrap_process_single_pdf(args):
    return process_single_pdf_standalone(*args)


class TextExtractor:
    def __init__(self, pdf_folder, output_excel_path, areas, ocr_settings, revision_regex=None, batch_threshold=100):
        self.header_column_map = None
        self.final_output_path = None
        self.pdf_folder = pdf_folder
        self.output_excel_path = output_excel_path
        self.areas = areas
        self.ocr_settings = ocr_settings

        self.tessdata_folder = ocr_settings.get("tessdata_folder") or find_tessdata()
        if not self.tessdata_folder:
            print("❌ Tessdata folder not found. OCR will not work.")

        self.revision_regex = re.compile(revision_regex, re.IGNORECASE) if revision_regex else REVISION_REGEX

        self.batch_threshold = batch_threshold

        # Initialize headers with fixed metadata columns
        self.headers = ["Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No", "Page Size"]

        # Dictionary to store unique header assignments per area
        self.unique_headers_mapping = {}  # Maps each rectangle index to a unique column header

        header_count = {}  # To count occurrences of each title

        for i, area in enumerate(self.areas):
            title = area.get("title", f"Area {i + 1}")  # Default title if not manually set

            # Ensure unique title for each rectangle
            if title not in header_count:
                header_count[title] = 1
                unique_title = title
            else:
                header_count[title] += 1
                unique_title = f"{title} ({header_count[title]})"

            self.headers.append(unique_title)  # Add unique header to headers list
            self.unique_headers_mapping[i] = unique_title  # Assign rectangle index to its specific column

        # Define application directory (where the script is located)


        if getattr(sys, 'frozen', False):  # Check if running as PyInstaller EXE
            app_directory = os.path.dirname(sys.executable)  # Use the real EXE location
        else:
            app_directory = os.path.dirname(os.path.abspath(__file__))  # Running as script

        # Create a secure temp folder with random name
        main_temp_folder = os.path.join(app_directory, "temp")
        os.makedirs(main_temp_folder, exist_ok=True)

        # Create random subfolder name (16 random chars)
        random_str = secrets.token_hex(8)
        self.temp_image_folder = os.path.join(main_temp_folder, random_str)

        # Revision Area
        self.revision_area = None  # Will be set externally if needed
        self.unid_counter = 10001


    def detect_column_indices(self, sample_rows, max_rows=3):
        col_scores = {}

        for row in sample_rows[:max_rows]:
            for col_idx, cell in enumerate(row):
                text = str(cell).strip().lower()
                if not text:
                    continue
                if col_idx not in col_scores:
                    col_scores[col_idx] = {"rev": 0, "date": 0, "desc": 0}

                if self.revision_regex.fullmatch(text.upper()):
                    col_scores[col_idx]["rev"] += 1
                elif DATE_REGEX.search(text):
                    col_scores[col_idx]["date"] += 1
                elif any(kw in text for kw in DESC_KEYWORDS):
                    col_scores[col_idx]["desc"] += 1

        # ─── Pick each column once (winner-takes-slot, then remove) ───
        scores = {k: v.copy() for k, v in col_scores.items()}  # make a mutable copy

        def pick(metric):
            """
            Return the column index that scores highest on the chosen metric,
            but only if its score is > 0. Remove that column so it can't be
            selected again.
            """
            if not scores:
                return None
            best = max(scores, key=lambda c: scores[c][metric])
            if scores[best][metric] == 0:  # no evidence → treat as 'not found'
                return None
            scores.pop(best)
            return best

        rev_idx = pick("rev")  # first pass
        desc_idx = pick("desc")  # second pass, cannot reuse rev_idx
        date_idx = pick("date")  # third pass, cannot reuse previous two

        #print(f"🔍 Column Indices Detected → Rev: {rev_idx}, Desc: {desc_idx}, Date: {date_idx}")
        return rev_idx, desc_idx, date_idx

    def parse_revision_row(self, row, rev_idx, desc_idx, date_idx):
        rev = row[rev_idx].strip() if rev_idx is not None and rev_idx < len(row) else None
        desc = row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else None
        date = row[date_idx].strip() if date_idx is not None and date_idx < len(row) else None

        print(f"🧪 Raw values → rev: {repr(rev)}, desc: {repr(desc)}, date: {repr(date)}")

        # Validate formats
        if rev and not self.revision_regex.fullmatch(rev.strip().upper()):
            print(f"⚠️ Rejected rev: {repr(rev)} by pattern: {self.revision_regex.pattern}")
            rev = None
        if date and not DATE_REGEX.search(date):
            print(f"⚠️ Rejected date: {repr(date)}")
            date = None

        print(f"🎯 Validated → rev: {rev}, desc: {desc}, date: {date}")
        return rev, desc, date

    def is_footer_or_header_row(self, row, rev_idx=0):
        """
        Detects if a row is a footer/header or not real revision data.

        Exclude rows that:
        - Are mostly empty, OR
        - Have unwanted keywords *and* do NOT have a valid revision pattern in the revision column.
        """

        normalized = [str(cell).strip().lower() if cell else "" for cell in row]

        unwanted_keywords = [
            "revision", "rev date description", "by chk'd",
            "date", "description", "no.", "rev", "checked by"
        ]

        empty_count = sum(1 for cell in normalized if cell == "")
        if empty_count >= len(normalized) * 0.75:
            return True

        rev_col_val = ""
        if rev_idx is not None and rev_idx < len(normalized):
            rev_col_val = normalized[rev_idx]

        # If the rev_col_val matches revision pattern, it's likely a real revision row
        if self.revision_regex.fullmatch(rev_col_val.upper()):
            # Valid revision found, so do NOT exclude
            return False

        # Otherwise, if any unwanted keyword is found in the row, exclude it
        for keyword in unwanted_keywords:
            for cell in normalized:
                if keyword in cell:
                    return True

        return False

    def extract_revision_history_from_page_obj(self, page, revision_coordinates):
        try:
            if not revision_coordinates or len(revision_coordinates) != 4:
                return []

            clip_rect = fitz.Rect(revision_coordinates)
            if clip_rect.is_empty or clip_rect.get_area() < 100:
                print(f"⚠️ Revision area too small or empty, skipping.")
                return []

            tables = page.find_tables(clip=clip_rect)
            if not tables or not tables.tables:
                print("❌ No tables found.")
                return []

            table = tables.tables[0]
            data = table.extract()
            if not data or len(data) < 2:
                print("❌ Not enough rows to detect header or parse data.")
                return []

            data = data[::-1]

            # Detect revision column index first
            rev_idx, desc_idx, date_idx = self.detect_column_indices(data)

            # Filter rows by passing detected revision index
            filtered_data = [row for row in data if not self.is_footer_or_header_row(row, rev_idx)]

            rev_idx, desc_idx, date_idx = self.detect_column_indices(filtered_data)

            extracted_list = []

            for row in filtered_data:
                if not any(row):
                    continue
                rev, desc, date = self.parse_revision_row(row, rev_idx, desc_idx, date_idx)
                if rev and desc and date:
                    extracted_list.append({"rev": rev, "desc": desc, "date": date})
                else:
                    print(f"⛔ Skipped incomplete row: {row}")

            return extracted_list

        except Exception as e:
            print(f"❌ Error during revision extraction: {e}")
            return []

    def clean_text(self, text):
        """Cleans text by replacing newlines, stripping, and removing illegal characters."""
        replacement_char = '■'  # Character to replace prohibited control characters

        # Step 1: Replace newline and carriage return characters with a space
        text = text.replace('\n', ' ').replace('\r', ' ')

        # Step 2: Strip leading and trailing whitespace
        text = text.strip()

        # Step 3: Replace prohibited control characters with a replacement character
        text = re.sub(r'[\x00-\x1F\x7F-\x9F]', replacement_char, text)

        # Step 4: Remove extra spaces between words
        return re.sub(r'\s+', ' ', text)

    def stream_to_excel(self, csv_path, final_output_path, max_revisions):
        wb = Workbook()
        ws = wb.active

        base_headers = ["Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No", "Page Size"]
        area_headers = [self.unique_headers_mapping[i] for i in range(len(self.areas))]
        extra_headers = ["Latest Revision", "Latest Description", "Latest Date"]
        revision_headers = [f"Rev{i + 1}" for i in range(max_revisions)]
        headers = ["UNID"] + base_headers + area_headers + extra_headers + revision_headers
        ws.append(headers)

        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header row

            ocr_font = Font(color="FF3300")
            area_start_col = 1 + len(["UNID"] + base_headers)

            for row in reader:
                unid = row[0]
                base = row[1:7]  # Size, Date, Folder, Filename, Page No, Page Size
                areas = row[7:7 + len(self.areas)]
                base_index = 7 + len(self.areas)
                latest_rev = row[base_index] if len(row) > base_index else ""
                latest_desc = row[base_index + 1] if len(row) > base_index + 1 else ""
                latest_date = row[base_index + 2] if len(row) > base_index + 2 else ""
                # revisions_flat is row[base_index + 3], you may or may not need to parse it here

                revisions = eval(row[-1]) if row[-1].startswith("[") else []

                row_data = [unid] + base + areas + [latest_rev, latest_desc, latest_date] + revisions + \
                           [""] * (max_revisions - len(revisions))

                ws.append(row_data)
                current_row = ws.max_row

                # Apply hyperlink styling to filename cell
                folder_col_idx = headers.index("Folder") + 1
                filename_col_idx = headers.index("Filename") + 1

                folder_cell = ws.cell(row=current_row, column=folder_col_idx)
                filename_cell = ws.cell(row=current_row, column=filename_col_idx)

                if folder_cell.value and filename_cell.value:
                    abs_path = os.path.abspath(os.path.join(self.pdf_folder, folder_cell.value, filename_cell.value))
                    filename_cell.hyperlink = abs_path
                    filename_cell.font = Font(color="0000FF", underline="single")

                # ✅ OCR styling + image embedding per cell
                for i in range(len(area_headers)):
                    cell = ws.cell(row=current_row, column=area_start_col + i)

                    if isinstance(cell.value, str) and "_OCR_" in cell.value:
                        cell.font = ocr_font
                        cell.value = cell.value.replace("_OCR_", "").strip()

                    if self.ocr_settings["enable_ocr"] == "Text1st+Image-beta":
                        folder = base[2]  # Folder
                        filename = base[3]  # Filename
                        page_no = base[4]  # Page No

                        image_filename = f"{filename}_page{page_no}_area{i}.png"
                        image_path = os.path.join(self.temp_image_folder, image_filename)

                        if os.path.exists(image_path):
                            try:
                                img = ExcelImage(image_path)
                                img.anchor = f"{get_column_letter(area_start_col + i)}{current_row}"
                                ws.add_image(img)
                            except Exception as e:
                                print(f"⚠️ Failed to embed image {image_filename}: {e}")


        # Save file
        if os.path.exists(self.output_excel_path):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name, file_ext = os.path.splitext(self.output_excel_path)
            output_path = f"{file_name}_{timestamp}{file_ext}"
        else:
            output_path = self.output_excel_path

        wb.save(output_path)
        final_output_path.value = output_path

        # ✅ Copy NDJSON from temp folder to match Excel path
        ndjson_temp_path = os.path.join(self.temp_image_folder, "streamed_revisions.ndjson")
        ndjson_output_path = output_path.replace(".xlsx", "_revisions.ndjson")

        try:
            if os.path.exists(ndjson_temp_path):
                shutil.copy(ndjson_temp_path, ndjson_output_path)
                #print(f"✅ NDJSON revision file saved → {ndjson_output_path}")
            else:
                print(f"⚠️ NDJSON temp file not found at {ndjson_temp_path}")
        except Exception as e:
            print(f"❌ Failed to copy NDJSON file: {e}")

        print(f"✅ Excel saved with hyperlinks → {output_path}")

    def combine_temp_files(self, final_output_path):
        temp_csv_files = sorted(glob.glob(os.path.join(self.temp_image_folder, "temp_*.csv")))
        temp_jsonl_files = sorted(glob.glob(os.path.join(self.temp_image_folder, "temp_*.ndjson")))

        combined_csv = os.path.join(self.temp_image_folder, "streamed_output.csv")
        combined_jsonl = os.path.join(self.temp_image_folder, "streamed_revisions.ndjson")

        # Combine CSV
        with open(combined_csv, "w", newline="", encoding="utf-8") as outfile:
            writer = csv.writer(outfile)
            writer.writerow(
                ["UNID", "Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No", "Page Size"] +
                [self.unique_headers_mapping[i] for i in range(len(self.areas))] +
                ["Latest Revision", "Latest Description", "Latest Date", "__revisions__"]
            )

            for f in temp_csv_files:
                with open(f, "r", encoding="utf-8") as infile:
                    reader = csv.reader(infile)
                    writer.writerows(reader)
                os.remove(f)  # Cleanup immediately

        # ✅ Only combine NDJSON if revision area was defined
        if self.revision_area:
            with open(combined_jsonl, "w", encoding="utf-8") as outfile:
                for f in temp_jsonl_files:
                    with open(f, "r", encoding="utf-8") as infile:
                        shutil.copyfileobj(infile, outfile)
                    os.remove(f)  # Cleanup

            # ✅ Save final copy next to Excel
            ndjson_output_path = final_output_path.value.replace(".xlsx", "_revisions.ndjson")
            try:
                shutil.copy(combined_jsonl, ndjson_output_path)
            except Exception as e:
                print(f"❌ Failed to copy NDJSON file: {e}")

        # Stream to Excel (existing efficient method)
        max_revisions = self.get_max_revisions(combined_csv)
        self.stream_to_excel(combined_csv, final_output_path, max_revisions)

    def get_max_revisions(self, csv_path):
        max_revs = 0
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                revisions = eval(row[-1]) if row[-1].startswith("[") else []
                max_revs = max(max_revs, len(revisions))
        return max_revs

    def start_extraction(self, progress_counter, total_files, final_output_path,selected_paths):
        n_workers = multiprocessing.cpu_count()
        pdf_files = selected_paths
        total_files.value = len(pdf_files)

        os.makedirs(self.temp_image_folder, exist_ok=True)

        jobs = [
            (pdf_path, self.areas, self.revision_area, self.ocr_settings, self.pdf_folder, self.temp_image_folder,
             str(10000 + idx), self.revision_regex.pattern)
            for idx, pdf_path in enumerate(pdf_files)
        ]

        with multiprocessing.Pool(processes=n_workers) as pool:
            for _ in pool.imap_unordered(_unwrap_process_single_pdf, jobs):
                with progress_counter.get_lock():
                    progress_counter.value += 1
                    if progress_counter.value % 50 == 0:
                        print_ram()

        # After multiprocessing ends, combine CSV/NDJSON files quickly
        self.combine_temp_files(final_output_path)



    def process_single_pdf(self, pdf_path):
        """Processes a single PDF file, extracting text and images as necessary."""
        if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
            print(f"Skipping missing or empty file: {pdf_path}")
            result_rows = [[
                0,
                "",
                "",
                os.path.basename(pdf_path),
                0,
                [("Error: File not found or empty", "")],  # list of one tuple
                []
            ]]

            return result_rows

        try:

            # Open the PDF file with additional validation and caching
            try:
                pdf_document = fitz.open(pdf_path, filetype="pdf")
                if not pdf_document.is_pdf:
                    raise ValueError("Not a valid PDF file")
                fitz.TOOLS.set_small_glyph_heights(True)
            except (fitz.FileDataError, fitz.EmptyFileError) as e:
                print(f"PDF corrupted or empty: {pdf_path}, {e}")
                result_rows = [[
                    0,
                    "",
                    "",
                    os.path.basename(pdf_path),
                    0,
                    f"Error: Corrupted or empty PDF: {str(e)}",
                    []
                ]]
                return result_rows

            except Exception as e:
                print(f"Error opening PDF {pdf_path}: {e}")
                result_rows = [[
                    0,
                    "",
                    "",
                    os.path.basename(pdf_path),
                    0,
                    f"Error: Invalid PDF: {str(e)}",
                    []
                ]]
                return result_rows


            file_size = os.path.getsize(pdf_path)
            last_modified_date = datetime.fromtimestamp(os.path.getmtime(pdf_path)).strftime('%Y-%m-%d %H:%M:%S')
            folder = os.path.relpath(os.path.dirname(pdf_path), self.pdf_folder)
            filename = os.path.basename(pdf_path)

            result_rows = []  # Collect each page's data here

            # Process each page safely
            for page_number in range(pdf_document.page_count):
                try:
                    page = pdf_document[page_number]
                    page.remove_rotation()
                    extracted_areas = []

                    for area_index, area in enumerate(self.areas):
                        coordinates = area["coordinates"]
                        text_area, img_path = self.extract_text_from_area(page, coordinates, pdf_path, page_number,
                                                                          area_index)

                        # If text_area is empty, treat it as blank rather than "Error"
                        extracted_areas.append((text_area if text_area != "Error" else "", img_path or ""))

                    revision_data = []
                    if self.revision_area:
                        revision_data = self.extract_revision_history_from_page_obj(
                            page, self.revision_area["coordinates"]
                        )

                    result_rows.append([
                        file_size, last_modified_date, folder, filename, page_number + 1,
                        extracted_areas, revision_data, list(page.rect)  # page size as 8th element
                    ])

                    del extracted_areas
                    del revision_data
                    del page
                    gc.collect()  # ✅ Force garbage collection

                except Exception as e:
                    print(f"Error processing page {page_number + 1} in {pdf_path}: {e}")
                    result_rows.append(
                        [file_size, last_modified_date, folder, filename, page_number + 1,
                         [("Error processing page", "")],
                         [],
                         None  # Add this to keep length consistent
                         ]
                    )

            pdf_document.close()
            del pdf_document
            gc.collect()

            return result_rows

        except Exception as e:
            print(f"Error processing {pdf_path}: {e}")
            # Add an error placeholder for this file in results
            result_rows = [[
                file_size if 'file_size' in locals() else 0,
                last_modified_date if 'last_modified_date' in locals() else "",
                folder if 'folder' in locals() else "",
                filename if 'filename' in locals() else os.path.basename(pdf_path),
                0,
                f"Error: File Processing Error: {str(e)}",
                [],
                None  # Add None here for page.rect to keep consistent
            ]]

            gc.collect()
            # total_files.value += 1  # Update progress counter even on error
            return result_rows

    def get_pdf_files(self, selected_paths=None):
        """Return only explicitly selected PDF files if provided, else fallback to scanning."""
        if selected_paths:
            return selected_paths
        # (optional fallback if needed)
        return []

    def extract_text_from_area(self, page, area_coordinates, pdf_path, page_number, area_index):
        """Extracts text or image content from a specified area in a PDF page."""
        adjusted_coordinates = adjust_coordinates_for_rotation(
            area_coordinates, page.rotation, page.rect.height, page.rect.width
        )
        text_area = ""
        img_path = None

        try:
            if self.ocr_settings["enable_ocr"] == "Default":
                text_area = page.get_text("text", clip=adjusted_coordinates)
                if not text_area.strip():  # Perform OCR only if text_area is empty
                    text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)

            # OCR for all areas, no images saved
            elif self.ocr_settings["enable_ocr"] == "OCR-All":
                text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)

            # OCR with image saving regardless of extracted text for Text1st+Image-beta
            elif self.ocr_settings["enable_ocr"] == "Text1st+Image-beta":
                text_area = page.get_text("text", clip=adjusted_coordinates)
                pix = page.get_pixmap(clip=area_coordinates, dpi=self.ocr_settings.get("dpi_value", 150))

                if not os.path.exists(self.temp_image_folder):
                    os.makedirs(self.temp_image_folder, exist_ok=True)

                # Validate and save image
                img_path = os.path.join(self.temp_image_folder, f"{os.path.basename(pdf_path)}_page{page_number + 1}_area{area_index}.png")

                # Save image with size validation (20MB max)
                pix.save(img_path)
                if os.path.getsize(img_path) > 10 * 1024 * 1024:  # 20MB
                    os.remove(img_path)
                    img_path = None
                    print(f"Warning: Skipped large image (>10MB) from {pdf_path} page {page_number + 1}")

                # Apply OCR if no text found
                if not text_area.strip():
                    try:
                        text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)
                    except Exception as e:
                        print(f"OCR failed for {pdf_path} page {page_number + 1}: {e}")
                        text_area = "OCR_ERROR"

            # Clean the extracted text
            text_area = self.clean_text(text_area)

            # Return empty string if text_area is blank after cleaning
            return (text_area if text_area.strip() else "", img_path)

        except fitz.EmptyFileError as e:
            print(f"Error: {pdf_path} is empty or corrupted: {e}")
            return f"EmptyFileError: {str(e)}", None

        except fitz.FileNotFoundError as e:
            print(f"Error: {pdf_path} not found: {e}")
            return f"FileNotFoundError: {str(e)}", None

        except RuntimeError as e:
            print(f"Runtime error on page {page_number + 1} in {pdf_path}: {e}")
            return f"RuntimeError: {str(e)}", None

        except Exception as e:
            print(f"Unexpected error extracting text from area {area_index} in {pdf_path}, Page {page_number + 1}: {e}")
            return f"UnexpectedError: {str(e)}", None

    def apply_ocr(self, page, coordinates, pdf_path, page_number, area_index):
        """Applies OCR on a specified area and returns the extracted text and image path."""
        if self.tessdata_folder is None:
            self.tessdata_folder = find_tessdata()
            if not self.tessdata_folder:
                print("❌ Tessdata folder not found. OCR cannot proceed.")
                return "", ""

        if not os.path.exists(self.temp_image_folder):
             os.makedirs(self.temp_image_folder, exist_ok=True)

        pix = page.get_pixmap(clip=coordinates, dpi=self.ocr_settings.get("dpi_value", 150))
        pdfdata = pix.pdfocr_tobytes(language="eng", tessdata=self.tessdata_folder)
        clipdoc = fitz.open("pdf", pdfdata)
        text_area = "_OCR_" + clipdoc[0].get_text()
        clipdoc.close()
        del clipdoc
        gc.collect()
        img_path = os.path.join(self.temp_image_folder,f"{os.path.basename(pdf_path)}_page{page_number + 1}_area{area_index}.png")
        pix.save(img_path)
        del pix
        return text_area, img_path

    def extract_revision_history_from_page(self, pdf_path, page, filename, page_number):
        try:
            if (
                    self.revision_area and
                    isinstance(self.revision_area, dict) and
                    "coordinates" in self.revision_area and
                    isinstance(self.revision_area["coordinates"], list) and
                    len(self.revision_area["coordinates"]) == 4
            ):

                revision_data = self.extract_revision_history_from_page_obj(page, self.revision_area["coordinates"])
                self.revision_data_mapping[(filename, page_number + 1)] = revision_data
            else:
                print(f"⚠️ Invalid revision area format for {filename} Page {page_number + 1}: {self.revision_area}")
        except Exception as e:
            print(f"⚠️ Revision extraction failed for {filename} Page {page_number + 1}: {e}")

    def consolidate_results(self, results, final_output_path):
        """Consolidates all extracted results into an Excel file with each page as a row."""
        wb = Workbook()
        ws = wb.active

        # Determine the maximum number of revision entries
        max_revision_columns = max(
            (len(row[6]) for result_pages in results for row in result_pages if
             isinstance(row, list) and len(row) == 7),
            default=0
        )

        # Extend headers dynamically with Rev1, Rev2, ...
        rev_headers = [f"Rev{i + 1}" for i in range(max_revision_columns)]
        self.headers.extend(rev_headers)

        self.header_column_map = {title: idx + 1 for idx, title in enumerate(self.headers)}

        ws.append(self.headers)  # Add headers to the first row

        # Precompute revision column indices to avoid repeated lookups
        rev_col_indices = {
            i: self.header_column_map[header]
            for i, header in enumerate(rev_headers)
        }

        for result_pages in results:
            for row_data in result_pages:
                try:
                    # if not isinstance(row_data, list) or len(row_data) != 7:
                    #     print(f"⚠️ Skipping malformed row_data: {row_data}")
                    #     continue

                    # Unpack metadata and extracted area data for the page
                    file_size, last_modified, folder, filename, page_number, page_rect, extracted_areas, revision_data = row_data

                    # Prepare row for insertion
                    row_index = ws.max_row + 1
                    metadata = [file_size, last_modified, folder, filename, page_number]

                    # Add page size string (e.g. "width x height")
                    page_size_str = f"{page_rect.width:.1f} x {page_rect.height:.1f}" if page_rect else ""

                    metadata.append(page_size_str)  # Now metadata has 6 items

                    # Write metadata (6 columns)
                    for col, data in enumerate(metadata, start=1):
                        cell = ws.cell(row=row_index, column=col)
                        cell.value = data
                        if col == 4:  # hyperlink on filename column remains at 4
                            absolute_path = os.path.abspath(os.path.join(self.pdf_folder, folder, filename))
                            cell.hyperlink = absolute_path
                            cell.style = "Hyperlink"

                    # Write extracted areas to columns starting from 6
                    for index, area in enumerate(self.areas):  # Use the updated self.areas list

                        # Get the unique title assigned to this rectangle
                        column_title = self.unique_headers_mapping.get(index, f"Area {index + 1}")
                        col_index = self.header_column_map[column_title]

                        # Retrieve text and image path from extracted areas
                        text, img_path = extracted_areas[index] if isinstance(extracted_areas, list) else (
                        extracted_areas, None)

                        text_cell = ws.cell(row=row_index, column=col_index)
                        cleaned_text = self.clean_text(text.replace("_OCR_", "")) if text != "Error" else "Error"
                        text_cell.value = cleaned_text

                        # Highlight OCR text in red if "_OCR_" was detected
                        if "_OCR_" in text:
                            text_cell.font = Font(color="FF3300")

                        # If there is an associated image, add it to the cell
                        if img_path and img_path != "Error":
                            img = ExcelImage(img_path)
                            img.anchor = f"{get_column_letter(col_index)}{row_index}"
                            ws.add_image(img)

                    # Append revisions if available
                    for i, revision_text in enumerate(revision_data):
                        if i in rev_col_indices:
                            col_index = rev_col_indices[i]
                            ws.cell(row=row_index, column=col_index).value = revision_text

                except Exception as e:
                    print(f"❌ Unexpected error: {e}")
                    print(f"   ↳ row_data was: {row_data!r}")

                    # pull metadata _directly_ out of this row_data (if possible)
                    size, mod, fld, fn, pg = (row_data + [None] * 5)[:5]

                    # reserve the next row
                    row_index = ws.max_row + 1

                    # rewrite metadata
                    for col, data in enumerate([size, mod, fld, fn, pg], start=1):
                        cell = ws.cell(row=row_index, column=col)
                        cell.value = data
                        if col == 4 and fn:
                            path = os.path.abspath(os.path.join(self.pdf_folder, fld or "", fn))
                            cell.hyperlink = path
                            cell.style = "Hyperlink"

                    # put the error message in column6
                    ws.cell(row=row_index, column=6).value = f"Error: {e}"

                    # blank out the rest of the area columns
                    for c in range(7, 7 + len(self.areas)):
                        ws.cell(row=row_index, column=c).value = ""

        # Generate a unique filename if the output file already exists
        output_filename = self.output_excel_path
        if os.path.exists(output_filename):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name, file_ext = os.path.splitext(output_filename)
            output_filename = f"{file_name}_{timestamp}{file_ext}"

        # Save to the new filename
        try:
            wb.save(output_filename)
            wb.close()
            final_output_path.value = output_filename  # ✅ Save to shared multiprocessing Value

            print(f"Consolidated results saved to {output_filename}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

        # Secure cleanup of temporary images
        if os.path.exists(self.temp_image_folder):
            try:
                print(f"Deleting temp folder: {self.temp_image_folder}")
                shutil.rmtree(self.temp_image_folder)
            except Exception as e:
                print(f"Warning: Failed to delete temp folder: {e}")
                # Try again on next run
