# gui.py
import multiprocessing
import os
import re
import time
import customtkinter as ctk
import sys
from tkinter import filedialog, messagebox, StringVar
from openpyxl import Workbook, load_workbook
from constants import *
from customtkinter import CTkImage
from typing import Any
from pdf_viewer import PDFViewer
from utils import create_tooltip, EditableTreeview
from utils import find_tessdata
from utils import REVISION_PATTERNS
from ttkwidgets import CheckboxTreeview
from tkinter import ttk
from PIL import Image, ImageTk  # Make sure this is at the top

from app.domain.models import AreaSpec, OcrSettings, ExtractionRequest
from app.controllers.extract_controller import ExtractController
from pathlib import Path

import logging
from app.logging_setup import configure_logging
logger = configure_logging()


def _is_spec(a: Any) -> bool:
    try:
        from app.domain.models import AreaSpec
        return isinstance(a, AreaSpec)
    except Exception:
        return False

def _area_title(a: Any) -> str:
    return a.title if _is_spec(a) else a["title"]

def _area_coords(a: Any):
    return list(a.rect) if _is_spec(a) else a["coordinates"]

def _to_spec(a: Any):
    from app.domain.models import AreaSpec
    return a if _is_spec(a) else AreaSpec(title=a["title"], rect=tuple(a["coordinates"]))

def _maybe_to_dict(a: Any) -> dict:
    # convenient when exporting, etc.
    return {"title": a.title, "coordinates": list(a.rect)} if _is_spec(a) else a

def _rev_area_to_spec(a: Any):
    if a is None:
        return None
    return _to_spec(a)


# DnD (safe import)
try:
    from tkinterdnd2 import TkinterDnD, DND_ALL
    DND_ENABLED = True
except Exception as e:
    print("tkdnd not available, drag & drop disabled:", e)
    TkinterDnD = None
    DND_ALL = None
    DND_ENABLED = False

class CTkDnD(ctk.CTk, *( (TkinterDnD.DnDWrapper,) if DND_ENABLED else tuple() )):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if DND_ENABLED:
            self.TkdndVersion = TkinterDnD._require(self)


def _preflight(self) -> tuple[bool, str]:
    if not self.pdf_viewer.areas:
        return False, "No areas defined. Please draw at least one area."
    if not self.pdf_folder or not os.path.isdir(self.pdf_folder):
        return False, "The selected PDF folder does not exist."
    if not self.output_excel_path:
        return False, "Please choose an Excel output path."
    out_dir = os.path.dirname(self.output_excel_path)
    if not os.path.isdir(out_dir):
        return False, "The output folder does not exist."
    if not os.access(out_dir, os.W_OK):
        return False, "You do not have write permission to the output folder."
    # at least one checked PDF:
    checked = [iid for iid in self.files_tree_widget.get_checked()
               if str(self.files_tree_widget.item(iid).get("text","")).lower().endswith(".pdf")]
    if not checked:
        return False, "Please check at least one PDF to extract."
    return True, ""

def resource_path(rel: str) -> str:
    """
    Return an absolute path to a bundled resource that works for:
    • normal `python main.py` runs
    • Nuitka --standalone builds
    • PyInstaller one-file / one-dir builds
    """
    # -------- if we are inside a frozen app ---------------------------
    if getattr(sys, "frozen", False):
        # PyInstaller defines _MEIPASS; other freezers (Nuitka, cx_Freeze) don't.
        base_dir = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    else:
        # running from source – use the directory where *this* file lives
        base_dir = Path(__file__).parent

    return str(base_dir / rel)



class XtractorGUI:
    def __init__(self, root):
        self.root = root

        self.extractor = ExtractController()

        # compute UI scale from Tk (dpi-aware thanks to main.py)
        self._ui_scale = float(self.root.tk.call('tk', 'scaling')) / (96 / 72)  # = 1.0 at 96 DPI
        def px(v: float) -> int: return int(round(v * self._ui_scale))
        self._px = px

        self.pdf_viewer = PDFViewer(self, self.root)  # Pass GUI instance and root window


        self.root.update_idletasks()


        self.pdf_folder = ''
        self.output_excel_path = ''
        self.ocr_settings = {'enable_ocr': 'Default', 'dpi_value': 150, 'tessdata_folder': TESSDATA_FOLDER}
        self.recent_pdf_path = None


        self.setup_widgets()
        self.ocr_menu_callback("Default")
        self.setup_bindings()
        self.setup_tooltips()
        self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)

    def _on_app_close(self):
        try:
            if getattr(self, "cancel_event", None):
                self.cancel_event.set()
            if getattr(self, "extraction_process", None) and self.extraction_process.is_alive():
                self.extraction_process.terminate()
                self.extraction_process.join(2)
        except Exception:
            pass
        try:
            if getattr(self, "current_manager", None):
                self.current_manager.shutdown()
        except Exception:
            pass
        self.root.destroy()

    def export_rectangles(self):
        export_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Rectangles As"
        )
        if not export_file_path:
            return

        try:
            wb = Workbook()
            ws_area = wb.active
            ws_area.title = "Rectangles"
            ws_area.append(["Title", "x0", "y0", "x1", "y1"])
            for area in self.pdf_viewer.areas:
                title = _area_title(area)
                x0, y0, x1, y1 = _area_coords(area)
                ws_area.append([title, x0, y0, x1, y1])

            if self.pdf_viewer.revision_area:
                ws_rev = wb.create_sheet("RevisionTable")
                ws_rev.append(["Title", "x0", "y0", "x1", "y1"])
                ra = self.pdf_viewer.revision_area
                title = _area_title(ra)
                x0, y0, x1, y1 = _area_coords(ra)
                ws_rev.append([title, x0, y0, x1, y1])

            wb.save(export_file_path)
            wb.close()
            print(f"Exported areas to {export_file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export areas: {e}")

    def import_rectangles(self):
        import_file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm"), ("All files", "*.*")],
            title="Import Rectangles"
        )
        if not import_file_path:
            return

        try:
            wb = load_workbook(import_file_path)
            ws_area = wb["Rectangles"] if "Rectangles" in wb.sheetnames else wb.active

            self.pdf_viewer.areas = []
            for row in ws_area.iter_rows(min_row=2, values_only=True):
                title, x0, y0, x1, y1 = row
                self.pdf_viewer.areas.append({"title": title, "coordinates": [x0, y0, x1, y1]})

            # Handle revision area
            revision_area_set = False
            if "RevisionTable" in wb.sheetnames:
                ws_rev = wb["RevisionTable"]
                for row in ws_rev.iter_rows(min_row=2, values_only=True):
                    title, x0, y0, x1, y1 = row
                    # ✅ Only set revision area if all coordinates are present
                    if all(isinstance(coord, (int, float)) for coord in [x0, y0, x1, y1]):
                        self.pdf_viewer.revision_area = {"title": title, "coordinates": [x0, y0, x1, y1]}
                        revision_area_set = True

            if not revision_area_set:
                self.pdf_viewer.revision_area = None  # ✅ Clear revision area if nothing valid was set

            self.pdf_viewer.update_rectangles()
            self.update_areas_treeview()
            print(f"Imported areas from {import_file_path}")
        except Exception as e:
            messagebox.showerror("Import Error", f"Could not import areas: {e}")

    def import_rectangles_from_file(self, file_path):
        try:
            wb = load_workbook(file_path)
            ws_area = wb["Rectangles"] if "Rectangles" in wb.sheetnames else wb.active

            self.pdf_viewer.areas = []
            for row in ws_area.iter_rows(min_row=2, values_only=True):
                title, x0, y0, x1, y1 = row
                self.pdf_viewer.areas.append({"title": title, "coordinates": [x0, y0, x1, y1]})

            # Handle revision area if present
            if "RevisionTable" in wb.sheetnames:
                ws_rev = wb["RevisionTable"]
                for row in ws_rev.iter_rows(min_row=2, values_only=True):
                    title, x0, y0, x1, y1 = row
                    if all(isinstance(coord, (int, float)) for coord in [x0, y0, x1, y1]):
                        self.pdf_viewer.revision_area = {"title": title, "coordinates": [x0, y0, x1, y1]}
                        break
            else:
                self.pdf_viewer.revision_area = None

            self.pdf_viewer.update_rectangles()
            self.update_areas_treeview()
            print(f"✅ Imported areas from dropped Excel file: {file_path}")

        except Exception as e:
            messagebox.showerror("Import Error", f"Could not import areas from dropped Excel file: {e}")

    def clear_all_areas(self):
        """Clears all areas and updates the display."""
        self.pdf_viewer.clear_areas()  # Clear area rectangles
        self.pdf_viewer.revision_area = None  # ✅ Clear revision area rectangle
        self.areas_tree.delete(*self.areas_tree.get_children())  # Clear all entries in the Treeview
        self.pdf_viewer.update_rectangles()  # ✅ Ensure canvas refresh reflects the change
        print("All areas and revision table cleared.")

    def update_areas_treeview(self):
        self.areas_tree.delete(*self.areas_tree.get_children())
        self.treeview_item_ids = {}
        for index, area in enumerate(self.pdf_viewer.areas):
            title = _area_title(area)
            x0, y0, x1, y1 = _area_coords(area)
            item_id = self.areas_tree.insert("", "end", values=(title, x0, y0, x1, y1))
            self.treeview_item_ids[item_id] = index

    def open_sample_pdf(self):
        # Opens a file dialog to select a PDF file, then displays it in the PDFViewer
        pdf_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if pdf_path:
            self.pdf_viewer.display_pdf(pdf_path)
            self.recent_pdf_path = pdf_path  # Store the recent PDF path
            print(f"Opened sample PDF: {pdf_path}")

    def open_recent_pdf(self):
        """Opens the most recently viewed PDF."""
        if self.recent_pdf_path:
            self.pdf_viewer.display_pdf(self.recent_pdf_path)
            print(f"Reopened recent PDF: {self.recent_pdf_path}")
        else:
            messagebox.showinfo("Info", "No recent PDF found.")

    def close_pdf(self):
        """Delegates PDF closing to the PDF viewer."""
        self.pdf_viewer.close_pdf()

    def remove_row(self):
        """Removes the selected row from the Treeview and updates the canvas to remove the associated rectangle."""
        selected_item = self.areas_tree.selection()
        if selected_item:
            # Get the rectangle index associated with the selected Treeview item
            index = self.treeview_item_ids.get(selected_item[0])
            if index is not None:
                # Remove the rectangle from the canvas
                rectangle_id = self.pdf_viewer.rectangle_list[index]
                self.pdf_viewer.canvas.delete(rectangle_id)
                # Remove the area from PDFViewer's areas and rectangle_list
                del self.pdf_viewer.areas[index]
                del self.pdf_viewer.rectangle_list[index]

                # Remove the item from Treeview
                self.areas_tree.delete(selected_item[0])

                # Update Treeview and canvas display
                self.update_areas_treeview()
                self.pdf_viewer.update_rectangles()

                print("Removed rectangle at index", index)

    def setup_widgets(self):
        # Create a tabbed panel on the left
        self.tab_view = ctk.CTkTabview(self.root, width=SIDEBAR_WIDTH)
        self.tab_view.pack(side="left", fill="y", padx=SIDEBAR_PADDING, pady=10)

        self.root.update_idletasks()  # Ensure sidebar dimensions are accurate

        # Zoom Slider
        self.zoom_var = ctk.DoubleVar(value=self.pdf_viewer.current_zoom)  # Initialize with the current zoom level
        self.zoom_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        self.zoom_out_label = ctk.CTkLabel(self.zoom_frame, text="➖", font=(BUTTON_FONT, 10))
        self.zoom_slider = ctk.CTkSlider(self.zoom_frame, from_=0.1, to=4, variable=self.zoom_var,
                                         command=self.update_zoom, width=self._px(170), height=self._px(8))
        self.zoom_in_label = ctk.CTkLabel(self.zoom_frame, text="➕", font=(BUTTON_FONT, 10))

        self.zoom_out_label.pack(side="left", padx=(5, 2))
        self.zoom_slider.pack(side="left")
        self.zoom_in_label.pack(side="left", padx=(2, 5))

        # Floating recent PDF and close PDF buttons (top-right)
        # Create a label that behaves like a button
        self.recent_pdf_button = ctk.CTkLabel(
            self.root,
            text="↩",
            font=(BUTTON_FONT, 12, "bold"),
            text_color="lightblue",
            cursor="hand2",
            width=self._px(24),
            height=self._px(24)
        )
        self.recent_pdf_button.pack(pady=2)
        self.recent_pdf_button.bind("<Button-1>", lambda e: self.open_recent_pdf())

        self.close_pdf_button = ctk.CTkLabel(
            self.root,
            text="X",
            font=(BUTTON_FONT, 10, "bold"),
            text_color="red",
            cursor="hand2",
            width=self._px(24),
            height=self._px(24)
        )
        self.close_pdf_button.pack(pady=2)
        self.close_pdf_button.bind("<Button-1>", lambda e: self.close_pdf())

        # Create each tab
        tab_files = self.tab_view.add("📁 Files")
        tab_rectangles = self.tab_view.add("🔲 Rectangles")
        tab_extract = self.tab_view.add("🚀 Extract")
        tab_tools = self.tab_view.add("🧰 Tools")

        # ======================= 📁 FILES TAB =======================

        # PDF Folder Drop Zone
        self.pdf_folder_button = ctk.CTkButton(tab_files,
                                               text="\n➕\n\nDrop Folder or Click to Browse",
                                               command=self.browse_pdf_folder,
                                               fg_color="transparent", border_width=2, width=240, height=80,
                                               hover_color="#444", text_color="white")
        self.pdf_folder_button.pack(pady=(10, 5))
        if DND_ENABLED:
            try:
                self.pdf_folder_button.drop_target_register(DND_ALL)
                self.pdf_folder_button.dnd_bind('<<Drop>>', self.drop_pdf_folder)
            except Exception as e:
                print("Could not enable DnD on pdf_folder_button:", e)

        self.pdf_folder_entry = ctk.CTkEntry(self.root, width=240, height=24, font=(BUTTON_FONT, 9),
                                             placeholder_text="Select Folder with PDFs", border_width=1,
                                             corner_radius=3)
        self.pdf_folder_entry.place(x=-200, y=-223) #hide haha
        # Files Tree Label
        self.files_tree_label = ctk.CTkLabel(tab_files, text="📄 Files in Folder", font=(BUTTON_FONT, 10))
        self.files_tree_label.pack(pady=(10, 0))

        # Create Frame to hold the tree
        self.files_tree_frame = ctk.CTkFrame(tab_files, height=140, width=240)
        self.files_tree_frame.pack(pady=(2, 5), fill="both", expand=True)

        # 🧱 Create scrollable frame ONCE
        self.files_tree_scrollframe = ctk.CTkScrollableFrame(
            self.files_tree_frame,
            orientation="horizontal",
            width=240,
            height=150,
            fg_color="transparent"
        )
        self.files_tree_scrollframe.pack(fill="both", expand=True)

        # Inner container for Treeview
        self.files_tree_container = ctk.CTkFrame(self.files_tree_scrollframe, fg_color="transparent")
        self.files_tree_container.pack(side="left", fill="both", expand=True)
        self.files_tree_scrollbar = None

        # Placeholder for Treeview widget
        self.files_tree_widget = None

        self.files_tree_widget = CheckboxTreeview(self.files_tree_container, show="tree", height=10)
        self.files_tree_widget.pack(side="left", fill="both", expand=True)
        self.files_tree_widget.column("#0", width=800, stretch=False)

        # Counter Label (created only once)
        self.pdf_counter_label = ctk.CTkLabel(self.files_tree_frame, text="Selected PDFs: 0", font=(BUTTON_FONT, 9))
        self.pdf_counter_label.pack(pady=(5, 0), anchor="center")

        # ======================= 🔲 RECTANGLES TAB =======================


        # 🔲 Frame for Mode Buttons (Area / Revision)
        mode_frame = ctk.CTkFrame(tab_rectangles, width=240, height=24, fg_color="transparent")
        mode_frame.pack_propagate(False)
        mode_frame.pack(pady=(5, 5))

        self.mode_area_btn = ctk.CTkButton(mode_frame, text="🟥 Area", command=self.set_mode_area,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.mode_area_btn.pack(side="left", padx=5)

        self.mode_revision_btn = ctk.CTkButton(mode_frame, text="🟩 Revision Table", command=self.set_mode_revision,
                                               font=(BUTTON_FONT, 9), width=115, height=24)
        self.mode_revision_btn.pack(side="left", padx=5)

        #revision pattern
        pattern_options = [f"{k} — {', '.join(v['examples'])}" for k, v in REVISION_PATTERNS.items()]
        self.revision_dropdown_map = {f"{k} — {', '.join(v['examples'])}": k for k, v in REVISION_PATTERNS.items()}
        self.revision_pattern_var = StringVar(value=pattern_options[-1])
        self.revision_pattern_menu = ctk.CTkOptionMenu(tab_rectangles,
                                                       font=("Verdana", 9),
                                                       values=pattern_options,
                                                       variable=self.revision_pattern_var,
                                                       width=240, height=24,
                                                       fg_color="red4", button_color="red4", text_color="white")
        self.revision_pattern_menu.pack(pady=(5, 5))

        # 🛈 How to Use Button (Revision Table)
        self.revision_help_button = ctk.CTkButton(
            tab_rectangles,
            text="How to Use?",
            command=self.show_revision_help,
            font=(BUTTON_FONT, 9),
            fg_color="gray20", hover_color="gray30",
            width=240,
            height=24
        )
        self.revision_help_button.pack(pady=(0, 5))


        # ⬇️ Frame for Import/Export Buttons
        action_frame = ctk.CTkFrame(tab_rectangles, width=240, height=24, fg_color="transparent")
        action_frame.pack_propagate(False)
        action_frame.pack(pady=(25, 5))

        self.import_button = ctk.CTkButton(action_frame, text="⬇️ Import Areas", command=self.import_rectangles,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.import_button.pack(side="left", padx=5)

        self.export_button = ctk.CTkButton(action_frame, text="⬆️ Export Areas", command=self.export_rectangles,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.export_button.pack(side="left", padx=5)

        # 🗑 Clear Button (Full width)
        # 🧹 Row of Clear Buttons
        clear_frame = ctk.CTkFrame(tab_rectangles, fg_color="transparent", width=240, height=24)
        clear_frame.pack(pady=(5, 15))

        # 🟥 Clear Extraction Areas
        self.clear_extraction_button = ctk.CTkButton(clear_frame, text="Clear Area",
                                                     command=self.clear_extraction_areas,
                                                     font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_extraction_button.pack(side="left", padx=5)

        # 🟩 Clear Revision Table
        self.clear_revision_button = ctk.CTkButton(clear_frame, text="Clear Rev",
                                                   command=self.clear_revision_area,
                                                   font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_revision_button.pack(side="left", padx=5)

        # 🗑 Clear Button (Full width)
        self.clear_areas_button = ctk.CTkButton(clear_frame, text="Clear All", command=self.clear_all_areas,
                                                font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_areas_button.pack(side="left", padx=5)

        # Treeview inside a sub-frame
        self.areas_frame = ctk.CTkFrame(tab_rectangles, width=240)
        self.areas_frame.pack_propagate(False)  # Optional to maintain control over child sizes

        self.areas_frame.pack(pady=5, fill="both", expand=True)
        self.areas_tree = EditableTreeview(self, self.areas_frame,
                                           columns=("Title", "x0", "y0", "x1", "y1"),
                                           show="headings", height=4)
        self.areas_tree.heading("Title", text="Title")
        self.areas_tree.column("Title", width=60, anchor="center")
        for col in ("x0", "y0", "x1", "y1"):
            self.areas_tree.heading(col, text=col)
            self.areas_tree.column(col, width=40, anchor="center")
        self.areas_tree.pack(side="left", fill="both", expand=True)

        scrollbar = ctk.CTkScrollbar(self.areas_frame, orientation="vertical", command=self.areas_tree.yview)
        self.areas_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.areas_tree.bind("<<TreeviewSelect>>", self.on_treeview_select)

        # ======================= 🚀 EXTRACT TAB =======================

        # 📦 Frame for OCR & DPI Settings
        extract_frame = ctk.CTkFrame(tab_extract, width=240, fg_color="transparent")
        extract_frame.pack(pady=(25, 5))

        # ─────────────── OCR Setting ───────────────
        ocr_row = ctk.CTkFrame(extract_frame, fg_color="transparent")
        ocr_row.pack(pady=(0, 5), fill="x")

        ocr_label = ctk.CTkLabel(ocr_row, text="OCR Mode:", font=(BUTTON_FONT, 9), width=80, anchor="w")
        ocr_label.pack(side="left", padx=(0, 5))

        self.ocr_menu_var = StringVar(value="Default")
        self.ocr_menu = ctk.CTkOptionMenu(ocr_row,
                                          values=[ "Default", "OCR-All", "Text1st+Image-beta"],
                                          command=self.ocr_menu_callback,
                                          font=("Verdana Bold", 9),
                                          variable=self.ocr_menu_var,
                                          width=140, height=24)
        self.ocr_menu.pack(side="left", padx=(0, 5))

        # Add help "?" button next to OCR dropdown
        self.ocr_help_button = ctk.CTkButton(
            ocr_row,
            text="?",
            width=24,
            height=24,
            fg_color="gray20", hover_color="gray30",
            font=(BUTTON_FONT, 10),
            command=self.show_ocr_help
        )
        self.ocr_help_button.pack(side="left", padx=(5, 0))

        # ─────────────── DPI Setting ───────────────
        dpi_row = ctk.CTkFrame(extract_frame, fg_color="transparent")
        dpi_row.pack(pady=(0, 5), fill="x")

        dpi_label = ctk.CTkLabel(dpi_row, text="DPI:", font=(BUTTON_FONT, 9), width=80, anchor="w")
        dpi_label.pack(side="left", padx=(0, 5))

        self.dpi_var = ctk.IntVar(value=150)
        self.dpi_menu = ctk.CTkOptionMenu(dpi_row,
                                          values=["50", "75", "150", "300", "450", "600"],
                                          command=self.dpi_callback,
                                          font=("Verdana Bold", 9),
                                          variable=self.dpi_var,
                                          width=140, height=24)
        self.dpi_menu.pack(side="left")

        self.output_path_entry = ctk.CTkEntry(tab_extract, width=240, height=24, font=(BUTTON_FONT, 9),
                                              placeholder_text="Select Excel Output Path", border_width=1,
                                              corner_radius=3)
        self.output_path_entry.pack(pady=(15, 2))
        self.output_path_button = ctk.CTkButton(tab_extract, text="📂 Browse Output Path", command=self.browse_output_path,
                                                font=(BUTTON_FONT, 9), width=240, height=24)
        self.output_path_button.pack(pady=2)

        self.extract_button = ctk.CTkButton(tab_extract, text="🚀 Extract Now", font=("Arial Black", 13),
                                            corner_radius=10, width=240, height=30, command=self.start_extraction)
        self.extract_button.pack(pady=35)

        self.extract_description_box = ctk.CTkTextbox(tab_extract, width=240, height=200, wrap="word",
                                                      font=(BUTTON_FONT, 9))
        self.extract_description_box.insert("end",
                                            "EXTRACTION SUMMARY\n\n"
                                            "Output Excel File (.xlsx):\n"
                                            "  - One row per PDF page\n"
                                            "  - Columns from selected areas\n"
                                            "  - Hyperlinks to original PDF files\n\n"
                                            "IF REVISION TABLE IS USED:\n"
                                            "  ️‼️‼️EXTRACTION TIME WILL INCREASE BY x3 or x4‼️‼️\n\n"
                                            "  - Adds revision rows (Rev, Desc, Date)\n"
                                            "  - Saves NDJSON with structured revision info\n\n"
                                            "IF TEXT IS OCR-ed:\n"
                                            "  - Texts will be colored red\n"
                                            )
        self.extract_description_box.configure(state="disabled")  # Make it read-only
        self.extract_description_box.pack(pady=(0, 10))

        # ======================= 🧰 TOOLS TAB =======================
        # Create a frame for all tools
        tool_frame = ctk.CTkFrame(tab_tools)
        tool_frame.pack(pady=10, fill="both", expand=True)

        for label, tool in tool_definitions.items():
            # Create the main tool button
            if label == "📐 PDF & DWG Checker":
                # Pass self.root explicitly when launching PDF/DWG Checker popup
                btn = ctk.CTkButton(
                    tool_frame,
                    text=label,
                    width=240,
                    height=28,
                    font=(BUTTON_FONT, 10),
                    command=lambda root=self.root, func=tool["action"]: func(root)
                )
            else:
                btn = ctk.CTkButton(
                    tool_frame,
                    text=label,
                    width=240,
                    height=28,
                    font=(BUTTON_FONT, 10),
                    command=tool["action"]
                )
            btn.pack(pady=(10, 2))

            # Create the "How to use?" help button next to it, capturing tool text correctly
            help_btn = ctk.CTkButton(
                tool_frame,
                text="How to use?",
                width=240,
                height=20,
                font=(BUTTON_FONT, 9),
                fg_color="gray20",
                hover_color="gray30",
                command=lambda t=tool: self.show_tool_instructions(t["instructions"])
            )
            help_btn.pack(pady=(0, 5))

        # Load PNG from style folder
        logo_image = Image.open(resource_path("style/xtractor-logo.png"))
        logo_image = logo_image.resize((180, 180), Image.Resampling.LANCZOS)
        logo_photo = CTkImage(light_image=logo_image, dark_image=logo_image, size=(180, 180))

        # Create a label to hold the image
        self.logo_label = ctk.CTkLabel(tab_tools, image=logo_photo, text="", width=180, height=45)
        self.logo_label.image = logo_photo  # Prevent garbage collection
        self.logo_label.pack(pady=(20, 10), anchor="center")

        self.version_label = ctk.CTkLabel(
            tab_tools,
            text=VERSION_TEXT,
            fg_color="transparent",
            text_color="gray59",
            font=(BUTTON_FONT, 9)
        )
        self.version_label.pack(pady=(0, 15), anchor="center")  # ⬅️ anchor set to center
        self.version_label.bind("<Button-1>", self.display_version_info)

        self.root.after_idle(self.update_floating_controls)

    def show_tool_instructions(self, text):
        window = ctk.CTkToplevel(self.root)
        window.title("Tool Instructions")
        window.geometry("500x400")
        text_box = ctk.CTkTextbox(window, wrap="word")
        text_box.insert("end", text)
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def show_ocr_help(self):
        window = ctk.CTkToplevel(self.root)
        window.title("OCR Mode Explanation")
        window.geometry("480x300")
        text_box = ctk.CTkTextbox(window, wrap="word", font=(BUTTON_FONT, 11))
        text_box.insert("end",
                        "🧠 OCR Modes:\n\n"
                        "• Default:\n"
                        "   Extracts text normally. If no text is extracted, OCR is used.\n\n"
                        "• OCR-All:\n"
                        "   Ignores normal text. OCR is always applied.\n\n"
                        "• Text1st+Image-beta:\n"
                        "   Default mode and\n"
                        "   also saves and embeds area image into Excel.\n"
                        )
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def on_treeview_select(self, event):
        # Clear previous selection (restore original red)
        if self.pdf_viewer.selected_rectangle_id is not None:
            self.pdf_viewer.canvas.itemconfig(self.pdf_viewer.selected_rectangle_id, outline="red")
            self.pdf_viewer.selected_rectangle_id = None

        # Get the selected Treeview item
        selected = self.areas_tree.selection()
        if not selected:
            return

        item_id = selected[0]
        rect_index = self.treeview_item_ids.get(item_id)

        for rect_id in self.pdf_viewer.rectangle_list:
            self.pdf_viewer.canvas.itemconfig(rect_id, outline="red", width=2)

        if rect_index is not None and rect_index < len(self.pdf_viewer.rectangle_list):
            rect_id = self.pdf_viewer.rectangle_list[rect_index]
            # Highlight rectangle in yellow
            self.pdf_viewer.canvas.itemconfig(rect_id, outline="yellow", width=3)
            self.pdf_viewer.selected_rectangle_id = rect_id

    def clear_extraction_areas(self):
        """Clears only extraction areas, leaving the revision area untouched."""
        self.pdf_viewer.areas.clear()
        for rect_id in self.pdf_viewer.rectangle_list:
            self.pdf_viewer.canvas.delete(rect_id)
        self.pdf_viewer.rectangle_list.clear()
        self.update_areas_treeview()
        self.pdf_viewer.update_rectangles()
        print("Cleared only extraction areas.")

    def clear_revision_area(self):
        """Clears only the revision area, leaving extraction areas untouched."""
        if self.pdf_viewer.revision_rectangle_id:
            self.pdf_viewer.canvas.delete(self.pdf_viewer.revision_rectangle_id)
        self.pdf_viewer.revision_rectangle_id = None
        self.pdf_viewer.revision_area = None
        self.pdf_viewer.update_rectangles()
        print("Cleared only revision table area.")

    def show_revision_help(self):
        window = ctk.CTkToplevel(self.root)
        window.title("Revision Table Area and Pattern Explanation")
        window.geometry("480x300")
        text_box = ctk.CTkTextbox(window, wrap="word", font=(BUTTON_FONT, 11))
        text_box.insert("end",
            "🟩 Revision Table Help\n\n"
            "  ️Note: EXTRACTION TIME WILL INCREASE BY x3 or x4 if you use this feature\n\n"
            "• Use 'Revision Table' mode to select a SINGLE AREA containing revision history -revision, date, description.\n\n"
            "• Do not include in the selection area the -header-/-footer- (the one with Rev Date Description).\n\n"
            "• After selecting the area, use the dropdown to choose the expected revision format (e.g. A, B, C or A1, B2).\n\n"
            "• Make sure the selected table area includes 3 columns: Revision, Description, and Date.\n\n")
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def update_floating_controls(self):
        """
        Pin Recent↩, Close X, and the Zoom slider to the same x-coordinate
        as the PDF canvas.  Because the canvas is already placed with
        resize_canvas(), this avoids guessing sidebar/border widths.
        """
        self.root.update_idletasks()
        canvas_x = self.pdf_viewer.canvas.winfo_x()
        win_h = self.root.winfo_height()
        gap = self._px(4)
        y_top = self._px(23)
        y_zoom = win_h - self._px(57)

        self.recent_pdf_button.place(x=canvas_x + gap, y=y_top)
        self.close_pdf_button.place(x=canvas_x + gap + self._px(30), y=y_top)
        self.zoom_frame.place(x=canvas_x + gap, y=y_zoom)

    def place_zoom_and_version_controls(self):
        sidebar_width = self.tab_view.winfo_width() + self._px(20)
        window_height = self.root.winfo_height()
        self.zoom_frame.place(x=sidebar_width + 0, y=window_height - self._px(57))
        self.recent_pdf_button.place(x=sidebar_width + 0, y=self._px(23))
        self.close_pdf_button.place(x=sidebar_width + self._px(30), y=self._px(23))

    def setup_bindings(self):
        self.pdf_folder_entry.bind("<KeyRelease>", self.update_pdf_folder)
        self.output_path_entry.bind("<KeyRelease>", self.update_output_path)
        self.root.bind("<Configure>", self.on_window_resize)

    def setup_tooltips(self):
        create_tooltip(self.ocr_menu, "OCR options - select an OCR mode for text extraction")
        create_tooltip(self.dpi_menu, "DPI resolution")
        create_tooltip(self.pdf_folder_entry, "Select the main folder containing PDF files")
        create_tooltip(self.output_path_entry, "Select folder for the Excel output")
        create_tooltip(self.extract_button, "Start the extraction process")
        create_tooltip(self.import_button, "Import a saved template of selected areas")
        create_tooltip(self.export_button, "Export the selected areas as a template")
        create_tooltip(self.clear_areas_button, "Clear all selected areas")
        create_tooltip(self.revision_pattern_menu, "Choose the revision format pattern")
        create_tooltip(self.recent_pdf_button, "Open recent PDF")
        create_tooltip(self.close_pdf_button, "Close the opened PDF")

    def build_folder_tree(self):
        # destroy old tree
        if self.files_tree_widget:
            self.files_tree_widget.destroy()
            self.files_tree_widget = None
        # destroy old scrollbar
        if self.files_tree_scrollbar:
            try:
                self.files_tree_scrollbar.destroy()
            except Exception:
                pass
            self.files_tree_scrollbar = None

        # guard
        if not self.pdf_folder or not os.path.isdir(self.pdf_folder):
            return

        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 7))
        style.configure("Treeview.Heading", font=("Arial", 8, "bold"))

        # new tree + scrollbar
        self.files_tree_widget = CheckboxTreeview(self.files_tree_container, show="tree", height=10)
        self.files_tree_widget.pack(side="left", fill="both", expand=True)
        self.files_tree_widget.column("#0", width=800, stretch=False)

        self.files_tree_scrollbar = ttk.Scrollbar(
            self.files_tree_container, orient="vertical", command=self.files_tree_widget.yview
        )
        self.files_tree_widget.configure(yscrollcommand=self.files_tree_scrollbar.set)
        self.files_tree_scrollbar.pack(side="right", fill="y")

        # root node
        root_text = os.path.basename(self.pdf_folder.rstrip(os.sep)) or self.pdf_folder
        root_node = self.files_tree_widget.insert("", "end", text=root_text, tags=("checked",))

        # helper to insert children (folders with PDFs and .pdf files)
        def insert_children(parent, folder_path):
            try:
                entries = sorted(os.listdir(folder_path))
            except Exception as e:
                print(f"Error accessing {folder_path}: {e}")
                return
            for entry in entries:
                full_path = os.path.join(folder_path, entry)
                if os.path.isdir(full_path):
                    if self.has_pdf(full_path):  # only show folders that contain PDFs somewhere under them
                        node = self.files_tree_widget.insert(parent, "end", text=entry, tags=("checked",))
                        insert_children(node, full_path)
                elif entry.lower().endswith(".pdf"):
                    # if we dropped a subset, only those should be checked; otherwise default to checked
                    is_dropped_subset = bool(getattr(self, "dropped_pdf_set", set()))
                    tags = ("checked",) if (not is_dropped_subset or full_path in self.dropped_pdf_set) else (
                    "unchecked",)
                    self.files_tree_widget.insert(parent, "end", text=entry, values=[full_path], tags=tags)

        insert_children(root_node, self.pdf_folder)

        # update counter when user clicks/changes checks
        self.files_tree_widget.bind("<<TreeviewSelect>>", lambda e: self.update_pdf_counter())
        self.files_tree_widget.bind("<ButtonRelease-1>", lambda e: self.root.after(100, self.update_pdf_counter))

        self.update_pdf_counter()

    def recursive_set_check_state(self, item_id):
        children = self.files_tree_widget.get_children(item_id)
        if not children:
            # Leaf node (file)
            item = self.files_tree_widget.item(item_id)
            values = item.get("values", [])
            if values:
                full_path = os.path.abspath(values[0])
                tag = "checked" if full_path in self.dropped_pdf_set else "unchecked"
                self.files_tree_widget.item(item_id, tags=(tag,))
                return tag == "checked"
            else:
                self.files_tree_widget.item(item_id, tags=("unchecked",))
                return False
        else:
            # Folder node
            any_child_checked = False
            for child in children:
                if self.recursive_set_check_state(child):
                    any_child_checked = True

            tag = "checked" if any_child_checked else "unchecked"
            self.files_tree_widget.item(item_id, tags=(tag,))
            return any_child_checked

    def drop_pdf_folder(self, event):
        raw_data = event.data.strip()
        raw_items = re.findall(r'{(.*?)}', raw_data) or [raw_data.strip()]
        cleaned_paths = [os.path.abspath(p.strip('"')) for p in raw_items]

        dropped_pdfs = [p for p in cleaned_paths if os.path.isfile(p) and p.lower().endswith(".pdf")]
        dropped_folders = [p for p in cleaned_paths if os.path.isdir(p)]

        def collect_pdfs_from_folder(folder_path):
            collected = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith(".pdf"):
                        collected.append(os.path.join(root, file))
            return collected

        all_pdfs = dropped_pdfs.copy()
        for folder in dropped_folders:
            all_pdfs.extend(collect_pdfs_from_folder(folder))

        if not all_pdfs:
            messagebox.showerror("Invalid Drop", "No PDF files found in the dropped folders or files.")
            return

        # choose a common root if possible
        try:
            common_root = os.path.commonpath(all_pdfs)
        except ValueError:
            # different drives on Windows -> just use the first file’s folder
            common_root = os.path.dirname(all_pdfs[0])

        if not os.path.isdir(common_root):
            common_root = os.path.dirname(all_pdfs[0])

        self.pdf_folder = common_root
        self.pdf_folder_entry.delete(0, ctk.END)
        self.pdf_folder_entry.insert(0, self.pdf_folder)

        # mark only the dropped files as "checked"
        self.dropped_pdf_set = set(all_pdfs)

        self.build_folder_tree()
        # (optional) if you want to force recalculating tag state top-down:
        # for iid in self.files_tree_widget.get_children(""):
        #     self.recursive_set_check_state(iid)
        self.update_pdf_counter()

        if not self.pdf_viewer.pdf_document:
            self.pdf_viewer.display_pdf(all_pdfs[0])
            self.recent_pdf_path = all_pdfs[0]
            print(f"Loaded first PDF from dropped items: {all_pdfs[0]}")

    def drop_sample_pdf(self, event):
        path = event.data.strip().replace("{", "").replace("}", "")
        if os.path.isfile(path) and path.lower().endswith(".pdf"):
            self.pdf_viewer.display_pdf(path)
            self.recent_pdf_path = path
            print(f"Dropped Sample PDF: {path}")
        else:
            messagebox.showerror("Invalid Drop", "Please drop a valid PDF file.")

    def set_mode_area(self):
        self.pdf_viewer.selection_mode = "area"
        print("🟥 Switched to Area Mode")
        self.highlight_mode_button(self.mode_area_btn)
        self.reset_mode_button(self.mode_revision_btn)

    def set_mode_revision(self):
        self.pdf_viewer.selection_mode = "revision"
        print("🟩 Switched to Revision History Mode")
        self.highlight_mode_button(self.mode_revision_btn)
        self.reset_mode_button(self.mode_area_btn)

    def highlight_mode_button(self, button):
        """Visually highlight the active mode button."""
        button.configure(fg_color="#3949AB", text_color="white", border_color="white", border_width=2)

    def reset_mode_button(self, button):
        """Reset the visual style of inactive buttons."""
        button.configure(fg_color="gray25", text_color="white", border_width=0)

    def update_pdf_counter(self):
        if not self.files_tree_widget:
            return

        def is_pdf(iid):
            item = self.files_tree_widget.item(iid)
            text = item.get("text", "")
            return text.lower().endswith(".pdf")

        checked = [iid for iid in self.files_tree_widget.get_checked() if is_pdf(iid)]
        count = len(checked)
        self.pdf_counter_label.configure(text=f"Selected PDFs: {count}")

    def has_pdf(self, folder_path):
        """Recursively checks if the folder or its subfolders contain any .pdf files."""
        try:
            for entry in os.listdir(folder_path):
                full_path = os.path.join(folder_path, entry)
                if os.path.isdir(full_path):
                    if self.has_pdf(full_path):
                        return True
                elif entry.lower().endswith(".pdf"):
                    return True
        except Exception as e:
            print(f"Error scanning {folder_path}: {e}")
        return False

    #not needed
    def ocr_menu_callback(self, choice):
        print("OCR menu dropdown clicked:", choice)

        def enable_ocr_menu(enabled):
            color = "red4" if enabled else "gray29"
            self.ocr_menu.configure(fg_color=color, button_color=color)
            self.dpi_menu.configure(state="normal" if enabled else "disabled", fg_color=color, button_color=color)

        if choice in ("Default", "OCR-All", "Text1st+Image-beta"):
            found_tesseract_path = find_tessdata()
            if found_tesseract_path:
                self.ocr_settings['tessdata_folder'] = found_tesseract_path
                enable_ocr_menu(True)
                if choice == "Default":
                    print("OCR will start if no text is extracted.")
                elif choice == "OCR-All":
                    print("OCR will be enabled for every area.")
                elif choice == "Text1st+Image-beta":
                    print("OCR will start if no text is extracted and images will also be extracted.")
            else:
                enable_ocr_menu(False)
                print("Tessdata folder not found. OCR disabled.")

        self.ocr_settings['enable_ocr'] = choice
        print("OCR mode:", self.ocr_settings['enable_ocr'])

    def dpi_callback(self, dpi_value):
        self.ocr_settings['dpi_value'] = int(dpi_value)
        print(f"DPI set to: {dpi_value}")

    def browse_pdf_folder(self):
        self.pdf_folder = filedialog.askdirectory()
        if not self.pdf_folder:
            return
        self.pdf_folder_entry.delete(0, ctk.END)
        self.pdf_folder_entry.insert(0, self.pdf_folder)
        # we are browsing a whole folder -> no subset filter
        self.dropped_pdf_set = set()
        self.build_folder_tree()

    def browse_output_path(self):
        """Opens a dialog to specify the output Excel file path."""
        self.output_excel_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if self.output_excel_path:  # Only update the entry if a file was selected
            self.output_path_entry.delete(0, ctk.END)
            self.output_path_entry.insert(0, self.output_excel_path)

    def update_zoom_slider(self, zoom_level):
        """Updates the zoom slider to reflect the current zoom level in PDFViewer."""
        self.zoom_var.set(zoom_level)

    def update_pdf_folder(self, event):
        self.pdf_folder = self.pdf_folder_entry.get()

    def update_output_path(self, event):
        self.output_excel_path = self.output_path_entry.get()

    def update_zoom(self, value):
        """Adjusts the zoom level of the PDFViewer based on slider input."""
        zoom_level = float(value)
        self.pdf_viewer.set_zoom(zoom_level)  # Update zoom in PDFViewer

    def start_extraction(self):
        # --- guards (unchanged) ---
        if not self.pdf_viewer.areas:
            messagebox.showerror("Extraction Error", "No areas defined. Please select areas before extracting.")
            return
        self.pdf_folder = self.pdf_folder_entry.get()
        if not self.pdf_folder or not os.path.isdir(self.pdf_folder):
            messagebox.showerror("Invalid Folder", "Select a valid folder.")
            return
        self.output_excel_path = self.output_path_entry.get()
        output_dir = os.path.dirname(self.output_excel_path)
        if not self.output_excel_path or not os.path.isdir(output_dir):
            messagebox.showerror("Invalid Output Path", "Select a valid folder.")
            return

        # collect checked PDFs
        checked = self.files_tree_widget.get_checked()
        selected_paths = []
        for iid in checked:
            item = self.files_tree_widget.item(iid)
            if item and "values" in item and item["values"]:
                path = item["values"][0]
                if path.lower().endswith(".pdf"):
                    selected_paths.append(Path(path))
        if not selected_paths:
            messagebox.showerror("No Files Selected", "Please check at least one PDF to extract.")
            return

        # close viewer to release any file locks
        self.pdf_viewer.close_pdf()

        # progress UI
        self.start_time = time.time()
        self.progress_window = ctk.CTkToplevel(self.root)
        self.progress_window.title("Progress")
        self.progress_window.geometry("320x150")
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        self.progress_window.attributes('-topmost', True)

        self.progress_label = ctk.CTkLabel(self.progress_window, text="Processing PDFs...")
        self.progress_label.pack(pady=(10, 2))
        self.total_files_label = ctk.CTkLabel(self.progress_window, text="Total: 0")
        self.total_files_label.pack(pady=2)

        self.progress_var = ctk.DoubleVar(value=0)
        self.progress_bar = ctk.CTkProgressBar(self.progress_window, variable=self.progress_var,
                                               orientation="horizontal", width=260)
        self.progress_bar.pack(pady=8)
        cancel_btn = ctk.CTkButton(self.progress_window, text="Cancel", width=80, command=self.on_cancel_extraction)
        cancel_btn.pack(pady=(2, 8))
        self.progress_window.protocol("WM_DELETE_WINDOW", self.on_cancel_extraction)

        # build domain request (accept both dict & dataclass areas)
        selected_pattern_display = self.revision_pattern_var.get()
        selected_pattern_key = self.revision_dropdown_map[selected_pattern_display]
        selected_revision_regex = REVISION_PATTERNS[selected_pattern_key]["pattern"]

        req = ExtractionRequest(
            pdf_paths=selected_paths,
            output_excel=Path(self.output_excel_path),
            areas=[_to_spec(a) for a in self.pdf_viewer.areas],
            revision_area=_rev_area_to_spec(self.pdf_viewer.revision_area),
            revision_regex=selected_revision_regex,
            ocr=OcrSettings(
                mode=self.ocr_settings['enable_ocr'],
                dpi=int(self.ocr_settings['dpi_value']),
                tessdata_dir=Path(self.ocr_settings['tessdata_folder']) if self.ocr_settings.get(
                    'tessdata_folder') else None,
                scale=self.ocr_settings.get("scale")
            ),
            pdf_root=Path(self.pdf_folder)
        )

        # start job via controller
        self._job = self.extractor.start(req)

        def _tick():
            polled = self.extractor.poll(self._job)
            if polled is not None:
                processed, total = polled
                if total > 0:
                    self.total_files_label.configure(text=f"Processed pages: {processed}/{total}")
                    self.progress_var.set(processed / max(total, 1))
                self.root.after(100, _tick)
            else:
                # finished (or cancelled / error)
                try:
                    self.progress_var.set(1)
                except Exception:
                    pass
                try:
                    self.progress_window.destroy()
                except Exception:
                    pass

                out = self.extractor.finish(self._job)  # Path | None
                self._job = None

                end_time = time.time()
                elapsed = end_time - self.start_time
                formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed))

                if not out or not out.exists():
                    # treat as cancelled or failed (controller does not return a separate status)
                    messagebox.showinfo("Stopped", f"Extraction was cancelled or failed after {formatted}.")
                    return

                if messagebox.askyesno("Extraction Complete", f"Completed in {formatted}.\nOpen the Excel file?"):
                    try:
                        os.startfile(out)
                    except Exception as e:
                        messagebox.showerror("Error", f"Could not open the Excel file: {e}")

        self.root.after(100, _tick)

    def on_cancel_extraction(self):
        if hasattr(self, "_job") and self._job:
            if messagebox.askyesno("Cancel extraction", "Stop the extraction now?"):
                self.extractor.cancel(self._job)
                try:
                    self.progress_label.configure(text="Cancelling…")
                except Exception:
                    pass

    def update_progress(self, progress_counter, total_files, extraction_process):
        if total_files.value > 0:
            processed = progress_counter.value
            self.total_files_label.configure(text=f"Processed: {processed}/{total_files.value}")
            self.progress_var.set(processed / max(total_files.value, 1))

        if extraction_process.is_alive():
            self.root.after(100, self.update_progress, progress_counter, total_files, extraction_process)
            return

        # process finished
        status = (self.shared or {}).get("status", "done")
        err = (self.shared or {}).get("error", "")
        try:
            self.progress_var.set(1)
        except Exception:
            pass
        if getattr(self, "progress_window", None):
            try:
                self.progress_window.destroy()
            except Exception:
                pass
            self.progress_window = None

        try:
            extraction_process.join(timeout=2)
            extraction_process.close()
        except Exception:
            pass
        self.extraction_process = None

        # shutdown manager safely
        try:
            if getattr(self, "current_manager", None):
                self.current_manager.shutdown()
        except Exception:
            pass
        self.current_manager = None

        final_output_file = (self.final_output_path.value or self.output_excel_path).strip()

        if status == "cancelled":
            logger.info("User cancelled; no dialog.")
            return
        if status == "error":
            logger.error("Worker error: %s", err)
            messagebox.showerror("Extraction failed", err or "An unknown error occurred. See logs for details.")
            return

        # success
        elapsed = time.time() - self.start_time
        formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed))
        if messagebox.askyesno("Extraction Complete", f"Completed in {formatted}.\nOpen the Excel file?"):
            if final_output_file and os.path.exists(final_output_file):
                try:
                    os.startfile(final_output_file)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not open the Excel file: {e}")

    def optionmenu_callback(self, choice):
        """Execute the corresponding function based on the selected option."""
        action = OPTION_ACTIONS.get(choice)
        if action:
            action()  # Call the function
        else:
            messagebox.showerror("Error", f"No action found for {choice}")

    def on_window_resize(self, event=None):
        new_width = self.root.winfo_width()
        new_height = self.root.winfo_height()

        if hasattr(self, "prev_width") and hasattr(self, "prev_height"):
            if new_width == self.prev_width and new_height == self.prev_height:
                return

        self.prev_width = new_width
        self.prev_height = new_height

        try:
            # Estimate the width of the left tabview panel

            sidebar_width = self.tab_view.winfo_width() + 20

            self.pdf_viewer.resize_canvas(self.root.winfo_width(), self.root.winfo_height(),x_offset=CANVAS_LEFT_MARGIN)

            self.pdf_viewer.update_rectangles()

            # Zoom and version controls
            self.zoom_frame.place_configure(x=sidebar_width + 0, y=new_height - 57)

            self.update_floating_controls()

        except Exception as e:
            print(f"Error resizing widgets: {e}")

    def display_version_info(self, event):
        version_text = """
        Created by: Rei Raphael Reveral
        
        Links:
        https://github.com/r-Yayap/MultiplePDF-Areas2Excel
        https://www.linkedin.com/in/rei-raphael-reveral
        """
        window = ctk.CTkToplevel(self.root)
        window.title("Version Info")
        text_widget = ctk.CTkTextbox(window, wrap="word", width=400, height=247)
        text_widget.insert("end", version_text)
        text_widget.pack(padx=10, pady=10, side="left")
        window.grab_set()



