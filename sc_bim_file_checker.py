#sc_bim_file_checker.py
import os
import pandas as pd
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# List of “standard” extensions
EXTS = ["rvt", "ifc", "nwc", "dwfx", "nwd", "xml", "dwg"]


def choose_directory(title="Select Directory"):
    root = Tk()
    root.withdraw()
    path = filedialog.askdirectory(title=title)
    root.destroy()
    return path


def choose_file_save_location(title="Save Excel File As"):
    root = Tk()
    root.withdraw()
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title=title
    )
    root.destroy()
    return path


def list_files(start_path):
    # one dict per EXT, plus “OTHER”
    file_dict = {ext.upper(): {} for ext in EXTS}
    file_dict["OTHER"] = {}

    for root_dir, dirs, files in os.walk(start_path):
        for filename in files:
            base, ext = os.path.splitext(filename)
            ext = ext.lstrip('.').lower()
            full_path = os.path.join(root_dir, filename)
            info = {
                'path': full_path,
                'name': base,
                'ext': ext,
                'size': os.path.getsize(full_path),
                'modified': datetime.fromtimestamp(os.path.getmtime(full_path))
            }
            if ext in EXTS:
                file_dict[ext.upper()][full_path] = info
            else:
                file_dict["OTHER"][full_path] = info

    return file_dict


def prepare_data_for_export(file_dict, root_folder):
    def rel_folder(path):
        return os.path.relpath(os.path.dirname(path), root_folder)

    # gather all unique base names
    all_names = set(
        info['name']
        for bucket in file_dict.values()
        for info in bucket.values()
    )

    rows = []
    for name in sorted(all_names):
        row = {'FileName': name}

        # standard extensions
        for ext in EXTS:
            key = ext.upper()
            matches = [info for info in file_dict[key].values() if info['name'] == name]
            if matches:
                row[key] = "✓"
                row[f"{key}_Path"] = "; ".join(rel_folder(info['path']) for info in matches)
            else:
                row[key] = ""
                row[f"{key}_Path"] = ""

        # OTHER
        other_matches = [info for info in file_dict["OTHER"].values() if info['name'] == name]
        if other_matches:
            exts = sorted({info['ext'] for info in other_matches})
            row["Other"] = "; ".join(exts)
            row["Other_Path"] = "; ".join(rel_folder(info['path']) for info in other_matches)
        else:
            row["Other"] = ""
            row["Other_Path"] = ""

        rows.append(row)

    df = pd.DataFrame(rows)
    # enforce column order
    cols = (
        ["FileName"]
        + [e.upper() for e in EXTS]
        + ["Other"]
        + [f"{e.upper()}_Path" for e in EXTS]
        + ["Other_Path"]
    )
    return df[cols]


def save_to_excel(df, excel_file, file_dict):
    df.to_excel(excel_file, index=False)
    wb = load_workbook(excel_file)
    ws = wb.active

    present_fill = PatternFill(fill_type="solid", start_color="ABF5DF", end_color="ABF5DF")
    empty_fill   = PatternFill(fill_type="solid", start_color="000000", end_color="000000")
    dup_fill     = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")

    # color the standard‐ext columns
    for ext in EXTS:
        col = ext.upper()
        col_idx = df.columns.get_loc(col) + 1

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            matches = [info for info in file_dict[col].values() if info['name'] == name]
            cell = ws.cell(row=row, column=col_idx)

            if len(matches) > 1:
                cell.fill = dup_fill
            elif len(matches) == 1:
                cell.fill = present_fill
            else:
                cell.fill = empty_fill

    # color the “Other” column (no duplicate‐red for OTHER)
    other_idx = df.columns.get_loc("Other") + 1
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        has_other = any(info['name'] == name for info in file_dict["OTHER"].values())
        cell = ws.cell(row=row, column=other_idx)
        cell.fill = present_fill if has_other else empty_fill

    wb.save(excel_file)
    wb.close()
    print(f"Directory listing exported to {excel_file}")


def main():
    folder = choose_directory()
    if not folder:
        print("No directory selected.")
        return

    file_dict = list_files(folder)
    save_path = choose_file_save_location()
    if not save_path:
        print("No file location selected.")
        return

    df = prepare_data_for_export(file_dict, folder)
    save_to_excel(df, save_path, file_dict)

    if messagebox.askyesno("Open Excel File", "Do you want to open the Excel file now?"):
        os.startfile(save_path)


if __name__ == "__main__":
    main()
