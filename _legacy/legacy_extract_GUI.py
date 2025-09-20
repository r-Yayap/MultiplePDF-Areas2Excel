'''
Changelog 14
--fixed multiple pages extraction
--fixed autoscroll speed

Changelog 13
- Tooltips added
- Fixed updating rectangles
- Autoscroll
- Right Click to Delete Rectangle

Changelog 12
- added Image extraction (would not work for PDFs with multiple pages)
- added Last Modified Date and Size for Extractor and other features
- updated to pymupdf v 1.24.10
- import/export areas
- recent pdf/close pdf

Changelog 11
- Fixed "Illegal Text" error
- Text sorting hotfix

Changelog 10
- Fixed OCR

Changelog 09
- Fixed Zoom
- OCR options and DPI options

Changelog 08
- Optimized-ish? XD
- Progress Bar (at last!)

Changelog 07
- Resize Display along with th windows
- Added Option button for other features
- Added Bulk Rename and Directory List
- Scroll & Shift+Scroll on Canvas

Changelog 06
- UI Overhaul
- Zoom implemented
- Rectangles stays when zoomed
- Removed progress bar
- Can now edit coordinates

Changelog 05
-  time lapsed counter
-  Progress bar during extraction
-  List files on a table (DWG and PDf counter) [INTEGRATED!!!]
-  excel output: add time created on filename
-  open generated excel file (or directory)
-  Includes all pages

Changelog 04
- added DWG and PDF counter (numbers only)
- include subfolder

Changelog 03
- coordinates based on pdf's rotation
- can now read text regardless of pdf inherent rotation

Changelog 02
- scrollbar (not placed well, but working)
- area selection now working (coordinates are now correct)
- text extraction is now working

Changelog 01
- scrollbar (not placed well, but working)
- area selection in display (areas not fixed yet)
'''

import time
import os
import threading
import tkinter as tk
import json
from tkinter import messagebox, ttk, filedialog
import customtkinter as ctk
import pymupdf  as fitz# PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from datetime import datetime
from CTkToolTip import *

# Importing python files
from standalone import sc_bulk_rename as brn, sc_dir_list as dlist, sc_pdf_dwg_list as pdl

# Label to display version
version_txt = "Version 0.231219-14"

# Global variables
include_subfolders = False
pdf_folder = ''
output_excel_path = ''
areas = []
rectangle_list = []
ws = None
areas_tree = None
rectangle = None  # Define rectangle globally
canvas = None  # Initialize canvas globally
zoom_slider = None  # Initialize zoom_slider globally
prev_width = None  # for Windows Resize Function
prev_height = None
pix = None
page = None
enable_ocr = None
tessdata_folder = None
recent_pdf_path = None
pdf_document = None
selected_rectangle_id = None

# Initial Window and Display settings
initial_width = 965
initial_height = initial_width / ((1 + 5 ** 0.5) / 2)
canvas_width = initial_width - 30
canvas_height = initial_height - 135
initial_x_position = 100
initial_y_position = 100
current_zoom = 2.0

last_resize_time = 0 # Global variable to hold the time of the last resize event
resize_delay = 0.35  # Delay in seconds before the function is called after the last resize event


#Buttons/Widgets Styling
button_font = "Verdana"

# Define DPI options globally
dpi_options = {
    "75": 75,
    "150": 150,
    "300": 300,
    "450": 450,
    "600": 600,
    "750": 750,
    "900": 900,
    "1200": 1200
}

# Define the option menu choices and their corresponding actions
option_actions = {
    #"PDF/DWG List": pdl.pdf_dwg_counter,
    "Directory List": dlist.generate_file_list_and_excel,
    "Bulk Renamer": brn.bulk_rename_gui
}

class EditableTreeview(ttk.Treeview):
    def __init__(self, *args, **kwargs):
        ttk.Treeview.__init__(self, *args, **kwargs)
        self._entry = None
        self._col = None

        # Bind right-click to show context menu
        self.bind("<Button-3>", self.show_context_menu)
        self.bind("<Double-Button-1>", self.on_double_click)

        # Create context menu
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Remove Row", command=self.remove_row)

    def on_double_click(self, event):
        item = self.focus()
        col = self.identify_column(event.x)
        if item and col and col != "#0":
            self._col = col
            cell_values = self.item(item, "values")
            if cell_values:
                col_index = int(col.split("#")[-1]) - 1
                cell_value = cell_values[col_index]
                self.edit_cell(item, col, cell_value)

    def show_context_menu(self, event):
        item = self.identify_row(event.y)
        if item:
            self.context_menu.post(event.x_root, event.y_root)

    def remove_row(self):
        item = self.focus()
        if item:
            self.delete(item)

    def on_focus_out(self, event):
        if self._entry is not None:
            self.stop_editing()

    def edit_cell(self, item, col, _):
        def on_ok():
            new_value = entry_var.get()
            if new_value:
                current_values = list(self.item(item, "values"))
                current_values[col_index] = new_value
                self.item(item, values=tuple(current_values))
                self.update_areas_list()  # Update areas list when cell is edited
            top.destroy()

        bbox = self.bbox(item, col)
        x, y, _, _ = bbox
        col_index = int(col.replace("#", "")) - 1  # Subtract 1 for 0-based indexing

        top = ctk.CTkToplevel(self)
        top.title("Edit Cell")

        entry_var = ctk.StringVar()
        entry_var.set(self.item(item, "values")[col_index])

        entry = ctk.CTkEntry(top, justify="center", textvariable=entry_var,
                             width=100, height=20, font=(button_font, 9),
                             border_width=1,
                             corner_radius=3)

        entry.pack(pady=5)

        ok_button = ctk.CTkButton(top, text="OK", command=on_ok)
        ok_button.pack()

        top.geometry(f"+{x}+{y}")
        top.transient(self)  # Set the transient master to the treeview
        top.grab_set()  # Make the pop-up modal

        entry.focus_set()
        top.wait_window(top)  # Wait for the window to be closed

    def stop_editing(self, event=None):
        if self._entry is not None:
            new_value = self._entry.get()
            col = int(self._col.replace("#", ""))
            item = self.focus()

            if event and getattr(event, "keysym", "") == "Return" and item:
                current_values = self.item(item, "values")
                updated_values = [new_value if i == 0 else val for i, val in enumerate(current_values)]
                self.item(item, values=updated_values)
                self.update_areas_list()  # Update areas list when cell is edited

            self._entry.destroy()
            self._entry = None
            self._col = None

    def update_areas_list(self):
        global areas

        # Clear the existing areas list
        areas.clear()

        # Iterate through items in the EditableTreeview and update areas
        for item in self.get_children():
            values = self.item(item, "values")
            areas.append([float(value) for value in values])


# Tooltips using the helper function
def create_tooltip(widget, message,
                        delay=0.3,
                        font=("Verdana", 9),
                        border_width=1,
                        border_color="gray50",
                        corner_radius=6,
                        justify="left"
                   ):

    return CTkToolTip(widget,
                      delay=delay,
                      justify=justify,
                      font=font,
                      border_width=border_width,
                      border_color=border_color,
                      corner_radius=corner_radius,
                      message=message)


def browse_pdf_folder():
    global pdf_folder

    pdf_folder = filedialog.askdirectory()
    pdf_folder_entry.delete(0, ctk.END)
    pdf_folder_entry.insert(0, pdf_folder)

    print(f"Selected PDF Folder: {pdf_folder}")


def toggle_include_subfolders():
    global include_subfolders
    include_subfolders = include_subfolders_var.get()

    # Check if include_subfolders has a valid value (1 for True, 0 for False)
    if include_subfolders in {0, 1}:
        print(f"Include Subfolders: {bool(include_subfolders)}")
    else:
        print("Include Subfolders: Invalid value")


def H7354():
    H7354 = "mi amor"


def edit_areas(event, areas_tree_ref):
    global areas

    # Create a new window for editing areas
    edit_window = ctk.CTkToplevel(root)
    edit_window.title("Edit Areas")

    # Define column names and headings
    column_headings = ["X0", "Y0", "X1", "Y1"]

    # Create an EditableTreeview widget
    editable_treeview = EditableTreeview(edit_window, columns=column_headings, show="headings")

    # Set column headings
    for col, heading in enumerate(column_headings, start=1):
        editable_treeview.heading(f"#{col}", text=heading)

    # Insert existing areas into the EditableTreeview
    for idx, area in enumerate(areas, start=1):
        values = [area[i] for i in range(len(column_headings))]
        editable_treeview.insert("", idx, values=values)

    editable_treeview.pack(padx=10, pady=10)

    def save_changes():
        global areas

        # Clear the existing areas list
        areas = []

        # Iterate through items in the EditableTreeview and update areas
        for item in editable_treeview.get_children():
            values = editable_treeview.item(item, "values")
            areas.append([float(value) for value in values])

        # Close the edit window
        edit_window.destroy()

        # Update the display
        update_display()

        # Update the areas_tree_ref in the main window
        areas_tree_ref.delete(*areas_tree_ref.get_children())
        for idx, area in enumerate(areas, start=1):
            areas_tree_ref.insert("", idx, values=(area[0], area[1], area[2], area[3]))


    # Create a Save button
    save_button = ctk.CTkButton(edit_window, text="Save Changes", command=save_changes)
    save_button.pack(pady=10)


def browse_output_path():
    global output_excel_path
    output_excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_path_entry.delete(0, ctk.END)
    output_path_entry.insert(0, output_excel_path)
    print(f"Selected Output Excel Path: {output_excel_path}")


def open_sample_pdf():
    global areas_tree, recent_pdf_path

    # Ask user to choose a sample PDF file
    sample_pdf_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])

    if sample_pdf_path:
        print(f"Opening Sample PDF: {sample_pdf_path}")

        # Store the path of the recently opened PDF
        recent_pdf_path = sample_pdf_path

        # Display the sample PDF and get its dimensions
        display_sample_pdf(sample_pdf_path)

    # Button to open the most recent PDF
    recent_pdf_button = ctk.CTkButton(root, text="Recent..", command=open_recent_pdf, font=(button_font, 9),
                                      fg_color=("gray29", "gray39"),
                                      #fg_color="#B30B00", # hover_color="#860A00",
                                       width=25, height=10)
    recent_pdf_button.place(x=87, y=35)  # Adjust the layout method as needed

    # Button to open the most recent PDF
    close_pdf_button = ctk.CTkButton(root, text="X", command=close_pdf, font=(button_font, 9),
                                     fg_color="#B30B00", # hover_color="#860A00",
                                       width=25, height=10)
    close_pdf_button.place(x=143, y=35)  # Adjust the layout method as needed

    tooltip_recent_pdf_button = create_tooltip(recent_pdf_button, "Open most recent opened template")
    tooltip_close_pdf_button = create_tooltip(close_pdf_button, "Close the PDF/remove PDF in the Display")


def start_rectangle(event):
    global original_coordinates, current_rectangle
    # Get the starting coordinates of the rectangle
    x, y = canvas.canvasx(event.x), canvas.canvasy(event.y)
    original_coordinates = [x, y]
    current_rectangle = canvas.create_rectangle(x, y, x, y, outline="red", width=2)


last_scroll_time = time.time()

def auto_scroll_canvas(x, y):
    global last_scroll_time
    scroll_margin = 20
    scroll_speed = 1
    delay = 0.05  # Delay in seconds between scrolls

    current_time = time.time()
    if current_time - last_scroll_time >= delay:
        # Only scroll if the delay period has passed
        if x >= canvas.winfo_width() - scroll_margin:
            canvas.xview_scroll(scroll_speed, "units")
        elif x <= scroll_margin:
            canvas.xview_scroll(-scroll_speed, "units")

        if y >= canvas.winfo_height() - scroll_margin:
            canvas.yview_scroll(scroll_speed, "units")
        elif y <= scroll_margin:
            canvas.yview_scroll(-scroll_speed, "units")

        last_scroll_time = current_time



def delete_rectangle():
    global selected_rectangle_id, areas, rectangle_list, areas_tree

    if selected_rectangle_id and selected_rectangle_id in rectangle_list:
        # Get the index of the rectangle in rectangle_list
        rect_index = rectangle_list.index(selected_rectangle_id)

        # Delete the rectangle from canvas
        canvas.delete(selected_rectangle_id)

        # Remove the rectangle data from areas and rectangle_list
        del areas[rect_index]
        del rectangle_list[rect_index]

        # Update the Treeview to reflect the deletion
        areas_tree.delete(areas_tree.get_children()[rect_index])

        # Clear the selected rectangle ID
        selected_rectangle_id = None
        print(f"Deleted rectangle at index {rect_index}")
    else:
        print("No rectangle selected for deletion.")


def show_popup_menu(event):
    global selected_rectangle_id

    # Get coordinates where right-click occurred
    x, y = canvas.canvasx(event.x), canvas.canvasy(event.y)

    # Find the closest rectangle item at the clicked coordinates
    item = canvas.find_closest(x, y)
    if item and item[0] in rectangle_list:
        # Only show the menu if a rectangle is clicked
        selected_rectangle_id = item[0]
        popup_menu.post(event.x_root, event.y_root)  # Show the popup menu at mouse position
    else:
        selected_rectangle_id = None  # Clear selection if not clicking a rectangle


def draw_rectangle(event):
    global original_coordinates, current_rectangle
    # Check if there's a current rectangle
    if current_rectangle:
        # Update the rectangle coordinates as the mouse is dragged
        x, y = canvas.canvasx(event.x), canvas.canvasy(event.y)
        canvas.coords(current_rectangle, original_coordinates[0], original_coordinates[1], x, y)

        # Call the auto-scroll function with the mouse coordinates (no other replacements here)
        auto_scroll_canvas(event.x, event.y)


def end_rectangle(event):
    global current_rectangle, areas, areas_tree, pdf_height, current_zoom

    if current_rectangle:
        # Get the final coordinates of the rectangle
        x, y = canvas.canvasx(event.x), canvas.canvasy(event.y)
        canvas.coords(current_rectangle, original_coordinates[0], original_coordinates[1], x, y)

        # Add the adjusted rectangle coordinates to the area table
        bbox = canvas.bbox(current_rectangle)
        if bbox is not None:
            x0, y0, x1, y1 = bbox

            # Adjust coordinates based on zoom level
            adjusted_coordinates = [
                x0 / current_zoom,
                y0 / current_zoom,
                x1 / current_zoom,
                y1 / current_zoom
            ]

            # append to areas
            areas.append(adjusted_coordinates)

            print("Updated Areas List:", areas)  # Print the areas list for debugging

            # Store the rectangle ID in the rectangle_list
            rectangle_list.append(current_rectangle)
            print("Updated Rectangle List:", rectangle_list)

            # Update the Treeview widget with the rectangle coordinates
            areas_tree.insert("", ctk.END, values=(adjusted_coordinates[0], adjusted_coordinates[1],
                                                   adjusted_coordinates[2], adjusted_coordinates[3]))
        else:
            print("Error: Failed to retrieve bounding box coordinates for current_rectangle")
    else:
        print("Error: 'current_rectangle' is None in end_rectangle")


def update_rectangles():
    global areas, current_zoom, canvas, rectangle_list, areas_tree

    # Delete existing rectangles on the canvas
    for rectangle_id in rectangle_list:
        canvas.delete(rectangle_id)
    rectangle_list.clear()

    # Clear the areas_tree
    for item in areas_tree.get_children():
        areas_tree.delete(item)

    # Update rectangles based on the new zoom level
    for stored_coords in areas:
        x0, y0, x1, y1 = stored_coords

        # Adjust coordinates based on the current zoom level
        adjusted_coords = [
            x0 * current_zoom,
            y0 * current_zoom,
            x1 * current_zoom,
            y1 * current_zoom
        ]

        # Draw the rectangle on the canvas
        rectangle_id = canvas.create_rectangle(
            adjusted_coords[0], adjusted_coords[1],
            adjusted_coords[2], adjusted_coords[3],
            outline="red", width=2
        )

        # Append the new rectangle ID to the rectangle_list
        rectangle_list.append(rectangle_id)

        # Update the areas_tree with the adjusted coordinates
        areas_tree.insert("", "end", values=(x0, y0, x1, y1))


def on_zoom_slider_change(value):
    global current_zoom
    current_zoom = float(value)
    update_display()


def check_resize(event):
    global last_resize_time, resize_delay
    # Get the current time
    current_time = time.time()

    # Calculate the time elapsed since the last resize event
    time_elapsed = current_time - last_resize_time

    if time_elapsed >= resize_delay:
        # Execute the function if the delay has passed since the last resize event
        # print("No resize event for the past", resize_delay, "seconds.")
        on_windowresize()


def on_windowresize(event=None):
    global prev_width, prev_height, last_resize_time, resize_delay

    current_width = root.winfo_width()
    current_height = root.winfo_height()


    if current_width != prev_width or current_height != prev_height:
        prev_width = current_width
        prev_height = current_height

        print(f'Display Resized: {root.winfo_width() - 30}, {root.winfo_height() - 135}')
        last_resize_time = time.time()
        update_display()


def update_display():
    global root, canvas, pdf_width, pdf_height, current_zoom, v_scrollbar, h_scrollbar, pix, page

    # Set canvas dimensions based on aspect ratio and desired size
    canvas_width = root.winfo_width() - 30
    canvas_height = root.winfo_height() - 135

    # Scrollbar reposition
    v_scrollbar.configure(command=canvas.yview, height=canvas_height)
    h_scrollbar.place_configure(x=canvas_width + 14, y=100)
    h_scrollbar.configure(command=canvas.xview, width=canvas_width)
    h_scrollbar.place_configure(x=10, y=canvas_height + 107)

    # Resize the canvas
    canvas.config(width=canvas_width, height=canvas_height)

    # Check if `page` is valid before continuing
    if page is None:
        print("No valid page to display.")
        return

    try:
        # Get the currently displayed image on the canvas
        current_image = getattr(canvas, 'pdf_image', None)

        # Clear the existing image on the canvas
        canvas.delete("all")

        # Get the pixmap from the page
        pix = page.get_pixmap(matrix=fitz.Matrix(current_zoom, current_zoom))
        img = pix.tobytes("ppm")
        img_tk = tk.PhotoImage(data=img)

        # Display the updated image on the canvas
        canvas.create_image(0, 0, anchor=tk.NW, image=img_tk, tags="pdf_image")

        # Update the image reference in the canvas
        canvas.pdf_image = img_tk

        # Adjust the canvas scroll region
        zoomed_width = int(pdf_width * current_zoom)
        zoomed_height = int(pdf_height * current_zoom)

        # Configure canvas to use scrollbars
        canvas.config(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set,
                      scrollregion=(0, 0, zoomed_width, zoomed_height))

    except ValueError as e:
        print(f"Error updating display: {e}")

    update_rectangles()


def display_sample_pdf(pdf_path):
    global current_zoom, canvas, zoom_slider, page, pdf_width, pdf_height, pix, pdf_document

    pdf_document = fitz.open(pdf_path)
    page = pdf_document[0]  # Access the first page

    # Print the bounding box information for the first page
    print(f"Page 1 Bounding Box: {page.rect}")

    # Get the size of the PDF page and round to the nearest integer
    pdf_width, pdf_height = int(round(page.rect.width)), int(round(page.rect.height))

    if not hasattr(canvas, 'pdf_image'):
        # Create a zoom slider
        zoom_slider = ctk.CTkSlider(root, from_=0.5, to=3.5, command=on_zoom_slider_change,
                                    height=10,
                                    width=150,
                                    border_width=0,
                                    # number_of_steps=12,
                                    # button_color="#B30B00",
                                    # button_hover_color="#860A00",
                                    orientation="horizontal")
        zoom_slider.set(current_zoom)  # Initial zoom level
        zoom_slider.place(x=800, y=80)

        zoom_slider_label = ctk.CTkLabel(root, text="Zoom:", font=(button_font, 9))
        zoom_slider_label.place(x=760, y=70)

        # Bind events to handle rectangle drawing
        canvas.bind("<ButtonPress-1>", start_rectangle)
        canvas.bind("<B1-Motion>", draw_rectangle)
        canvas.bind("<ButtonRelease-1>", end_rectangle)

    print(f"Displayed PDF: {pdf_path}")

    update_display()

    canvas.yview_moveto(1.0)
    canvas.xview_moveto(1.0)

    print(f"PDF Dimensions: {pdf_width} x {pdf_height}")
    print(f"PDF Rotation: {page.rotation} x {page.rotation_matrix}")

    return pdf_width, pdf_height, page


def close_pdf():
    global canvas, pdf_document

    # Remove the PDF image using its tag or ID
    canvas.delete("pdf_image")

    # Close the PDF document if it's open
    try:
        if pdf_document:
            pdf_document.close()
            print("PDF document closed.")
        else:
            print("No PDF document to close.")
    except Exception as e:
        print(f"Error closing PDF: {e}")

    # Reset the reference to the PDF document
    pdf_document = None


def clear_all_areas():
    global areas, areas_tree

    # Clear the areas list
    areas = []

    # Clear the areas Treeview widget
    for item in areas_tree.get_children():
        areas_tree.delete(item)

    update_display()

    print("Cleared All Areas")


def get_cell_dimensions(sheet, cell):
    col_letter = get_column_letter(cell.column)
    col_width = sheet.column_dimensions[col_letter].width or 8.43  # Default width if not set
    row_height = sheet.row_dimensions[cell.row].height or 15  # Default height if not set
    return col_width * 7, row_height  # Approximate width in pixels


def adjust_coordinates_for_rotation(coordinates, rotation, pdf_height, pdf_width):
    if rotation == 0:
        return coordinates
    elif rotation == 90:
        x0, y0, x1, y1 = coordinates
        return [y0, pdf_width - x1, y1, pdf_width - x0]
    elif rotation == 180:
        x0, y0, x1, y1 = coordinates
        return [pdf_width - x1, pdf_height - y1, pdf_width - x0, pdf_height - y0]
    elif rotation == 270:
        x0, y0, x1, y1 = coordinates
        return [pdf_height - y1, x0, pdf_height - y0, x1]


def start_extraction_thread():
    extraction_thread = threading.Thread(target=extract_text)
    extraction_thread.start()


def extract_text():
    close_pdf()
    start_time = time.time()
    global areas, ws, pdf_height, pdf_width, include_subfolders, enable_ocr, tessdata_folder, dpi_value

    if not areas:
        messagebox.showwarning("Error", "No areas defined. Please draw rectangles.")
        return

    pdf_folder_value = pdf_folder_entry.get()
    output_path_value = output_path_entry.get()

    if not pdf_folder_value or not output_path_value or not os.path.isdir(pdf_folder_value):
        messagebox.showerror("Invalid Folder", "The specified folder does not exist.")
        return

    # Initialize Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.append(["Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page Number"] +
              [f"Area {i + 1}" for i in range(len(areas))])

    progress_window = ctk.CTkToplevel(root)
    progress_window.title("Extraction in Progress...")
    progress_window.geometry("300x90")
    progress_window.attributes('-topmost', True)
    progress_window.lift()

    progress_label = ctk.CTkLabel(progress_window, text="Loading PDFs...")
    progress_label.pack()

    progress = ctk.CTkProgressBar(progress_window, orientation="horizontal", mode="determinate",
                                  progress_color='limegreen', width=150, height=15)
    progress.pack()

    # Determine the total number of iterations (PDF files)
    if include_subfolders:
        total_files = sum(
            1 for root, _, files in os.walk(pdf_folder) for file in files if file.lower().endswith('.pdf'))
    else:
        _, _, files = next(os.walk(pdf_folder))
        total_files = sum(1 for file in files if file.lower().endswith('.pdf'))

    processed_files = 0
    temp_image_paths = []  # Track temporary images if images are used in extraction

    for root_folder, subfolders, files in os.walk(pdf_folder):
        if not include_subfolders:
            subfolders.clear()

        for pdf_filename in files:
            if pdf_filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(root_folder, pdf_filename)

                try:
                    file_size = os.path.getsize(pdf_path)
                    last_modified_date = datetime.fromtimestamp(os.path.getmtime(pdf_path)).strftime(
                        '%Y-%m-%d %H:%M:%S')

                    pdf_document = fitz.open(pdf_path)
                    for page_number in range(pdf_document.page_count):
                        page = pdf_document[page_number]
                        page_specific_row = [
                            file_size,  # Size (Bytes)
                            last_modified_date,  # Date Last Modified
                            os.path.relpath(root_folder, pdf_folder),  # Folder
                            pdf_filename,  # Filename
                            page_number + 1  # Page Number
                        ]

                        for i, area_coordinates in enumerate(areas, start=1):
                            pdf_width, pdf_height = page.rect.width, page.rect.height
                            adjusted_coordinates = adjust_coordinates_for_rotation(area_coordinates, page.rotation,
                                                                                   pdf_height, pdf_width)
                            try:
                                if enable_ocr == "Text-first":
                                    text_area = page.get_text("text", clip=adjusted_coordinates)
                                    if not text_area.strip():
                                        pix = page.get_pixmap(clip=area_coordinates, dpi=dpi_value)
                                        pdfdata = pix.pdfocr_tobytes(language="eng", tessdata=tessdata_folder)
                                        clipdoc = fitz.open("pdf", pdfdata)
                                        text_area = "_OCR_" + clipdoc[0].get_text()
                                elif enable_ocr == "OCR-All":
                                    pix = page.get_pixmap(clip=area_coordinates, dpi=dpi_value)
                                    pdfdata = pix.pdfocr_tobytes(language="eng", tessdata=tessdata_folder)
                                    clipdoc = fitz.open("pdf", pdfdata)
                                    text_area = "_OCR_" + clipdoc[0].get_text()
                                elif enable_ocr == "Text1st+Image-beta":
                                    text_area = page.get_text("text", clip=adjusted_coordinates)
                                    pix = page.get_pixmap(clip=area_coordinates, dpi=dpi_value)
                                    if not text_area.strip():
                                        pdfdata = pix.pdfocr_tobytes(language="eng", tessdata=tessdata_folder)
                                        clipdoc = fitz.open("pdf", pdfdata)
                                        text_area = "_OCR_" + clipdoc[0].get_text()
                                    img_path = f'{pdf_filename}_{page_number}_{i}.png'
                                    pix.save(img_path)
                                    temp_image_paths.append(img_path)
                                    img = ExcelImage(img_path)
                                    ws.add_image(img,
                                                 f"{get_column_letter(len(page_specific_row) + 1)}{ws.max_row + 1}")
                                else:
                                    text_area = page.get_text("text", clip=adjusted_coordinates)

                                text_area = text_area.replace('\n', ' ').strip().replace('  ', ' ')
                                page_specific_row.append(text_area)

                            except Exception as e:
                                print(
                                    f"Error extracting text from area {i} in {pdf_filename}, Page {page_number + 1}: {e}")
                                page_specific_row.append("Error extracting text")

                        ws.append(page_specific_row)

                        pdf_filename_cell = ws.cell(row=ws.max_row, column=4)
                        pdf_filename_cell.value = pdf_filename
                        pdf_filename_cell.font = Font(color="0000FF")
                        pdf_filename_cell.hyperlink = Hyperlink(target=pdf_path, ref=f"D{ws.max_row}")

                    processed_files += 1
                    progress_label.configure(text=f"Processing PDFs... ({processed_files}/{total_files})")
                    progress.set(processed_files / total_files)
                    progress_window.update_idletasks()

                except Exception as e:
                    print(f"Error processing {pdf_filename}: {e}")

    try:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if "_OCR_" in str(cell.value):
                    cell.value = cell.value.replace("_OCR_", "")
                    cell.font = Font(color="FF3300")

        wb.save(output_path_value)

        for temp_img_path in temp_image_paths:
            try:
                os.remove(temp_img_path)
            except OSError as e:
                print(f"Error: {temp_img_path} : {e.strerror}")

    except PermissionError:
        timestamp = time.strftime("%Y%m%d%H%M%S")
        timestamped_output_path = f"{os.path.splitext(output_path_value)[0]}_{timestamp}.xlsx"
        wb.save(timestamped_output_path)
        print(f"A copy has been created: {timestamped_output_path}")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Text extraction completed in {elapsed_time:.2f} seconds.")

    # Prompt to open the Excel file
    open_file = messagebox.askyesno("Open Excel File", f"Elapsed Time: {elapsed_time:.2f} seconds\n\nDo you want to open the Excel file now?")
    if open_file:
        os.startfile(output_path_value)

    # Close progress window
    progress_window.destroy()


def after_command():
    root.bind("<Configure>", check_resize)
    canvas.bind("<MouseWheel>", on_mousewheel)
    canvas.bind("<Shift-MouseWheel>", on_mousewheel)  # Shift + Scroll
    # Bind right-click on the canvas to display the popup menu
    canvas.bind("<Button-3>", show_popup_menu)
    pdf_folder_entry.bind("<KeyRelease>", update_pdf_folder)
    output_path_entry.bind("<KeyRelease>", update_output_path)


def update_pdf_folder(event):
    global pdf_folder
    pdf_folder = pdf_folder_entry.get()
    print(f"PDF root path: {pdf_folder}")


def update_output_path(event):
    global output_excel_path
    output_excel_path = output_path_entry.get()
    print(f"output path: {output_excel_path}")


def find_tessdata():
    global tessdata_folder  # Use the global variable

    # Define the subdirectories
    tesseract_subdirectory = "Tesseract-OCR"
    tessdata_subdirectory = "tessdata"

    # Check Program Files directory
    program_files_dir = os.path.join("C:", os.sep, "Program Files", tesseract_subdirectory, tessdata_subdirectory)
    if os.path.exists(program_files_dir):
        tessdata_folder = program_files_dir
        os.environ["TESSDATA_PREFIX"] = program_files_dir
        return tessdata_folder

    # Get the local application data directory
    local_programs_dir = os.path.join(os.getenv("LOCALAPPDATA"), "Programs")

    # Search in the local Programs directory
    local_programs_path = os.path.join(local_programs_dir, tesseract_subdirectory, tessdata_subdirectory)
    if os.path.exists(local_programs_path):
        tessdata_folder = local_programs_path
        os.environ["TESSDATA_PREFIX"] = local_programs_path
        return tessdata_folder

    # Get the platform-independent local application data directory
    app_data_dir = os.path.join(os.getenv("APPDATA"), tesseract_subdirectory, tessdata_subdirectory)

    # Check if the folder exists
    if os.path.exists(app_data_dir):
        tessdata_folder = app_data_dir
        os.environ["TESSDATA_PREFIX"] = app_data_dir
        return tessdata_folder

    # If not found, prompt the user to browse manually
    manual_path = filedialog.askdirectory(title="Select Tesseract TESSDATA folder manually")

    if os.path.exists(manual_path):
        tessdata_folder = manual_path
        os.environ["TESSDATA_PREFIX"] = manual_path
        return tessdata_folder
    else:
        print("Invalid path. Tesseract tessdata folder not found.")
        return None


def ocr_menu_callback(choice):
    global enable_ocr

    print("OCR menu dropdown clicked:", choice)

    def enable_ocr_menu(enabled):
        color = "green4" if enabled else "gray29"
        ocr_menu.configure(fg_color=color, button_color=color)
        dpi_menu.configure(state="normal" if enabled else "disabled", fg_color=color, button_color=color)

    if choice == "Off":
        enable_ocr_menu(False)
        print("OCR disabled.")

    elif choice in ("Text-first", "OCR-All", "Text1st+Image-beta"):
        found_tesseract_path = find_tessdata()

        if found_tesseract_path:
            enable_ocr_menu(True)
            if choice == "Text-first":
                print("OCR will start if no text are extracted.")
            else:
                print("OCR will be enabled for every area.")
        else:
            enable_ocr_menu(False)

    else:
        # Handle other options
        print("Selected option:", choice)

    enable_ocr = choice
    print("OCR mode:", enable_ocr)


def dpi_callback(choice):
    global dpi_value
    dpi_value = dpi_options.get(choice, 150)
    print("DPI:", dpi_value)
    return dpi_value


def optionmenu_callback(choice):
    print("Option:", choice)

    # Retrieve the corresponding action from the dictionary and execute it
    action = option_actions.get(choice)
    if action:
        action()
    else:
        print("Selected option:", choice)


def on_mousewheel(event):
    if event.state & 0x1:  # Check if the Shift key is being held down
        canvas.xview_scroll(-1 * int(event.delta / 120), "units")
    else:
        canvas.yview_scroll(-1 * int(event.delta / 120), "units")


def export_rectangles():
    global areas

    # Ask the user where to save the file
    export_file_path = filedialog.asksaveasfilename(
        title="Save Rectangles As",
        defaultextension=".json",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )

    if export_file_path:
        try:
            # Save the list of areas (rectangle coordinates) to a JSON file
            with open(export_file_path, 'w') as json_file:
                json.dump(areas, json_file)
            print(f"Rectangles exported to {export_file_path}")
        except Exception as e:
            print(f"Error while exporting rectangles: {e}")
    else:
        print("Export canceled by the user.")


def import_rectangles():
    global areas, canvas, current_zoom, rectangle_list, areas_tree

    # Ask the user to select the JSON file to import
    import_file_path = filedialog.askopenfilename(
        title="Import Rectangles",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )

    if import_file_path:
        try:
            # Load the rectangles (areas) from the JSON file
            with open(import_file_path, 'r') as json_file:
                imported_areas = json.load(json_file)

            areas = imported_areas  # Update the global areas list

            # Clear existing rectangles and their IDs
            for rectangle_id in rectangle_list:
                canvas.delete(rectangle_id)
            rectangle_list.clear()

            # Clear the areas_tree
            for item in areas_tree.get_children():
                areas_tree.delete(item)

            # Redraw rectangles on the canvas and update the areas_tree
            update_rectangles()

            print(f"Rectangles imported from {import_file_path}")
        except Exception as e:
            print(f"Error while importing rectangles: {e}")
    else:
        print("Import canceled by the user.")


def open_recent_pdf():
    global recent_pdf_path

    if recent_pdf_path and os.path.exists(recent_pdf_path):
        print(f"Opening recent PDF: {recent_pdf_path}")
        display_sample_pdf(recent_pdf_path)
    else:
        print(f"File not found: {recent_pdf_path}")



# Create main window
root = ctk.CTk()
root.title("Xtractor by RRR")
root.geometry(f"{initial_width}x{initial_height}+{initial_x_position}+{initial_y_position}")

# Create the popup menu
popup_menu = tk.Menu(root, tearoff=0)
popup_menu.add_command(label="Delete Rectangle", command=lambda: delete_rectangle())

#OCR Widgets
ocr_menu_var = ctk.StringVar(value="OCR-OFF")
ocr_menu = ctk.CTkOptionMenu(root, values=["Off", "Text-first", "OCR-All","Text1st+Image-beta"],
                               command=ocr_menu_callback, font=("Verdana Bold", 9),
                               button_color=("gray29", "gray39"), fg_color=("gray29", "gray39"),
                               dropdown_font=(button_font, 9), dynamic_resizing=False,
                               #state="disabled", #OCR Mode currently disabled
                               variable=ocr_menu_var, width=85, height=18)
ocr_menu.place(x=330, y=10)


#DPI Widgets
dpi_var = ctk.IntVar(value=150)
dpi_value = 150
dpi_menu_values = list(dpi_options.keys())  # Use dpi_options here
dpi_menu = ctk.CTkOptionMenu(root, values=dpi_menu_values,
                               command=dpi_callback, font=("Verdana Bold", 7),
                               button_color=("gray29", "gray39"), fg_color=("gray29", "gray39"),
                               state="disabled", dropdown_font=(button_font, 8),
                               dynamic_resizing=False, variable=dpi_var, width=43, height=14)
dpi_menu.place(x=372, y=30)

dpi_label = ctk.CTkLabel(root, text="DPI:", text_color="gray59",
                             padx=0, pady=0, anchor="nw",
                             font=("Verdana Bold", 8))
dpi_label.place(x=348, y=32)


# PDF Folder
pdf_folder_entry = ctk.CTkEntry(root, width=270, height=20, font=(button_font, 9),
                                placeholder_text="Select Folder with PDFs", border_width=1,
                                corner_radius=3)
pdf_folder_entry.place(x=50, y=10)
pdf_folder_button = ctk.CTkButton(root, text="...", command=browse_pdf_folder, font=(button_font, 9),
                                  # fg_color="#B30B00", # hover_color="#860A00",
                                  width=25, height=10)
pdf_folder_button.place(x=20, y=10)


# Open Sample PDF Button
open_sample_button = ctk.CTkButton(root, text="Open PDF", command=open_sample_pdf, font=(button_font, 9),
                                   # fg_color="#B30B00", # hover_color="#860A00",
                                   width=25, height=10)
open_sample_button.place(x=20, y=35)





# Output Excel Path
output_path_entry = ctk.CTkEntry(root, width=270, height=20, font=(button_font, 9),
                                 placeholder_text="Select Folder for Excel output",
                                 border_width=1, corner_radius=3)
output_path_entry.place(x=50, y=60)
output_path_button = ctk.CTkButton(root, text="...", command=browse_output_path, font=(button_font, 9),
                                   fg_color="#217346", hover_color="#6AD49A",
                                   width=25, height=10)
output_path_button.place(x=20, y=60)


# Bind events to the pdf_folder_entry and output_path_entry widgets
pdf_folder_entry.bind("<KeyRelease>", update_pdf_folder)
output_path_entry.bind("<KeyRelease>", update_output_path)


#
include_subfolders_var = ctk.IntVar()

include_subfolders_checkbox = ctk.CTkCheckBox(root, text="Include Subfolders?", variable=include_subfolders_var,
                                              command=toggle_include_subfolders,
                                              checkbox_width=17,
                                              checkbox_height=17,
                                              border_width=1,
                                              # fg_color="#B30B00",hover_color="#860A00",
                                              font=(button_font, 9))

include_subfolders_checkbox.place(x=196, y=34)  # Adjusted the y-coordinate

# Initialize include_subfolders
include_subfolders = include_subfolders_var.get()


# Extract Text Button
extract_button = ctk.CTkButton(root, text="EXTRACT", font=("Arial Black", 12),
                               # fg_color="#217346",
                               # hover_color="#6AD49A",
                               corner_radius=10,
                               width=75, height=30, command=start_extraction_thread)
extract_button.place(x=330, y=55)

root.after(2500, after_command)



# PDF Display
canvas = ctk.CTkCanvas(root, width=canvas_width, height=canvas_height)
canvas.place(x=10, y=100)

v_scrollbar = ctk.CTkScrollbar(root, orientation="vertical", command=canvas.yview, height=canvas_height)
v_scrollbar.place(x=canvas_width + 14, y=100)
h_scrollbar = ctk.CTkScrollbar(root, orientation="horizontal", command=canvas.xview, width=canvas_width)
h_scrollbar.place(x=10, y=canvas_height + 105)



# Areas Table
areas_frame = ctk.CTkFrame(root, height=1, width=200,
                           # label_font=(button_font,8),
                           border_width=0)
areas_frame.place(x=425, y=10)
areas_frame.bind("<Double-Button-1>", lambda event: edit_areas())

style = ttk.Style()
style.configure("mystyle.Treeview", bd=0, font=('Verdana', 5))
style.configure("mystyle.Treeview.Heading", font=('Verdana', 6, 'bold'))


# Areas Table Treeview
area_columns_data = [("x0", "x0"),
                     ("y0", "y0"),
                     ("x1", "x1"),
                     ("y1", "y1")]

areas_tree = ttk.Treeview(areas_frame, columns=[column[0] for column in area_columns_data], show="headings", style="mystyle.Treeview", height=3)

for heading, column_id in area_columns_data:
    areas_tree.heading(column_id, text=heading)
    areas_tree.column(column_id, minwidth=0, width=50)

areas_tree.pack(side="left")
areas_tree.bind("<Double-Button-1>", lambda event, areas_tree_ref=areas_tree, areas=areas: edit_areas(event, areas_tree_ref))


# scrollbar for table
tree_scrollbar = ctk.CTkScrollbar(areas_frame, orientation="vertical", command=areas_tree.yview, minimum_pixel_length=3, height=20)
tree_scrollbar.pack(side="right", fill="y")
areas_tree.configure(yscrollcommand=tree_scrollbar.set)


# Import Rectangles Button
import_button = ctk.CTkButton(root, text="Import Areas", command=import_rectangles,
                              font=(button_font, 9), width=88, height=10)
import_button.place(x=650, y=15)

# Export Rectangles Button
export_button = ctk.CTkButton(root, text="Export Areas", command=export_rectangles,
                              font=(button_font, 9), width=88, height=10)
export_button.place(x=650, y=40)

# Clear Areas Button
clear_areas_button = ctk.CTkButton(root, text="Clear Areas", command=clear_all_areas, font=(button_font, 9),
                                   fg_color=("gray29", "gray39"), width=88, height=10)
clear_areas_button.place(x=650, y=65)



# Create the option menu
optionmenu_var = ctk.StringVar(value="Other Features")
optionmenu = ctk.CTkOptionMenu(root, values=list(option_actions.keys()),
                               command=optionmenu_callback,
                               font=(button_font, 9), dropdown_font=(button_font, 9),
                               dynamic_resizing=False,
                               variable=optionmenu_var, width=105, height=15)
optionmenu.place(x=850, y=10)

def version_text(event):
    version_text = """
    Created by:
    Rei Raphael Reveral
    
    Links:
    https://github.com/r-Yayap/MultiplePDF-Areas2Excel
    https://www.linkedin.com/in/rei-raphael-reveral
    
    """

    # Create a Toplevel window
    window = ctk.CTkToplevel(root)
    window.title("---")

    # Create a Text widget
    text_widget = ctk.CTkTextbox(window, wrap="word", width=400, height=247)
    text_widget.insert(tk.END, version_text)

    # Pack the Text widget and scrollbar
    text_widget.pack(padx=10, pady=10, side="left")

    # Bring the specific text window to the front
    window.grab_set()


version_label = ctk.CTkLabel(root, text=version_txt, fg_color="transparent", text_color="gray59", padx=0, pady=0,anchor="nw", font=(button_font, 9.5))
version_label.place(x=835, y=30)
version_label.bind("<Button-1>", version_text)


# Tooltips using the helper function
tooltip_ocr_menu = create_tooltip(ocr_menu, """Off                  -   no OCR
Text-first       -   Will OCR if extracted is blank
OCR-All         -   OCR all selected Areas
Text-Image   -   Extract Text along and a clipped image""")

tooltip_dpi_menu = create_tooltip(dpi_menu, "DPI/resolution", font=("Verdana", 8))

tooltip_pdf_folder_entry = create_tooltip(pdf_folder_entry, "Main folder where all PDF files are located")
tooltip_pdf_folder_button = create_tooltip(pdf_folder_button, "Select the Main folder where all PDF files are located")

tooltip_open_sample_button = create_tooltip(open_sample_button, "Open a PDF with the Title Block Template")

tooltip_output_path_entry = create_tooltip(output_path_entry, "Folder where result will be saved")
tooltip_output_path_button = create_tooltip(output_path_button, "Choose the folder where result will be saved")

tooltip_include_subfolders_checkbox = create_tooltip(include_subfolders_checkbox,
                                                     "If your Main Folder contains PDF in its subfolders then tick the box")

tooltip_extract_button = create_tooltip(extract_button, "Start Extraction")

tooltip_import_button = create_tooltip(import_button, "Load a saved Template of selected areas")
tooltip_export_button = create_tooltip(export_button, "Save a Template of the selected areas")
tooltip_clear_areas_button = create_tooltip(clear_areas_button, "Clear all selected areas")

tooltip_optionmenu = create_tooltip(optionmenu, """Features:

PDF/DWG List   -   List all PDF and DWG side by side to check if they have the same filename
Directory List    -   List all files in the selected directory
Bulk Renamer    -   Bulk rename all files with CSV - Column A: old names, Column B: new names""", font=("Verdana", 10))

tooltip_version_label = create_tooltip(version_label, version_txt, font=("Verdana", 12))

# Run the main loop
root.mainloop()

