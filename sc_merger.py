# Conflux V1
import os
import re
import tkinter as tk
import sys
import customtkinter as ctk
import pandas as pd
from difflib import SequenceMatcher
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from tkinterdnd2 import TkinterDnD, DND_ALL

def resource_path(relative_path):
    """Get the absolute path to a resource (used for PyInstaller compatibility)"""
    try:
        base_path = sys._MEIPASS  # PyInstaller's temp folder
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

theme_path = resource_path("style/conflux-dark-red.json")
ctk.set_default_color_theme(theme_path)
ctk.set_appearance_mode("dark")

# Helper to auto-select a header based on keywords.
def auto_select_header(headers, keywords):
    for header in headers:
        lower_header = header.lower()
        for kw in keywords:
            if kw in lower_header:
                return header
    return headers[0] if headers else ""


class ExcelMerger:
    """Handles Excel merging logic, including conditional formatting, hyperlink handling,
    and title rich-text highlighting (using tokenization and dynamic programming)."""

    @staticmethod
    def merge_3_excels(
            excel1_path, excel2_path, excel3_path,
            ref_column1, ref_column2, ref_column3,
            output_path,
            title_column1=None, title_column2=None, title_column3=None,
            compare_excel2=False, compare_excel3=False,
            status_column=None, status_value=None,
            project_column=None, project_value=None,
            custom_checks=None, filename_column=None

    ):

        # Read Excel files
        df1 = pd.read_excel(excel1_path, engine='openpyxl', dtype=str).fillna("")
        df2 = pd.read_excel(excel2_path, engine='openpyxl', dtype=str).fillna("")
        df3 = pd.read_excel(excel3_path, engine='openpyxl', dtype=str).fillna("")

        # Ensure the reference columns exist
        for df, ref_col in zip([df1, df2, df3],
                               [ref_column1, ref_column2, ref_column3]):
            if ref_col not in df.columns:
                raise KeyError(f"Reference column '{ref_col}' not found in one of the Excel files.")

        # For hyperlink purposes, add original row indices to df1
        df1 = ExcelMerger.add_original_row_index_to_dataframe(df1, excel1_path)
        # Extract hyperlinks from Excel1
        hyperlinks = ExcelMerger._extract_hyperlinks(excel1_path)

        # Compute occurrence counts
        df1['refno_count'] = df1.groupby(ref_column1).cumcount()
        df2['refno_count'] = df2.groupby(ref_column2).cumcount()
        df3['refno_count'] = df3.groupby(ref_column3).cumcount()

        # Rename reference columns to fixed names
        df1 = df1.rename(columns={ref_column1: 'number_1'})
        df2 = df2.rename(columns={ref_column2: 'number_2'})
        df3 = df3.rename(columns={ref_column3: 'number_3'})

        # Create a common key column from the reference values
        df1['common_ref'] = df1['number_1']
        df2['common_ref'] = df2['number_2']
        df3['common_ref'] = df3['number_3']

        # Rename title columns if provided
        if title_column1:
            df1 = df1.rename(columns={title_column1: 'title_excel1'})
        if title_column2:
            df2 = df2.rename(columns={title_column2: 'title_excel2'})
        if title_column3:
            df3 = df3.rename(columns={title_column3: 'title_excel3'})

        # Merge df1 and df2 on common_ref and refno_count
        merged_df = pd.merge(
            df1, df2,
            on=['common_ref', 'refno_count'],
            how='outer',
            suffixes=('_1', '_2')
        ).fillna("")

        # Merge the result with df3
        merged_df = pd.merge(
            merged_df, df3,
            on=['common_ref', 'refno_count'],
            how='outer',
            suffixes=("", "_3")
        ).drop(columns=['refno_count']).fillna("")

        if 'original_row_index' in merged_df.columns:
            merged_df['original_row_index'] = pd.to_numeric(
                merged_df['original_row_index'], errors='coerce').fillna(0).astype(int)

        print("Merged DataFrame columns (3 files):")
        print(merged_df.columns.tolist())

        # --- NEW: Add title_match column ---
        if title_column1 and title_column2 and title_column3 and compare_excel2 and compare_excel3:
            merged_df['title_match'] = merged_df.apply(
                lambda
                    row: f"{ExcelMerger.titles_match(str(row.get('title_excel1', '')).strip(), str(row.get('title_excel2', '')).strip())}, " \
                         f"{ExcelMerger.titles_match(str(row.get('title_excel1', '')).strip(), str(row.get('title_excel3', '')).strip())}",
                axis=1
            )
        elif title_column1 and title_column2 and compare_excel2:
            merged_df['title_match'] = merged_df.apply(
                lambda
                    row: f"{ExcelMerger.titles_match(row.get('title_excel1', ''), row.get('title_excel2', ''))}",
                axis=1
            )
        elif title_column1 and title_column3 and compare_excel3:
            merged_df['title_match'] = merged_df.apply(
                lambda
                    row: f"{ExcelMerger.titles_match(row.get('title_excel1', ''), row.get('title_excel3', ''))}",
                axis=1
            )
        else:
            merged_df['title_match'] = "N/A, N/A"

        # ✅ Ensure Comments_1 is updated even if title comparison is off
        merged_df = ExcelMerger.update_comments_column(
            merged_df,
            status_column=status_column, status_value=status_value,
            project_column=project_column, project_value=project_value,
            custom_checks=custom_checks,
            filename_column=filename_column
        )

        temp_file_path = ExcelMerger._save_merged_to_excel(merged_df, output_path)

        # *** IMPORTANT: First, reorder columns ***
        ExcelMerger.reorder_columns(temp_file_path)

        # Reload headers to reflect changes after reorder_columns
        wb = load_workbook(temp_file_path)
        ws = wb.active
        updated_headers = [cell.value for cell in ws[1]]
        col_name_to_excel_idx = {col: idx + 1 for idx, col in enumerate(updated_headers)}
        wb.close()


        # *** Then apply formatting and Hyperlinks ***
        ExcelMerger._apply_formatting_and_hyperlinks(
            temp_file_path, hyperlinks, merged_df,
            status_column=status_column, status_value=status_value,
            project_column=project_column, project_value=project_value,
            custom_checks=custom_checks,
            col_name_to_excel_idx=col_name_to_excel_idx  # <-- Pass the mapping here
        )

        # *** Then apply colors to True or False ***
        ExcelMerger.apply_title_match_highlighting(temp_file_path, merged_df)


        # *** Then apply title highlighting ***

        if title_column1 and title_column2 and title_column3 and compare_excel2 and compare_excel3:
            if title_column1 and title_column2:
                ExcelMerger.apply_title_highlighting(
                    temp_file_path, merged_df, title_col1='title_excel1',title_col2='title_excel2',
                    reorder=False, update_baseline=True,
                    status_column=status_column, status_value=status_value,
                    project_column=project_column, project_value=project_value,
                    custom_checks=custom_checks, filename_column=filename_column

                )

            ExcelMerger.apply_title_highlighting(
                temp_file_path, merged_df, 'title_excel1', 'title_excel3',
                reorder=False, update_baseline=True, filename_column=filename_column

            )

        # Explicitly handle Excel 2 independently
        elif title_column1 and title_column2 and compare_excel2:
            ExcelMerger.apply_title_highlighting(
                temp_file_path, merged_df, 'title_excel1', 'title_excel2',
                reorder=False, update_baseline=True, filename_column=filename_column

            )

        # Explicitly handle Excel 3 independently
        elif title_column1 and title_column3 and compare_excel3:
            ExcelMerger.apply_title_highlighting(
                temp_file_path, merged_df, 'title_excel1', 'title_excel3',
                reorder=False, update_baseline=True, filename_column=filename_column

            )

        return temp_file_path, merged_df

    @staticmethod
    def merge_excels(
            excel1_path, excel2_path,
            ref_column1, ref_column2,
            output_path,
            title_column1=None, title_column2=None,
            status_column=None, status_value=None,
            project_column=None, project_value=None,
            custom_checks=None, filename_column=None

    ):

        # Two-file merge code
        excel1 = pd.read_excel(excel1_path, engine='openpyxl', dtype=str).fillna("")
        excel2 = pd.read_excel(excel2_path, engine='openpyxl', dtype=str).fillna("")

        if ref_column1 not in excel1.columns or ref_column2 not in excel2.columns:
            raise KeyError("Reference columns not found in one or both Excel files.")

        excel1 = ExcelMerger.add_original_row_index_to_dataframe(excel1, excel1_path)
        hyperlinks = ExcelMerger._extract_hyperlinks(excel1_path)

        excel1['refno_count'] = excel1.groupby(ref_column1).cumcount()
        excel2['refno_count'] = excel2.groupby(ref_column2).cumcount()

        excel1 = excel1.rename(columns={ref_column1: 'number_1'})
        excel2 = excel2.rename(columns={ref_column2: 'number_2'})
        excel1['common_ref'] = excel1['number_1']
        excel2['common_ref'] = excel2['number_2']

        if title_column1:
            excel1 = excel1.rename(columns={title_column1: 'title_excel1'})
        if title_column2:
            excel2 = excel2.rename(columns={title_column2: 'title_excel2'})

        merged_df = pd.merge(
            excel1, excel2,
            on=['common_ref', 'refno_count'],
            how='outer',
            suffixes=('_1', '_2')
        ).drop(columns=['refno_count']).fillna("")

        merged_df['original_row_index'] = pd.to_numeric(
            merged_df['original_row_index'], errors='coerce').fillna(0).astype(int)

        print("Merged DataFrame columns:")
        print(merged_df.columns.tolist())

        # ✅ Add Comments_1 if needed
        merged_df = ExcelMerger.update_comments_column(
            merged_df,
            status_column=status_column, status_value=status_value,
            project_column=project_column, project_value=project_value,
            custom_checks=custom_checks,
            filename_column=filename_column
        )

        if title_column1 and title_column2:
            merged_df['title_match'] = merged_df.apply(
                lambda row: "True" if ExcelMerger.titles_match(
                    str(row.get('title_excel1', '')).strip(), str(row.get('title_excel2', '')).strip()
                ) else "False",
                axis=1
            )
        else:
            merged_df['title_match'] = "N/A"

        temp_file_path = ExcelMerger._save_merged_to_excel(merged_df, output_path)

        # ✅ Reorder columns and reload headers
        ExcelMerger.reorder_columns(temp_file_path)
        wb = load_workbook(temp_file_path)
        ws = wb.active
        updated_headers = [cell.value for cell in ws[1]]
        col_name_to_excel_idx = {col: idx + 1 for idx, col in enumerate(updated_headers)}
        wb.close()

        # ✅ Apply formatting and hyperlinks
        ExcelMerger._apply_formatting_and_hyperlinks(
            temp_file_path, hyperlinks, merged_df,
            status_column=status_column, status_value=status_value,
            project_column=project_column, project_value=project_value,
            custom_checks=custom_checks,
            col_name_to_excel_idx=col_name_to_excel_idx
        )

        # ✅ Apply rich text coloring on title_match
        ExcelMerger.apply_title_match_highlighting(temp_file_path, merged_df)

        # ✅ Apply title highlighting
        if title_column1 and title_column2:
            ExcelMerger.apply_title_highlighting(
                temp_file_path, merged_df,
                title_col1='title_excel1',
                title_col2='title_excel2',
                reorder=False,
                update_baseline=True,
                status_column=status_column, status_value=status_value,
                project_column=project_column, project_value=project_value,
                custom_checks=custom_checks,
                filename_column=filename_column
            )

        return temp_file_path, merged_df

    @staticmethod
    def reorder_columns(file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        # Extract current headers and data
        existing_headers = [cell.value for cell in ws[1]]
        data = list(ws.iter_rows(min_row=2, values_only=True))

        # Separate title-related and final special columns
        title_cols = [col for col in ["title_excel1", "title_excel2", "title_excel3"] if col in existing_headers]
        final_cols = [col for col in ["title_match", "Comments_1", "Case"] if col in existing_headers]

        # All columns that should be excluded from the general pool
        excluded = {"common_ref", "original_row_index"} | set(title_cols) | set(final_cols)

        # General non-title, non-final columns
        other_columns = [col for col in existing_headers if col not in excluded]

        # Build final order:
        desired_order = other_columns  # first: general columns
        if "common_ref" in existing_headers:
            desired_order.append("common_ref")  # second: place common_ref
        desired_order.extend(title_cols)  # third: title columns
        desired_order.extend(final_cols)  # fourth: final columns

        # Remove all current columns
        ws.delete_cols(1, ws.max_column)

        # Reinsert columns clearly in desired order
        for col_idx, header in enumerate(desired_order, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Re-map original data to the new column order
        header_to_idx = {h: existing_headers.index(h) for h in existing_headers}
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(desired_order, start=1):
                original_idx = header_to_idx.get(header)
                if original_idx is not None:
                    ws.cell(row=row_idx, column=col_idx, value=row_data[original_idx])
                else:
                    ws.cell(row=row_idx, column=col_idx, value=None)  # If header wasn't in original data

        wb.save(file_path)
        wb.close()
        print("✅ Column reordering completed and file saved.")

    @staticmethod
    def _extract_hyperlinks(file_path):
        """
        Extract hyperlinks from the original Excel file mapped by original_row_index and column headers.
        """
        print("Extracting hyperlinks from:", file_path)
        hyperlinks = {}
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_number = row[0].row
            hyperlinks[row_number] = {}
            for idx, cell in enumerate(row):
                if cell.hyperlink:
                    col_name = headers[idx]  # Use column header names instead of index
                    print(f"Hyperlink found at row {row_number}, column '{col_name}': {cell.hyperlink.target}")
                    hyperlinks[row_number][col_name] = cell.hyperlink.target

        print("Extracted Hyperlinks:", hyperlinks)
        return hyperlinks

    @staticmethod
    def add_original_row_index_to_dataframe(df, file_path):
        print(f"Adding original row indices from file: {file_path}")
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        original_row_indices = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_values = [cell.value for cell in row]
            if any(row_values):
                original_row_indices.append(row[0].row)
        if len(original_row_indices) != len(df):
            raise ValueError(
                f"Mismatch between extracted row indices ({len(original_row_indices)}) and DataFrame rows ({len(df)}). Ensure no extra blank rows in Excel."
            )
        df['original_row_index'] = original_row_indices
        print(f"Assigned original row indices: {original_row_indices}")
        return df

    @staticmethod
    def _save_merged_to_excel(df, output_path):
        """Ensure Comments_1 exists before saving the merged Excel file."""
        if "Comments_1" not in df.columns:
            df["Comments_1"] = ""  # ✅ Add default value if missing

        print("DEBUG: Final Columns Before Saving:", df.columns.tolist())  # ✅ Verify columns before saving

        temp_file_path = output_path if not os.path.isdir(output_path) else os.path.join(output_path,
                                                                                         'merged_result_temp.xlsx')
        df.to_excel(temp_file_path, index=False, header=True)

        print(f"✅ Saved merged file with Comments_1: {temp_file_path}")  # ✅ Confirm file is saved
        return temp_file_path

    @staticmethod
    def _apply_formatting_and_hyperlinks(file_path, hyperlinks, merged_df,
                                         status_column=None, status_value=None,
                                         project_column=None, project_value=None,
                                         custom_checks=None,
                                         col_name_to_excel_idx=None):
        if col_name_to_excel_idx is None:
            wb_temp = load_workbook(file_path)
            ws_temp = wb_temp.active
            updated_headers = [cell.value for cell in ws_temp[1]]
            col_name_to_excel_idx = {col: idx + 1 for idx, col in enumerate(updated_headers)}
            wb_temp.close()

        wb = load_workbook(file_path)
        ws = wb.active
        print("Applying formatting and hyperlinks to:", file_path)

        duplicate_font = Font(bold=True, color="FF3300")
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        header_row = [cell.value for cell in ws[1]]
        col_name_to_excel_idx = {col_name: idx + 1 for idx, col_name in enumerate(header_row)}

        common_ref_col_idx = col_name_to_excel_idx.get('common_ref')
        common_ref_duplicates = merged_df['common_ref'][merged_df['common_ref'].duplicated(keep=False)].tolist()

        if 'number_3' in merged_df.columns:
            refno1_col_idx = col_name_to_excel_idx.get('number_1')
            refno2_col_idx = col_name_to_excel_idx.get('number_2')
            refno3_col_idx = col_name_to_excel_idx.get('number_3')

            fill_styles = {
                (True, False, False): PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
                (False, True, False): PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid"),
                (False, False, True): PatternFill(start_color="AACCFF", end_color="AACCFF", fill_type="solid"),
                (True, True, False): PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"),
                (True, False, True): PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
                (False, True, True): PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
                (True, True, True): None,
                (False, False, False): None
            }

            refno1_duplicates = merged_df['number_1'][merged_df['number_1'].duplicated(keep=False)].tolist()
            refno2_duplicates = merged_df['number_2'][merged_df['number_2'].duplicated(keep=False)].tolist()
            refno3_duplicates = merged_df['number_3'][merged_df['number_3'].duplicated(keep=False)].tolist()

            for row_idx in range(2, len(merged_df) + 2):
                num1 = ws.cell(row=row_idx, column=refno1_col_idx).value
                num2 = ws.cell(row=row_idx, column=refno2_col_idx).value
                num3 = ws.cell(row=row_idx, column=refno3_col_idx).value
                presence = (bool(num1), bool(num2), bool(num3))
                fill = fill_styles.get(presence)

                if fill:
                    if num1:
                        ws.cell(row=row_idx, column=refno1_col_idx).fill = fill
                    if num2:
                        ws.cell(row=row_idx, column=refno2_col_idx).fill = fill
                    if num3:
                        ws.cell(row=row_idx, column=refno3_col_idx).fill = fill
                    if common_ref_col_idx:
                        ws.cell(row=row_idx, column=common_ref_col_idx).fill = fill

                if num1 in refno1_duplicates:
                    ws.cell(row=row_idx, column=refno1_col_idx).font = duplicate_font
                if num2 in refno2_duplicates:
                    ws.cell(row=row_idx, column=refno2_col_idx).font = duplicate_font
                if num3 in refno3_duplicates:
                    ws.cell(row=row_idx, column=refno3_col_idx).font = duplicate_font

                if common_ref_col_idx and ws.cell(row=row_idx, column=common_ref_col_idx).value in common_ref_duplicates:
                    ws.cell(row=row_idx, column=common_ref_col_idx).font = duplicate_font

            status_col_idx = col_name_to_excel_idx.get(status_column) if status_column else None
            project_col_idx = col_name_to_excel_idx.get(project_column) if project_column else None
            custom_col_indices = {
                col: col_name_to_excel_idx.get(col) for col, _ in custom_checks if col in col_name_to_excel_idx
            } if custom_checks else {}

            for i, row in merged_df.iterrows():
                excel_row = i + 2
                if status_col_idx and row[status_column] != status_value:
                    ws.cell(row=excel_row, column=status_col_idx).fill = light_red_fill
                if project_col_idx and row[project_column] != project_value:
                    ws.cell(row=excel_row, column=project_col_idx).fill = light_red_fill
                for custom_col, col_idx in custom_col_indices.items():
                    if row[custom_col] != dict(custom_checks).get(custom_col, ""):
                        ws.cell(row=excel_row, column=col_idx).fill = light_red_fill

            instance_mapping = {
                (True, False, False): "PDF is provided but not listed in LOD",
                (False, True, False): "LOD_2 only",
                (False, False, True): "LOD_3 only",
                (True, True, False): "PDF is provided and number_2",
                (True, False, True): "PDF is provided and number_3",
                (False, True, True): "No PDF but found in LOD_2 and LOD_3",
                (True, True, True): "",
                (False, False, False): "None"
            }
            instance_col_idx = ws.max_column + 1
            ws.cell(row=1, column=instance_col_idx, value="Case")
            for row_idx in range(2, len(merged_df) + 2):
                num1 = ws.cell(row=row_idx, column=refno1_col_idx).value
                num2 = ws.cell(row=row_idx, column=refno2_col_idx).value
                num3 = ws.cell(row=row_idx, column=refno3_col_idx).value
                presence = (bool(num1), bool(num2), bool(num3))
                ws.cell(row=row_idx, column=instance_col_idx, value=instance_mapping.get(presence, "Unknown"))

        else:
            refno1_col_idx = col_name_to_excel_idx.get('number_1')
            refno2_col_idx = col_name_to_excel_idx.get('number_2')

            fill_styles_2 = {
                (True, False): PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid"),
                (False, True): PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid"),
                (True, True): None,
                (False, False): None
            }

            refno1_duplicates = merged_df['number_1'][merged_df['number_1'].duplicated(keep=False)].tolist()
            refno2_duplicates = merged_df['number_2'][merged_df['number_2'].duplicated(keep=False)].tolist()

            for row_idx in range(2, len(merged_df) + 2):
                num1 = ws.cell(row=row_idx, column=refno1_col_idx).value
                num2 = ws.cell(row=row_idx, column=refno2_col_idx).value
                presence = (bool(num1), bool(num2))
                fill = fill_styles_2.get(presence)

                if fill:
                    if num1:
                        ws.cell(row=row_idx, column=refno1_col_idx).fill = fill
                    if num2:
                        ws.cell(row=row_idx, column=refno2_col_idx).fill = fill
                    if common_ref_col_idx:
                        ws.cell(row=row_idx, column=common_ref_col_idx).fill = fill

                if num1 in refno1_duplicates:
                    ws.cell(row=row_idx, column=refno1_col_idx).font = duplicate_font
                if num2 in refno2_duplicates:
                    ws.cell(row=row_idx, column=refno2_col_idx).font = duplicate_font
                if common_ref_col_idx and ws.cell(row=row_idx, column=common_ref_col_idx).value in common_ref_duplicates:
                    ws.cell(row=row_idx, column=common_ref_col_idx).font = duplicate_font

            instance_mapping_2 = {
                (True, False): "PDF is provided but not listed in LOD",
                (False, True): "LOD_2",
                (True, True): "",
                (False, False): "None"
            }
            instance_col_idx = ws.max_column + 1
            ws.cell(row=1, column=instance_col_idx, value="Instance")
            for row_idx in range(2, len(merged_df) + 2):
                num1 = ws.cell(row=row_idx, column=refno1_col_idx).value
                num2 = ws.cell(row=row_idx, column=refno2_col_idx).value
                presence = (bool(num1), bool(num2))
                ws.cell(row=row_idx, column=instance_col_idx, value=instance_mapping_2.get(presence, "Unknown"))

        for original_row_index, columns in hyperlinks.items():
            new_row = merged_df[merged_df['original_row_index'] == original_row_index]
            if new_row.empty:
                print(f"No matching row for original_row_index: {original_row_index}")
                continue
            new_row_idx = new_row.index[0] + 2
            for col_name, hyperlink in columns.items():
                excel_col_idx = col_name_to_excel_idx.get(col_name)
                if excel_col_idx:
                    try:
                        ws.cell(row=new_row_idx, column=excel_col_idx).hyperlink = hyperlink
                        ws.cell(row=new_row_idx, column=excel_col_idx).style = "Hyperlink"
                        print(f"✅ Applied hyperlink at row {new_row_idx}, column '{col_name}', link: {hyperlink}")
                    except Exception as e:
                        print(f"❌ Error applying hyperlink at row {new_row_idx}, column '{col_name}': {e}")
                else:
                    print(f"⚠️ Column '{col_name}' not found in Excel worksheet headers.")

        wb.save(file_path)
        wb.close()

    @staticmethod
    def build_colored_title_match(text):
        """Create a rich text object for title_match like 'True, False'."""
        rich = CellRichText()
        tokens = [t.strip() for t in text.split(",")]

        for i, token in enumerate(tokens):
            font = InlineFont(rFont="Calibri", sz=11)

            if token.lower() == "true":
                font.color = "008000"  # Green
            elif token.lower() == "false":
                font.color = "FF0000"  # Red
            else:
                font.color = "000000"  # Default black

            rich.append(TextBlock(font, token))  # ✅ proper text block

            # ✅ Add comma separator with default font (not string!)
            if i < len(tokens) - 1:
                sep_font = InlineFont(rFont="Calibri", sz=11, color="000000")
                rich.append(TextBlock(sep_font, ", "))

        return rich

    @staticmethod
    def apply_title_match_highlighting(file_path, merged_df):
        print("🎯 Applying rich-text to 'title_match'...")

        wb = load_workbook(file_path, rich_text=True)
        ws = wb.active

        col_idx = ExcelMerger.get_ws_column_index(ws, "title_match")
        if col_idx is None:
            print("❌ 'title_match' column not found.")
            wb.close()
            return

        for i, row in merged_df.iterrows():
            cell_value = str(row.get("title_match", ""))
            rich = ExcelMerger.build_colored_title_match(cell_value)
            print(type(rich), rich)
            ws.cell(row=i + 2, column=col_idx).value = rich

        wb.save(file_path)
        wb.close()
        print("✅ title_match rich-text coloring applied.")

    @staticmethod
    def tokenize_with_indices(text):
        tokens = []
        # Updated regex to separate underscores along with punctuation (hyphen, en dash, em dash)
        for match in re.finditer(r'[A-Za-z0-9]+|[_–—-]|[^\w\s]', text, re.UNICODE):
            tokens.append((match.group(), match.start()))
        return tokens

    @staticmethod
    def dp_align_tokens(tokens1, tokens2):
        token_strs1 = [t[0] for t in tokens1]
        token_strs2 = [t[0] for t in tokens2]
        n, m = len(token_strs1), len(token_strs2)
        dp = [[0] * (m + 1) for _ in range(n + 1)]
        backtrack = [[None] * (m + 1) for _ in range(n + 1)]
        for i in range(1, n + 1):
            dp[i][0] = i
            backtrack[i][0] = 'UP'
        for j in range(1, m + 1):
            dp[0][j] = j
            backtrack[0][j] = 'LEFT'
        for i in range(1, n + 1):
            for j in range(1, m + 1):
                similarity = SequenceMatcher(None, token_strs1[i - 1], token_strs2[j - 1]).ratio()
                if token_strs1[i - 1] == token_strs2[j - 1]:
                    replace_cost = dp[i - 1][j - 1]
                    flag = "EXACT"
                elif token_strs1[i - 1].lower() == token_strs2[j - 1].lower():
                    replace_cost = dp[i - 1][j - 1] + 0.5
                    flag = "CASE_ONLY"
                elif similarity >= 0.8:
                    replace_cost = dp[i - 1][j - 1] + (1 - similarity)
                    flag = "CHAR_LEVEL"
                elif similarity >= 0.4:
                    replace_cost = dp[i - 1][j - 1] + 2
                    flag = "CHAR_LEVEL"
                else:
                    replace_cost = float('inf')
                delete_cost = dp[i - 1][j] + 1
                insert_cost = dp[i][j - 1] + 1
                dp[i][j] = min(replace_cost, delete_cost, insert_cost)
                if dp[i][j] == replace_cost:
                    backtrack[i][j] = 'DIAG'
                elif dp[i][j] == delete_cost:
                    backtrack[i][j] = 'UP'
                else:
                    backtrack[i][j] = 'LEFT'
        aligned_tokens1, aligned_tokens2, flags = [], [], []
        i, j = n, m
        while i > 0 or j > 0:
            if i > 0 and j > 0 and backtrack[i][j] == 'DIAG':
                aligned_tokens1.append(tokens1[i - 1])
                aligned_tokens2.append(tokens2[j - 1])
                if token_strs1[i - 1] == token_strs2[j - 1]:
                    flags.append("EXACT")
                elif token_strs1[i - 1].lower() == token_strs2[j - 1].lower():
                    flags.append("CASE_ONLY")
                else:
                    flags.append("CHAR_LEVEL")
                i -= 1
                j -= 1
            elif i > 0 and (j == 0 or backtrack[i][j] == 'UP'):
                aligned_tokens1.append(tokens1[i - 1])
                aligned_tokens2.append((None, None))
                flags.append("MISSING_2")
                i -= 1
            elif j > 0 and (i == 0 or backtrack[i][j] == 'LEFT'):
                aligned_tokens1.append((None, None))
                aligned_tokens2.append(tokens2[j - 1])
                flags.append("MISSING_1")
                j -= 1
        return aligned_tokens1[::-1], aligned_tokens2[::-1], flags[::-1]

    @staticmethod
    def create_rich_text(original_text, aligned_tokens, aligned_tokens_other, flags):
        rich_text = CellRichText()
        last_index = 0

        for ((token, idx), (token_other, _), flag) in zip(aligned_tokens, aligned_tokens_other, flags):
            if token is None:
                continue

            if idx > last_index:
                rich_text.append(original_text[last_index:idx])

            if flag == "EXACT":
                inline_font = InlineFont(rFont="Calibri", sz=11)
                rich_text.append(TextBlock(inline_font, token))

            elif flag == "CASE_ONLY":
                inline_font = InlineFont(rFont="Calibri", sz=11, color="808080")
                rich_text.append(TextBlock(inline_font, token))

            elif flag == "CHAR_LEVEL":
                # Here, character-level highlighting
                for c1, c2 in zip(token, token_other or ""):
                    inline_font = InlineFont(rFont="Calibri", sz=11)
                    if c1 == c2:
                        inline_font.color = "000000"  # matching char (default black)
                    else:
                        inline_font.color = "FFA500"  # mismatched char (orange)
                    rich_text.append(TextBlock(inline_font, c1))

                # If token is longer than the compared token
                if len(token) > len(token_other or ""):
                    for c1 in token[len(token_other or ""):]:
                        inline_font = InlineFont(rFont="Calibri", sz=11, color="FFA500")
                        rich_text.append(TextBlock(inline_font, c1))

            elif flag in ["MISSING_1", "MISSING_2"]:
                inline_font = InlineFont(rFont="Calibri", sz=11, color="FF0000")
                rich_text.append(TextBlock(inline_font, token))

            last_index = idx + len(token)

        if last_index < len(original_text):
            rich_text.append(original_text[last_index:])

        return rich_text

    @staticmethod
    def get_ws_column_index(ws, header_name):
        header_name = header_name.strip().lower()
        print(f"Looking for header '{header_name}' in row 1:")
        for cell in ws[1]:
            if isinstance(cell.value, str):
                print(f"  Found header '{cell.value.strip().lower()}' at column {cell.column}")
                if cell.value.strip().lower() == header_name:
                    return cell.column  # openpyxl uses 1-based indexing
        return None

    @staticmethod
    def titles_match(title1, title2):
        """Check exact match between two titles after trimming and normalizing."""
        return str(title1).strip().lower() == str(title2).strip().lower()

    @staticmethod
    def update_comments_column(merged_df, status_column=None, status_value=None, project_column=None,
                               project_value=None, custom_checks=None, filename_column=None):
        """Ensures Comments_1 is correctly populated with mismatches before saving.
           Leaves cells empty if there are no issues. Uses '\n' for better readability."""

        # ✅ Ensure "Comments_1" exists
        if "Comments_1" not in merged_df.columns:
            merged_df["Comments_1"] = ""

        def append_comment(existing, new_comment):
            """Appends new_comment with '\n' for better readability."""
            if new_comment.strip():  # Only append if there's a real issue
                return f"{existing}\n{new_comment}".strip() if existing else new_comment.strip()
            return existing  # Keep existing value if there's no new comment

        # ✅ Perform Filename Check
        if filename_column and filename_column in merged_df.columns:
            for i, row in merged_df.iterrows():
                number_1 = str(row.get("number_1", "")).strip()
                filename = str(row.get(filename_column, "")).strip()
                comment = row.get("Comments_1", "")

                # Skip if any is missing
                if not number_1 or not filename:
                    continue

                # Check mismatch
                if not filename.startswith(number_1):
                    mismatch_msg = f"Filename & Drawing Number Mismatch: {filename} vs {number_1}"
                    merged_df.at[i, "Comments_1"] = append_comment(comment, mismatch_msg)

        # ✅ Perform Status Check
        if status_column and status_column in merged_df.columns:
            merged_df["Comments_1"] = merged_df.apply(
                lambda row: append_comment(row["Comments_1"],
                                           f"{status_column} Mismatch: {row[status_column]} <--> {status_value}")
                if pd.notna(row["number_1"]) and  # ✅ Skip rows where number_1 is empty
                   (pd.isna(row[status_column]) or  # ✅ Flags None/NaN as Mismatch
                    (pd.notna(row[status_column]) and pd.notna(status_value) and
                     str(row[status_column]).strip().lower() != str(status_value).strip().lower()))
                else row["Comments_1"],
                axis=1
            )

        # ✅ Perform Project Name Check
        if project_column and project_column in merged_df.columns:
            merged_df["Comments_1"] = merged_df.apply(
                lambda row: append_comment(row["Comments_1"],
                                           f"{project_column} Mismatch: {row[project_column]} <--> {project_value}")
                if pd.notna(row["number_1"]) and  # ✅ Skip rows where number_1 is empty
                   (pd.isna(row[project_column]) or  # ✅ Flags None/NaN as Mismatch
                    (pd.notna(row[project_column]) and pd.notna(project_value) and
                     str(row[project_column]).strip().lower() != str(project_value).strip().lower()))
                else row["Comments_1"],
                axis=1
            )

        # ✅ Perform Custom Checks
        if custom_checks:
            print(f"🧪 Detected custom_checks: {custom_checks}")
            print(f"🧾 DataFrame columns: {merged_df.columns.tolist()}")

            for custom_col, custom_value in custom_checks:
                print(f"🔍 Checking Custom: {custom_col=} {custom_value=}")
                if custom_col not in merged_df.columns:
                    print(f"❌ Column '{custom_col}' not found in merged_df. Skipping.")
                    continue

                def check_and_comment(row):
                    value = row.get(custom_col, "")
                    number_1 = row.get("number_1", "")
                    if pd.notna(number_1):
                        mismatch = (
                                pd.isna(value) or
                                (pd.notna(value) and pd.notna(custom_value) and
                                 str(value).strip().lower() != str(custom_value).strip().lower())
                        )
                        if mismatch:
                            print(f"❗ Row {row.name + 2}: {custom_col} mismatch -> {value} ≠ {custom_value}")
                            return append_comment(
                                row["Comments_1"],
                                f"{custom_col} Mismatch: {value} <--> {custom_value}"
                            )
                    return row["Comments_1"]

                merged_df["Comments_1"] = merged_df.apply(check_and_comment, axis=1)

        return merged_df



    @staticmethod
    def apply_title_highlighting(file_path, merged_df, title_col1, title_col2, reorder=True, update_baseline=True,
                                 status_column=None, status_value=None,
                                 project_column=None, project_value=None,
                                 custom_checks=None, filename_column=None):

        """Applies title highlighting and ensures 'Comments_1' updates persist in the final Excel file."""
        wb = load_workbook(file_path, rich_text=True)
        ws = wb.active

        # Get current column indices
        col_idx1 = ExcelMerger.get_ws_column_index(ws, title_col1.strip().lower())
        col_idx2 = ExcelMerger.get_ws_column_index(ws, title_col2.strip().lower())

        comments_col_idx = ExcelMerger.get_ws_column_index(ws, "Comments_1")  # Ensure column exists

        missing_cols = []
        if col_idx1 is None:
            missing_cols.append(title_col1)
        if col_idx2 is None:
            missing_cols.append(title_col2)
        if comments_col_idx is None:
            missing_cols.append("Comments_1")

        if missing_cols:
            print(f"⚠️ Warning: Columns not found: {missing_cols}. Highlighting aborted.")
            wb.close()
            return

        if col_idx1 is None or col_idx2 is None:
            print(f"Could not find header {title_col1} or {title_col2} in the worksheet.")
            wb.close()
            return


        print(f"Applying title highlighting on columns '{title_col1}' and '{title_col2}':")
        for i, row in merged_df.iterrows():
            excel_row = i + 2  # account for header row
            baseline_text = str(row.get(title_col1, ""))
            other_text = str(row.get(title_col2, ""))
            tokens1 = ExcelMerger.tokenize_with_indices(baseline_text)
            tokens2 = ExcelMerger.tokenize_with_indices(other_text)
            aligned_tokens1, aligned_tokens2, flags = ExcelMerger.dp_align_tokens(tokens1, tokens2)
            rich_text1 = ExcelMerger.create_rich_text(baseline_text, aligned_tokens1, aligned_tokens2, flags)
            rich_text2 = ExcelMerger.create_rich_text(other_text, aligned_tokens2, aligned_tokens1, flags)

            # ✅ Ensure Comments_1 updates persist in the Excel file
            if comments_col_idx:
                ws.cell(row=excel_row, column=comments_col_idx).value = row["Comments_1"]

            # Update baseline column only if update_baseline is True.
            if update_baseline:
                ws.cell(row=excel_row, column=col_idx1).value = rich_text1
            # Always update the second column.
            ws.cell(row=excel_row, column=col_idx2).value = rich_text2

        # ✅ Debug Before Saving
        print("✅ DEBUG: Final 'Comments_1' Column Before Saving:")
        print(merged_df["Comments_1"].value_counts(dropna=False))

        wb.save(file_path)
        wb.close()
        print("Title highlighting applied and workbook saved:", file_path)


class CTkDnD(ctk.CTk, TkinterDnD.DnDWrapper):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.TkdndVersion = TkinterDnD._require(self)


class MergerGUI:
    """Handles the GUI for the Excel merger with drag-and-drop file selection and theme toggle."""

    def __init__(self, master=None):

        #set_custom_theme("dark")  # or "light" if you prefer the light theme

        ctk.set_appearance_mode("dark")  # Set dark mode at startup


        # Use our custom CTkDnD main window for drag-and-drop support.
        self.mergerApp = CTkDnD() if master is None else ctk.CTkToplevel(master)
        self.mergerApp.title("Conflux")
        self.mergerApp.iconbitmap(resource_path("style/xtractor-logo.ico"))

        # File paths for three Excel files and the output file.
        self.excel1_path = tk.StringVar()
        self.excel2_path = tk.StringVar()
        self.excel3_path = tk.StringVar()
        self.output_path = tk.StringVar()

        # Header selections for each file (via drop-down lists).
        self.ref_column1 = tk.StringVar()
        self.title_column1 = tk.StringVar()
        self.ref_column2 = tk.StringVar()
        self.title_column2 = tk.StringVar()
        self.ref_column3 = tk.StringVar()
        self.title_column3 = tk.StringVar()

        # Boolean variables for report options and comparing Excel3 title.
        self.compare_excel2_title = tk.BooleanVar(value=False)
        self.generate_word_report = tk.BooleanVar(value=False)
        self.compare_excel3_title = tk.BooleanVar(value=False)

        # Boolean variable for theme mode; True = dark mode.
        self.theme_mode = tk.BooleanVar(value=True)

        self.excel1_headers = []
        self.excel2_headers = []
        self.excel3_headers = []

        self._build_gui()

    def _build_gui(self):
        self.mergerApp.grid_rowconfigure(0, weight=1)
        self.mergerApp.grid_columnconfigure(0, weight=1)
        self.mergerApp.grid_columnconfigure(1, weight=1)
        self.mergerApp.grid_columnconfigure(2, weight=1)

        # Create three frames for Excel 1, 2, and 3.
        self.excel1_frame = ctk.CTkFrame(self.mergerApp)
        self.excel1_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.excel2_frame = ctk.CTkFrame(self.mergerApp)
        self.excel2_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        self.excel3_frame = ctk.CTkFrame(self.mergerApp)
        self.excel3_frame.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")

        for frame in (self.excel1_frame, self.excel2_frame, self.excel3_frame):
            frame.grid_rowconfigure(0, weight=0)
            frame.grid_columnconfigure(0, weight=0)
            frame.grid_columnconfigure(1, weight=1)

        self._build_excel1_section(self.excel1_frame)
        self._build_excel2_section(self.excel2_frame)
        self._build_excel3_section(self.excel3_frame)


        # Create the comparison check frame
        self.comparison_frame = ctk.CTkFrame(self.mergerApp)
        self.comparison_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=(10, 10), sticky="nsew")
        self._build_comparison_checks(self.comparison_frame)

        # Build Filename Checker frame to the right
        self.filename_frame = ctk.CTkFrame(self.mergerApp)
        self.filename_frame.grid(row=1, column=2, padx=10, pady=(10, 10), sticky="nsew")
        self._build_filename_checker(self.filename_frame)

        # Create controls frame (move down)
        self.controls_frame = ctk.CTkFrame(self.mergerApp)
        self.controls_frame.grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
        self._build_controls(self.controls_frame)

    def _build_excel1_section(self, parent_frame):
        font_name = "Helvetica"
        font_size = 12
        self.excel1_button = ctk.CTkButton(parent_frame,
            text="\n➕\n\nSelect Extracted Excel or\nDrag & Drop Here",
            command=self._browse_excel1,
            border_width=3,
            fg_color="transparent",
            hover_color=("#D6D6D6", "#505050"),  # Light and dark hover color
            text_color=("#333333", "#FFFFFF"),
            corner_radius=10,
            width=200,
            height=150)
        self.excel1_button.grid(row=0, column=0, columnspan=2, padx=33, pady=33, sticky="ew")
        # Enable drag and drop on Excel1 button.
        self.excel1_button.drop_target_register(DND_ALL)
        self.excel1_button.dnd_bind('<<Drop>>', self.drop_excel1)
        ctk.CTkLabel(parent_frame, text="Reference Column:", font=(font_name, font_size)).grid(
            row=1, column=0, padx=5, pady=2, sticky="e")
        self.ref_option_menu1 = ctk.CTkOptionMenu(parent_frame, variable=self.ref_column1, values=[])
        self.ref_option_menu1.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        ctk.CTkLabel(parent_frame, text="", font=(font_name, font_size)).grid(
            row=2, column=0, padx=5, pady=2, sticky="e")
        ctk.CTkLabel(parent_frame, text="Drawing Title:", font=(font_name, font_size)).grid(
            row=3, column=0, padx=5, pady=2, sticky="e")
        self.title_option_menu1 = ctk.CTkOptionMenu(parent_frame, variable=self.title_column1, values=[], state="disabled")
        self.title_option_menu1.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

    def _build_excel2_section(self, parent_frame):
        font_name = "Helvetica"
        font_size = 12
        self.excel2_button = ctk.CTkButton(parent_frame,
            text="\n➕\n\nSelect DC_LOD Excel or\nDrag & Drop Here",
            command=self._browse_excel2,
            border_width=3,
            fg_color="transparent",
            hover_color=("#D6D6D6", "#505050"),  # Light and dark hover color
            text_color=("#333333", "#FFFFFF"),
            corner_radius=10,
            width=200,
            height=150)
        self.excel2_button.grid(row=0, column=0, columnspan=2, padx=33, pady=33, sticky="ew")
        self.excel2_button.drop_target_register(DND_ALL)
        self.excel2_button.dnd_bind('<<Drop>>', self.drop_excel2)
        ctk.CTkLabel(parent_frame, text="Reference Column:", font=(font_name, font_size)).grid(
            row=1, column=0, padx=5, pady=2, sticky="e")
        self.ref_option_menu2 = ctk.CTkOptionMenu(parent_frame, variable=self.ref_column2, values=[])
        self.ref_option_menu2.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        ctk.CTkCheckBox(parent_frame, text="Compare Title 2",
                        variable=self.compare_excel2_title, command=self._toggle_title_entries).grid(
            row=2, column=0, columnspan=2, padx=5, pady=2, sticky="w")

        ctk.CTkLabel(parent_frame, text="Drawing Title:", font=(font_name, font_size)).grid(
            row=3, column=0, padx=5, pady=2, sticky="e")
        self.title_option_menu2 = ctk.CTkOptionMenu(parent_frame, variable=self.title_column2, values=[], state="disabled")
        self.title_option_menu2.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

    def _build_excel3_section(self, parent_frame):
        font_name = "Helvetica"
        font_size = 12
        self.excel3_button = ctk.CTkButton(parent_frame,
            text="\n➕\n\nSelect DD_LOD Excel or\nDrag & Drop Here",
            command=self._browse_excel3,
            border_width=3,
            fg_color="transparent",
            hover_color=("#D6D6D6", "#505050"),  # Light and dark hover color
            text_color=("#333333", "#FFFFFF"),
            corner_radius=10,
            width=200,
            height=150)
        self.excel3_button.grid(row=0, column=0, columnspan=2, padx=33, pady=33, sticky="ew")
        self.excel3_button.drop_target_register(DND_ALL)
        self.excel3_button.dnd_bind('<<Drop>>', self.drop_excel3)
        ctk.CTkLabel(parent_frame, text="Reference Column:", font=(font_name, font_size)).grid(
            row=1, column=0, padx=5, pady=2, sticky="e")
        self.ref_option_menu3 = ctk.CTkOptionMenu(parent_frame, variable=self.ref_column3, values=[])
        self.ref_option_menu3.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        # Place the compare title checkbox above the title dropdown.
        ctk.CTkCheckBox(parent_frame, text="Compare Title 3",
                        variable=self.compare_excel3_title, command=self._toggle_title_entries).grid(
            row=2, column=0, columnspan=2, padx=5, pady=2, sticky="w")
        ctk.CTkLabel(parent_frame, text="Drawing Title:", font=(font_name, font_size)).grid(
            row=3, column=0, padx=5, pady=2, sticky="e")
        self.title_option_menu3 = ctk.CTkOptionMenu(parent_frame, variable=self.title_column3, values=[], state="disabled")
        self.title_option_menu3.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

    def _build_comparison_checks(self, parent_frame):
        """Builds the checkboxes, dropdowns, and textboxes for additional validation."""
        # Add column labels
        ctk.CTkLabel(parent_frame, text="Enable", font=("Helvetica", 12, "bold")).grid(row=0, column=0, padx=5, pady=2,
                                                                                       sticky="w")
        ctk.CTkLabel(parent_frame, text="Column Name", font=("Helvetica", 12, "bold")).grid(row=0, column=1, padx=5,
                                                                                            pady=2, sticky="ew")
        ctk.CTkLabel(parent_frame, text="Expected Value", font=("Helvetica", 12, "bold")).grid(row=0, column=2, padx=5,
                                                                                               pady=2, sticky="ew")

        # Status Check (Moved to row=1)
        self.status_enabled = tk.BooleanVar(value=False)
        self.status_column = tk.StringVar()
        self.status_column.trace_add("write",
                                     lambda *a: self._update_preview_combo(self.status_column, self.status_combo))

        self.status_value = tk.StringVar()

        self.status_check = ctk.CTkCheckBox(
            parent_frame, text="Check 1",
            variable=self.status_enabled, command=self._toggle_status
        )
        self.status_check.grid(row=1, column=0, padx=5, pady=2, sticky="w")

        self.status_dropdown = ctk.CTkOptionMenu(parent_frame, variable=self.status_column, values=[], state="disabled")
        self.status_dropdown.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        self.status_combo = ctk.CTkComboBox(
            parent_frame,
            variable=self.status_value,
            values=[],
            state="disabled",
            justify="left"
        )
        self.status_combo.grid(row=1, column=2, padx=5, pady=2, sticky="ew")

        # Project Name Check (Moved to row=2)
        self.project_enabled = tk.BooleanVar(value=False)
        self.project_column = tk.StringVar()
        self.project_column.trace_add("write",
                                      lambda *a: self._update_preview_combo(self.project_column, self.project_combo))

        self.project_value = tk.StringVar()

        self.project_check = ctk.CTkCheckBox(
            parent_frame, text="Check 2",
            variable=self.project_enabled, command=self._toggle_project
        )
        self.project_check.grid(row=2, column=0, padx=5, pady=2, sticky="w")

        self.project_dropdown = ctk.CTkOptionMenu(parent_frame, variable=self.project_column, values=[],
                                                  state="disabled")
        self.project_dropdown.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        self.project_combo = ctk.CTkComboBox(
            parent_frame,
            variable=self.project_value,
            values=[],
            state="disabled",
            justify="left"
        )
        self.project_combo.grid(row=2, column=2, padx=5, pady=2, sticky="ew")

        # Add Custom Checks Button (Moved to row=3)
        self.custom_checks = []
        self.add_check_button = ctk.CTkButton(parent_frame, text="+ Add Check", command=self._add_custom_check)
        self.add_check_button.grid(row=3, column=0, columnspan=3, padx=5, pady=2, sticky="ew")

    def _build_filename_checker(self, parent_frame):
        """Builds the UI for filename comparison against reference column."""

        self.filename_enabled = tk.BooleanVar(value=False)
        self.filename_column = tk.StringVar()

        ctk.CTkLabel(parent_frame, text="Filename vs Reference Number", font=("Helvetica", 12, "bold")).grid(
            row=0, column=0, columnspan=2, padx=5, pady=(5, 2), sticky="w")

        self.filename_check = ctk.CTkCheckBox(
            parent_frame, text="Enable Filename Check",
            variable=self.filename_enabled,
            command=self._toggle_filename_check
        )
        self.filename_check.grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky="w")

        ctk.CTkLabel(parent_frame, text="Filename Column:", font=("Helvetica", 11)).grid(
            row=2, column=0, padx=5, pady=2, sticky="e")

        self.filename_dropdown = ctk.CTkOptionMenu(parent_frame, variable=self.filename_column, values=[],
                                                   state="disabled")
        self.filename_dropdown.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

    def _build_controls(self, parent_frame):
        font_name = "Helvetica"
        font_size = 12
        ctk.CTkLabel(parent_frame, text="Output Path:", font=(font_name, font_size)).grid(
            row=0, column=0, padx=5, pady=2, sticky="e")
        ctk.CTkEntry(parent_frame, textvariable=self.output_path, width=300).grid(
            row=0, column=1, padx=5, pady=2, sticky="ew")
        ctk.CTkButton(parent_frame, text="Use Excel 1 Path", command=self._use_excel1_path).grid(
            row=0, column=2, padx=5, pady=2)
        # Theme toggle switch (no label) placed at the bottom-right of the controls frame.
        self.theme_switch = ctk.CTkSwitch(parent_frame, text="", variable=self.theme_mode,
                                          command=self.toggle_theme, switch_width=20, switch_height=10)
        # Use place to anchor it at the bottom-right corner of the parent frame.
        self.theme_switch.place(relx=1.0, rely=1.0, anchor="se")
        ctk.CTkButton(parent_frame, text="Start Merge", command=self._start_merge).grid(
            row=2, column=0, columnspan=3, pady=10)
        parent_frame.grid_columnconfigure(1, weight=1)

    def _update_preview_combo(self, column_var, combo_widget):
        col_name = column_var.get()
        values = self.preview_values_by_column.get(col_name, [])
        combo_widget.configure(values=values)

    def toggle_theme(self):
        if self.theme_mode.get():
            ctk.set_appearance_mode("dark")
            print("Theme set to dark mode")
        else:
            ctk.set_appearance_mode("light")
            print("Theme set to light mode")

    def _toggle_status(self):
        """Enable or disable status comparison fields and update custom checks."""
        state = "normal" if self.status_enabled.get() else "disabled"
        self.status_dropdown.configure(state=state)
        self.status_combo.configure(state=state)
        if state == "normal" and self.status_column.get():
            col = self.status_column.get()
            self.status_combo.configure(values=self.preview_values_by_column.get(col, []))

        # ✅ Update all custom checks
        self._toggle_custom_checks()

    def _toggle_project(self):
        """Enable or disable project name comparison fields and update custom checks."""
        state = "normal" if self.project_enabled.get() else "disabled"
        self.project_dropdown.configure(state=state)
        self.project_combo.configure(state=state)
        if state == "normal" and self.project_column.get():
            col = self.project_column.get()
            self.project_combo.configure(values=self.preview_values_by_column.get(col, []))

        # ✅ Update all custom checks
        self._toggle_custom_checks()

    def _toggle_custom_checks(self):
        """Enable or disable custom check dropdowns and entry fields dynamically."""
        for enabled_var, column_var, value_var, dropdown_widget, entry_widget in self.custom_checks:
            state = "normal" if enabled_var.get() else "disabled"
            dropdown_widget.configure(state=state)
            entry_widget.configure(state=state)

    def _toggle_title_entries(self):
        # Enable title_option_menu1 if either checkbox is checked
        state_1 = "normal" if (self.compare_excel2_title.get() or self.compare_excel3_title.get()) else "disabled"
        self.title_option_menu1.configure(state=state_1)

        # Enable title_option_menu2 only if compare_excel2 is checked
        state_2 = "normal" if self.compare_excel2_title.get() else "disabled"
        self.title_option_menu2.configure(state=state_2)

        # Enable title_option_menu3 only if compare_excel3_title is checked
        state_3 = "normal" if self.compare_excel3_title.get() else "disabled"
        self.title_option_menu3.configure(state=state_3)

    def _toggle_filename_check(self):
        state = "normal" if self.filename_enabled.get() else "disabled"
        self.filename_dropdown.configure(state=state)

    def _add_custom_check(self):
        """Adds a new custom check row dynamically and ensures dropdown values are populated."""
        row_idx = len(self.custom_checks) + 3  # Start after Status and Project Name

        enabled_var = tk.BooleanVar(value=False)
        column_var = tk.StringVar()
        value_var = tk.StringVar()

        def update_this_combo(*_):
            col_name = column_var.get()
            values = self.preview_values_by_column.get(col_name, [])
            combo.configure(values=values)

        column_var.trace_add("write", update_this_combo)

        check = ctk.CTkCheckBox(
            self.comparison_frame, text=f"Check {row_idx}",
            variable=enabled_var
        )
        check.grid(row=row_idx, column=0, padx=5, pady=2, sticky="w")

        dropdown = ctk.CTkOptionMenu(self.comparison_frame, variable=column_var, values=[], state="disabled")
        dropdown.grid(row=row_idx, column=1, padx=5, pady=2, sticky="ew")

        combo = ctk.CTkComboBox(
            self.comparison_frame,
            variable=value_var,
            values=[],
            state="disabled",
            justify="left"
        )
        combo.grid(row=row_idx, column=2, padx=5, pady=2, sticky="ew")

        # ✅ Store all required elements (including dropdown widget) to update later
        self.custom_checks.append((enabled_var, column_var, value_var, dropdown, combo))

        # ✅ If Excel is already loaded, populate dropdown values immediately
        if self.excel1_headers:
            dropdown.configure(values=self.excel1_headers)
            column_var.set(auto_select_header(self.excel1_headers, ["status", "project"]))

        # ✅ Ensure new checks enable/disable properly when toggled
        enabled_var.trace_add("write", lambda *args: self._toggle_custom_checks())

        # ✅ Move the + button down
        self.add_check_button.grid(row=row_idx + 1, column=0, columnspan=3, padx=5, pady=2, sticky="ew")

    # --- Drag and Drop Handlers ---
    def drop_excel1(self, event):
        file_path = event.data.strip().replace("{", "").replace("}", "")
        if file_path.lower().endswith((".xlsx", ".xlsm")):
            self.excel1_path.set(file_path)
            self._load_excel1_headers(file_path)
            filename = os.path.basename(file_path)
            self.excel1_button.configure(text=filename, fg_color="#990d10")
        else:
            messagebox.showerror("Error", "Please drag and drop a valid Excel file (.xlsx or .xlsm).")

    def drop_excel2(self, event):
        file_path = event.data.strip().replace("{", "").replace("}", "")
        if file_path.lower().endswith((".xlsx", ".xlsm")):
            self.excel2_path.set(file_path)
            self._load_excel2_headers(file_path)
            filename = os.path.basename(file_path)
            self.excel2_button.configure(text=filename, fg_color="#990d10")
        else:
            messagebox.showerror("Error", "Please drag and drop a valid Excel file (.xlsx or .xlsm).")

    def drop_excel3(self, event):
        file_path = event.data.strip().replace("{", "").replace("}", "")
        if file_path.lower().endswith((".xlsx", ".xlsm")):
            self.excel3_path.set(file_path)
            self._load_excel3_headers(file_path)
            filename = os.path.basename(file_path)
            self.excel3_button.configure(text=filename, fg_color="#990d10")
        else:
            messagebox.showerror("Error", "Please drag and drop a valid Excel file (.xlsx or .xlsm).")

    def _browse_excel1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")],
                                               title="Select Excel File 1")
        if file_path:
            self.excel1_path.set(file_path)
            self._load_excel1_headers(file_path)
            import os
            filename = os.path.basename(file_path)
            self.excel1_button.configure(text=filename, fg_color="#217346")

    def _browse_excel2(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")],
                                               title="Select Excel File 2")
        if file_path:
            self.excel2_path.set(file_path)
            self._load_excel2_headers(file_path)
            import os
            filename = os.path.basename(file_path)
            self.excel2_button.configure(text=filename, fg_color="#217346")

    def _browse_excel3(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")],
                                               title="Select Excel File 3")
        if file_path:
            self.excel3_path.set(file_path)
            self._load_excel3_headers(file_path)
            import os
            filename = os.path.basename(file_path)
            self.excel3_button.configure(text=filename, fg_color="#217346")

    def _load_excel1_headers(self, file_path):
        try:
            df_preview = pd.read_excel(file_path, engine='openpyxl', nrows=6)
            df = df_preview.head(0)  # only headers

            headers = list(df.columns)

            # Populate existing dropdowns
            self.excel1_headers = headers
            self.preview_values_by_column = {
                header: list(dict.fromkeys(df_preview[header].dropna().astype(str).tolist()))
                for header in headers
            }

            self.ref_option_menu1.configure(values=headers)
            self.title_option_menu1.configure(values=headers)
            self.ref_column1.set(auto_select_header(headers, ["drawing", "sheet", "ref", "number"]))
            self.title_column1.set(auto_select_header(headers, ["title"]))

            # Populate Status and Project dropdowns
            self.status_dropdown.configure(values=headers)
            self.status_column.set(auto_select_header(headers, ["status"]))

            self.project_dropdown.configure(values=headers)
            self.project_column.set(auto_select_header(headers, ["project"]))

            # Populate filename dropdown and auto-select
            self.filename_dropdown.configure(values=headers)
            self.filename_column.set(auto_select_header(headers, ["filename", "file name"]))

            # ✅ Ensure correct widgets are updated for custom checks
            for check in self.custom_checks:
                enabled_var, column_var, value_var, dropdown_widget, _ = check
                dropdown_widget.configure(values=headers)  # Populate dropdown options
                column_var.set(auto_select_header(headers, ["status", "project"]))  # Auto-select if applicable

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load headers from Excel File 1: {e}")

    def _load_excel2_headers(self, file_path):
        try:
            df = pd.read_excel(file_path, engine='openpyxl', nrows=0)
            headers = list(df.columns)
            self.excel2_headers = headers
            self.ref_option_menu2.configure(values=headers)
            self.title_option_menu2.configure(values=headers)
            self.ref_column2.set(auto_select_header(headers, ["drawing", "sheet", "ref", "number"]))
            self.title_column2.set(auto_select_header(headers, ["title"]))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load headers from Excel File 2: {e}")

    def _load_excel3_headers(self, file_path):
        try:
            df = pd.read_excel(file_path, engine='openpyxl', nrows=0)
            headers = list(df.columns)
            self.excel3_headers = headers
            self.ref_option_menu3.configure(values=headers)
            self.title_option_menu3.configure(values=headers)
            self.ref_column3.set(auto_select_header(headers, ["drawing", "sheet", "ref", "number"]))
            self.title_column3.set(auto_select_header(headers, ["title"]))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load headers from Excel File 3: {e}")

    def _use_excel1_path(self):
        excel1 = self.excel1_path.get()
        if excel1:
            directory, file_name = os.path.split(excel1)
            name, ext = os.path.splitext(file_name)
            self.output_path.set(os.path.join(directory, f"{name}_merged{ext}"))

    def _start_merge(self):
        excel1_path = self.excel1_path.get()
        excel2_path = self.excel2_path.get()
        excel3_path = self.excel3_path.get().strip()
        ref_column1 = self.ref_column1.get()
        ref_column2 = self.ref_column2.get()
        ref_column3 = self.ref_column3.get() if excel3_path else None

        output_path = self.output_path.get().strip()
        if not output_path:
            messagebox.showerror("Error", "Please provide a valid output path.")
            return

        if os.path.exists(output_path):
            directory, file_name = os.path.split(output_path)
            base, ext = os.path.splitext(file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(directory, f"{base}_{timestamp}{ext}")

        try:
            title_col1 = self.title_column1.get() if (self.compare_excel2_title.get() or self.compare_excel3_title.get()) else None
            title_col2 = self.title_column2.get() if self.compare_excel2_title.get() else None
            title_col3 = self.title_column3.get() if self.compare_excel3_title.get() and excel3_path else None

            # Filename check
            filename_col = self.filename_column.get() if self.filename_enabled.get() else None

            # Collect status & project name checks
            status_col = self.status_column.get() if self.status_enabled.get() else None
            status_value = self.status_value.get() if self.status_enabled.get() else None
            project_col = self.project_column.get() if self.project_enabled.get() else None
            project_value = self.project_value.get() if self.project_enabled.get() else None

            # Collect custom checks
            custom_checks = [
                (column_var.get(), value_var.get()) for enabled_var, column_var, value_var, _, _ in self.custom_checks
                if enabled_var.get() and column_var.get()
            ]

            # Call `ExcelMerger`
            if excel3_path:
                merged_file_path, merged_df = ExcelMerger.merge_3_excels(
                    excel1_path, excel2_path, excel3_path,
                    ref_column1, ref_column2, ref_column3,
                    output_path,
                    title_column1=title_col1, title_column2=title_col2, title_column3=title_col3,
                    compare_excel2=self.compare_excel2_title.get(),
                    compare_excel3=self.compare_excel3_title.get(),
                    status_column=status_col,
                    status_value=status_value,
                    project_column=project_col,
                    project_value=project_value,
                    custom_checks=custom_checks,
                    filename_column=filename_col
                )
            else:
                merged_file_path, merged_df = ExcelMerger.merge_excels(
                    excel1_path, excel2_path,
                    ref_column1, ref_column2,
                    output_path,
                    title_column1=title_col1,
                    title_column2=title_col2,
                    status_column=status_col,
                    status_value=status_value,
                    project_column=project_col,
                    project_value=project_value,
                    custom_checks=custom_checks,
                    filename_column=filename_col
                )

            messagebox.showinfo("Success", f"Merged file saved at {merged_file_path}")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def run(self):
        if isinstance(self.mergerApp, ctk.CTk):
            self.mergerApp.mainloop()



if __name__ == "__main__":
    app = MergerGUI()
    app.run()



