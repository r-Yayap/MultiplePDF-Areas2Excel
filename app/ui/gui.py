# gui.py
import os
import re
import time
import customtkinter as ctk
import sys
from tkinter import filedialog, messagebox, StringVar
import re
from openpyxl import load_workbook
from app.ui.constants import *
from customtkinter import CTkImage
from typing import Any
from ttkwidgets import CheckboxTreeview
from tkinter import ttk
from PIL import Image  # Make sure this is at the top

from app.ui.pdf_viewer import PDFViewer
from app.domain.models import OcrSettings, ExtractionRequest
from app.controllers.extract_controller import ExtractController
from app.infra.pdf_adapter import PdfAdapter
from pathlib import Path

from app.ui.ui_utils import create_tooltip, EditableTreeview, CTkOptionMenuNoArrow
from app.common.ocr import find_tessdata
from app.domain.revision_patterns import REVISION_PATTERNS
from app.services.revision_parser import RevisionParser


from app.logging_setup import configure_logging
logger = configure_logging()


def _is_spec(a: Any) -> bool:
    try:
        from app.domain.models import AreaSpec
        return isinstance(a, AreaSpec)
    except Exception:
        return False

def _area_title(a: Any) -> str:
    return a.title if _is_spec(a) else a["title"]

def _area_coords(a: Any):
    return list(a.rect) if _is_spec(a) else a["coordinates"]

def _to_spec(a: Any):
    from app.domain.models import AreaSpec
    return a if _is_spec(a) else AreaSpec(title=a["title"], rect=tuple(a["coordinates"]))

def _maybe_to_dict(a: Any) -> dict:
    # convenient when exporting, etc.
    return {"title": a.title, "coordinates": list(a.rect)} if _is_spec(a) else a

def _rev_area_to_spec(a: Any):
    if a is None:
        return None
    return _to_spec(a)

def _area_get_title_and_coords(area):
    """Return (title, (x0,y0,x1,y1)) from either a GUI dict or an AreaSpec"""
    try:
        # AreaSpec path
        from app.domain.models import AreaSpec  # local import to avoid cycles
        if isinstance(area, AreaSpec):
            t = area.title or "Area"
            x0, y0, x1, y1 = area.rect
            return t, (x0, y0, x1, y1)
    except Exception:
        pass

    # dict path
    if isinstance(area, dict):
        t = area.get("title", "Area")
        coords = area.get("coordinates") or area.get("rect") or area.get("bbox")
        if coords and len(coords) == 4:
            x0, y0, x1, y1 = coords
            return t, (x0, y0, x1, y1)

    raise TypeError(f"Unsupported area object: {area!r}")

def _as_gui_area(area):
    """Always return a GUI-style dict {'title':..., 'coordinates':[...]}"""
    t, (x0, y0, x1, y1) = _area_get_title_and_coords(area)
    return {"title": t, "coordinates": [x0, y0, x1, y1]}

# DnD (safe import)
try:
    from tkinterdnd2 import TkinterDnD, DND_ALL
    DND_ENABLED = True
except Exception as e:
    print("tkdnd not available, drag & drop disabled:", e)
    TkinterDnD = None
    DND_ALL = None
    DND_ENABLED = False

class CTkDnD(ctk.CTk, *( (TkinterDnD.DnDWrapper,) if DND_ENABLED else tuple() )):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if DND_ENABLED:
            self.TkdndVersion = TkinterDnD._require(self)

def resource_path(rel: str) -> str:
    """
    Return an absolute path to a bundled resource that works for:
    • normal `python main.py` runs
    • Nuitka --standalone builds
    • PyInstaller one-file / one-dir builds
    """
    # -------- if we are inside a frozen app ---------------------------
    if getattr(sys, "frozen", False):
        # PyInstaller defines _MEIPASS; other freezers (Nuitka, cx_Freeze) don't.
        base_dir = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    else:
        # running from source – use the directory where *this* file lives
        base_dir = Path(__file__).parent

    return str(base_dir / rel)

class XtractorGUI:
    def __init__(self, root):
        self.root = root

        self.extractor = ExtractController()

        # compute UI scale from Tk (dpi-aware thanks to main.py)
        self._ui_scale = float(self.root.tk.call('tk', 'scaling')) / (96 / 72)  # = 1.0 at 96 DPI
        def px(v: float) -> int: return int(round(v * self._ui_scale))
        self._px = px

        self.pdf_viewer = PDFViewer(self, self.root)  # Pass GUI instance and root window


        self.root.update_idletasks()


        self.pdf_folder = ''
        self.output_excel_path = ''
        self.ocr_settings = {'enable_ocr': 'Default', 'dpi_value': 150, 'tessdata_folder': TESSDATA_FOLDER}
        self.recent_pdf_path = None
        self.revision_column_selection = {"rev": None, "desc": None, "date": None}

        self.setup_widgets()
        self.ocr_menu_callback("Default")
        self.setup_bindings()
        self.setup_tooltips()
        self.update_revision_pattern_controls()
        self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)

    def _on_app_close(self):
        try:
            if hasattr(self, "_job") and self._job:
                self.extractor.cancel(self._job)
        except Exception:
            pass
        self.root.destroy()

    # inside class XtractorGUI
    # inside class XtractorGUI

    def _launch_tool(self, label: str, tool: dict):
        try:
            fn = tool["action"]
            if tool.get("needs_master", False):
                fn(self.root)
            else:
                fn()
        except Exception as e:
            logger.exception("Failed to open tool: %s", label)
            messagebox.showerror("Tool failed", f"{label}\n\n{e}")



    def export_rectangles(self):
        export_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Rectangles As"
        )
        if not export_file_path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws_area = wb.active
            ws_area.title = "Rectangles"
            ws_area.append(["Title", "x0", "y0", "x1", "y1"])
            for area in self.pdf_viewer.areas:
                title, (x0, y0, x1, y1) = _area_get_title_and_coords(area)
                ws_area.append([title, x0, y0, x1, y1])

            if self.pdf_viewer.revision_area:
                ws_rev = wb.create_sheet("RevisionTable")
                ws_rev.append(["Title", "x0", "y0", "x1", "y1"])
                title, (x0, y0, x1, y1) = _area_get_title_and_coords(self.pdf_viewer.revision_area)
                ws_rev.append([title, x0, y0, x1, y1])

            wb.save(export_file_path)
            wb.close()
            print(f"Exported areas to {export_file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export areas: {e}")

    def import_rectangles(self):
        import_file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm"), ("All files", "*.*")],
            title="Import Rectangles"
        )
        if not import_file_path:
            return
        self.import_rectangles_from_file(import_file_path)

    def import_rectangles_from_file(self, file_path):
        try:
            wb = load_workbook(file_path)
            ws_area = wb["Rectangles"] if "Rectangles" in wb.sheetnames else wb.active

            # always store GUI dicts in PDFViewer.areas
            areas = []
            for row in ws_area.iter_rows(min_row=2, values_only=True):
                title, x0, y0, x1, y1 = row
                if None in (x0, y0, x1, y1):
                    continue
                areas.append({"title": title or "Area", "coordinates": [float(x0), float(y0), float(x1), float(y1)]})
            self.pdf_viewer.set_gui_areas(areas)

            # Revision area (optional)
            self.pdf_viewer.revision_area = None
            if "RevisionTable" in wb.sheetnames:
                ws_rev = wb["RevisionTable"]
                for row in ws_rev.iter_rows(min_row=2, values_only=True):
                    title, x0, y0, x1, y1 = row
                    if None in (x0, y0, x1, y1):
                        continue
                    self.pdf_viewer.set_gui_revision_area({
                        "title": title or "Revision Table",
                        "coordinates": [float(x0), float(y0), float(x1), float(y1)]
                    })
                    break

            print(f"✅ Imported areas from dropped Excel file: {file_path}")

        except Exception as e:
            messagebox.showerror("Import Error", f"Could not import areas from dropped Excel file: {e}")

    def clear_all_areas(self):
        """Clears all areas and updates the display."""
        self.pdf_viewer.clear_areas()
        self.pdf_viewer.set_gui_revision_area(None)  # also clears the green rectangle
        self.revision_column_selection = {"rev": None, "desc": None, "date": None}
        # Tree refresh is triggered by update_rectangles() inside clear/set,
        # but keeping the Treeview wipe is harmless:
        self.areas_tree.delete(*self.areas_tree.get_children())
        print("All areas and revision table cleared.")

    def update_areas_treeview(self):
        self.areas_tree.delete(*self.areas_tree.get_children())
        self.treeview_item_ids = {}
        for index, area in enumerate(self.pdf_viewer.areas):
            title, (x0, y0, x1, y1) = _area_get_title_and_coords(area)
            item_id = self.areas_tree.insert("", "end", values=(title, x0, y0, x1, y1))
            self.treeview_item_ids[item_id] = index

    def _prompt_revision_columns(self, rows):
        if not rows:
            print("[DetectPattern] _prompt_revision_columns called with no rows.")
            return None

        column_count = max((len(r) for r in rows if r), default=0)
        if column_count == 0:
            print("[DetectPattern] No columns found in detected table rows.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return None

        parser = RevisionParser(None)
        suggested_rev, suggested_desc, suggested_date = parser.detect_column_indices(rows)
        print(
            "[DetectPattern] Suggested column indices from parser:",
            f"rev={suggested_rev}, desc={suggested_desc}, date={suggested_date}"
        )

        prev_rev = self.revision_column_selection.get("rev") if self.revision_column_selection else None
        prev_desc = self.revision_column_selection.get("desc") if self.revision_column_selection else None
        prev_date = self.revision_column_selection.get("date") if self.revision_column_selection else None

        default_rev_idx = prev_rev if isinstance(prev_rev, int) and 0 <= prev_rev < column_count else None
        if column_count <= 0:
            default_rev_idx = 0
        elif default_rev_idx is None:
            default_rev_idx = 0
        else:
            default_rev_idx = min(max(default_rev_idx, 0), column_count - 1)

        default_desc_idx = prev_desc if isinstance(prev_desc, int) and 0 <= prev_desc < column_count else suggested_desc
        if default_desc_idx is not None and not (0 <= default_desc_idx < column_count):
            default_desc_idx = None

        default_date_idx = prev_date if isinstance(prev_date, int) and 0 <= prev_date < column_count else suggested_date
        if default_date_idx is not None and not (0 <= default_date_idx < column_count):
            default_date_idx = None

        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Verify Revision Columns")
        dialog.geometry("560x380")
        dialog.transient(self.root)
        dialog.grab_set()

        info = ctk.CTkLabel(
            dialog,
            text="Verify the detected table columns and confirm which index holds each value.",
            wraplength=520,
            justify="left"
        )
        info.pack(padx=15, pady=(12, 6), anchor="w")

        tree = ttk.Treeview(dialog, columns=[str(i) for i in range(column_count)], show="headings", height=min(len(rows), 8))
        for idx in range(column_count):
            col_name = f"Column {idx}"
            tree.heading(str(idx), text=col_name)
            tree.column(str(idx), width=160, anchor="w")

        preview_rows = rows[:8]
        for row in preview_rows:
            values = [str(row[i]).strip() if i < len(row) else "" for i in range(column_count)]
            tree.insert("", "end", values=values)

        tree.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        option_values = [str(i) for i in range(column_count)]
        option_with_none = option_values + ["None"]

        rev_var = StringVar(value=str(default_rev_idx))
        desc_var = StringVar(value=str(default_desc_idx) if default_desc_idx is not None else "None")
        date_var = StringVar(value=str(default_date_idx) if default_date_idx is not None else "None")

        options_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        options_frame.pack(fill="x", padx=15, pady=(0, 10))

        def build_option(row_label, variable, values):
            row_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
            row_frame.pack(fill="x", pady=4)
            label = ctk.CTkLabel(row_frame, text=row_label, width=140, anchor="w")
            label.pack(side="left")
            menu = ctk.CTkOptionMenu(row_frame, values=values, variable=variable, width=120)
            menu.pack(side="left", padx=(10, 0))
            return menu

        build_option("Revision column", rev_var, option_values)
        build_option("Description column", desc_var, option_with_none)
        build_option("Date column", date_var, option_with_none)

        result = {"selection": None}

        def on_confirm():
            rev_choice = rev_var.get()
            if rev_choice not in option_values:
                print("[DetectPattern] Revision column not selected during confirmation dialog.")
                messagebox.showinfo("Detect Pattern", "Select a revision column before continuing.")
                return
            try:
                rev_idx = int(rev_choice)
                desc_choice = desc_var.get()
                date_choice = date_var.get()
                desc_idx = None if desc_choice == "None" else int(desc_choice)
                date_idx = None if date_choice == "None" else int(date_choice)
            except ValueError:
                print("[DetectPattern] Failed to parse selected column indices.")
                messagebox.showinfo("Detect Pattern", "Column selections are invalid.")
                return

            result["selection"] = (rev_idx, desc_idx, date_idx)
            print(
                "[DetectPattern] User confirmed column indices:",
                f"rev={rev_idx}, desc={desc_idx}, date={date_idx}"
            )
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(pady=(0, 12))

        confirm_btn = ctk.CTkButton(button_frame, text="Confirm", command=on_confirm, width=100)
        confirm_btn.pack(side="left", padx=6)
        cancel_btn = ctk.CTkButton(button_frame, text="Cancel", command=on_cancel, width=100)
        cancel_btn.pack(side="left", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        self.root.wait_window(dialog)

        return result.get("selection")

    def detect_revision_pattern(self):
        rev_area = self.pdf_viewer.revision_area
        pdf_path = getattr(self.pdf_viewer, "current_pdf_path", None) or self.recent_pdf_path

        if not rev_area or not pdf_path:
            print(
                "[DetectPattern] Cannot run detection –",
                f"rev_area={'present' if rev_area else 'missing'}, pdf_path={pdf_path}"
            )
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        coords = rev_area.get("coordinates") if isinstance(rev_area, dict) else None
        if not coords or len(coords) != 4:
            print(f"[DetectPattern] Invalid revision area coordinates: {coords}")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        adapter = PdfAdapter()
        try:
            with adapter.open(pdf_path) as doc:
                if doc.page_count == 0:
                    print("[DetectPattern] PDF appears to have no pages.")
                    raise ValueError("PDF has no pages")
                page_number = getattr(getattr(self.pdf_viewer, "page", None), "number", 0) or 0
                page_number = min(max(page_number, 0), doc.page_count - 1)
                print(
                    f"[DetectPattern] Analysing PDF '{pdf_path}' page {page_number + 1}/{doc.page_count}"
                )
                page = doc.load_page(page_number)
                rows = adapter.find_table_rows(page, tuple(coords))
                print(
                    "[DetectPattern] PdfAdapter returned",
                    f"{len(rows) if rows else 0} rows"
                )
        except Exception as exc:
            print(f"[DetectPattern] Exception during table detection: {exc}")
            messagebox.showerror("Detect Pattern", f"Failed to analyse revision table: {exc}")
            return

        if not rows:
            print("[DetectPattern] No rows extracted from detected table region.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        selection = self._prompt_revision_columns(rows)
        if selection is None:
            print("[DetectPattern] User cancelled or dialog failed to return column selection.")
            return

        rev_idx, desc_idx, date_idx = selection
        self.revision_column_selection = {"rev": rev_idx, "desc": desc_idx, "date": date_idx}

        first_column = []
        for row in rows:
            if not row:
                continue
            cell = ""
            if isinstance(rev_idx, int) and 0 <= rev_idx < len(row):
                cell = str(row[rev_idx]).strip()
            elif row:
                cell = str(row[0]).strip()
            if not cell:
                continue
            first_column.append(cell)

        if not first_column:
            print("[DetectPattern] Extracted rows had no usable revision values.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        data_values = [c for c in first_column if not re.search(r"\brev(ision)?", c, re.IGNORECASE)]
        if not data_values:
            data_values = first_column

        normalized_values = []
        for value in data_values:
            cleaned = re.sub(r"\s+", "", value)
            if cleaned:
                normalized_values.append(cleaned)

        if not normalized_values:
            print("[DetectPattern] No normalized revision values available for pattern matching.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        total = len(normalized_values)
        best_key = None
        best_ratio = 0.0
        best_matches = 0

        for key, info in REVISION_PATTERNS.items():
            try:
                regex = re.compile(info["pattern"], re.IGNORECASE)
            except re.error:
                continue
            matches = sum(1 for value in normalized_values if regex.fullmatch(value))
            if not matches:
                continue
            ratio = matches / total if total else 0
            if (ratio > best_ratio) or (abs(ratio - best_ratio) < 1e-9 and matches > best_matches):
                best_key = key
                best_ratio = ratio
                best_matches = matches

        if best_key is None:
            print("[DetectPattern] Pattern matching found no candidates.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        required_ratio = 1.0 if total <= 2 else 0.6
        if best_ratio < required_ratio:
            print(
                "[DetectPattern] Best match ratio",
                f"{best_ratio:.2f} below threshold {required_ratio:.2f}"
            )
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        display_value = self.revision_key_to_display.get(best_key)
        if not display_value:
            print(f"[DetectPattern] No dropdown display value mapped for key '{best_key}'.")
            messagebox.showinfo("Detect Pattern", "Revision pattern could not be identified.")
            return

        self.revision_pattern_var.set(display_value)
        self.revision_pattern_menu.set(display_value)
        print(f"Detected revision pattern '{best_key}' using {best_matches}/{total} samples.")

    def update_revision_pattern_controls(self):
        if not hasattr(self, "detect_revision_pattern_button"):
            return

        has_revision_area = bool(self.pdf_viewer.revision_area)
        pdf_available = bool(getattr(self.pdf_viewer, "current_pdf_path", None) or self.recent_pdf_path)
        state = "normal" if (has_revision_area and pdf_available) else "disabled"
        self.detect_revision_pattern_button.configure(state=state)

    def on_revision_area_changed(self):
        if not self.pdf_viewer.revision_area:
            self.revision_column_selection = {"rev": None, "desc": None, "date": None}
        self.update_revision_pattern_controls()

    def on_pdf_loaded(self, path: str):
        self.recent_pdf_path = path
        self.update_revision_pattern_controls()

    def on_pdf_closed(self):
        self.update_revision_pattern_controls()

    def open_sample_pdf(self):
        # Opens a file dialog to select a PDF file, then displays it in the PDFViewer
        pdf_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if pdf_path:
            self.pdf_viewer.display_pdf(pdf_path)
            self.recent_pdf_path = pdf_path  # Store the recent PDF path
            self.update_revision_pattern_controls()
            print(f"Opened sample PDF: {pdf_path}")

    def open_recent_pdf(self):
        """Opens the most recently viewed PDF."""
        if self.recent_pdf_path:
            self.pdf_viewer.display_pdf(self.recent_pdf_path)
            print(f"Reopened recent PDF: {self.recent_pdf_path}")
        else:
            messagebox.showinfo("Info", "No recent PDF found.")

    def close_pdf(self):
        """Delegates PDF closing to the PDF viewer."""
        self.pdf_viewer.close_pdf()

    def remove_row(self):
        selected = self.areas_tree.selection()
        if not selected:
            return
        index = self.treeview_item_ids.get(selected[0])
        if index is None:
            return

        areas = self.pdf_viewer.get_gui_areas()
        if 0 <= index < len(areas):
            del areas[index]
            self.pdf_viewer.set_gui_areas(areas)
            self.areas_tree.delete(selected[0])
            self.update_areas_treeview()
            print("Removed rectangle at index", index)

    def setup_widgets(self):
        # Create a tabbed panel on the left
        self.tab_view = ctk.CTkTabview(self.root, width=SIDEBAR_WIDTH)
        self.tab_view.pack(side="left", fill="y", padx=SIDEBAR_PADDING, pady=10)

        self.root.update_idletasks()  # Ensure sidebar dimensions are accurate

        # Zoom Slider
        self.zoom_var = ctk.DoubleVar(value=self.pdf_viewer.current_zoom)  # Initialize with the current zoom level
        self.zoom_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        self.zoom_out_label = ctk.CTkLabel(self.zoom_frame, text="➖", font=(BUTTON_FONT, 10))
        self.zoom_slider = ctk.CTkSlider(self.zoom_frame, from_=0.1, to=4, variable=self.zoom_var,
                                         command=self.update_zoom, width=self._px(170), height=self._px(8))
        self.zoom_in_label = ctk.CTkLabel(self.zoom_frame, text="➕", font=(BUTTON_FONT, 10))

        self.zoom_out_label.pack(side="left", padx=(5, 2))
        self.zoom_slider.pack(side="left")
        self.zoom_in_label.pack(side="left", padx=(2, 5))

        # Floating recent PDF and close PDF buttons (top-right)
        # Create a label that behaves like a button
        self.recent_pdf_button = ctk.CTkLabel(
            self.root,
            text="↩",
            font=(BUTTON_FONT, 12, "bold"),
            text_color="lightblue",
            cursor="hand2",
            width=self._px(24),
            height=self._px(24)
        )
        self.recent_pdf_button.pack(pady=2)
        self.recent_pdf_button.bind("<Button-1>", lambda e: self.open_recent_pdf())

        self.close_pdf_button = ctk.CTkLabel(
            self.root,
            text="X",
            font=(BUTTON_FONT, 10, "bold"),
            text_color="red",
            cursor="hand2",
            width=self._px(24),
            height=self._px(24)
        )
        self.close_pdf_button.pack(pady=2)
        self.close_pdf_button.bind("<Button-1>", lambda e: self.close_pdf())

        # Create each tab
        tab_files = self.tab_view.add("Files")
        tab_rectangles = self.tab_view.add("Rectangles")
        tab_extract = self.tab_view.add("Extract")
        tab_tools = self.tab_view.add("Tools")

        # --- Extract overlay (drawn in the main canvas area, where the PDF viewer sits) ---
        self.extract_overlay = ctk.CTkFrame(self.root, fg_color="transparent", width=1, height=1)
        self._build_extract_mode_cards(self.extract_overlay)

        self.extract_overlay.place_forget()

        # make the overlay content responsive
        self.extract_overlay.bind("<Configure>", self._on_extract_overlay_configure)
        self.extract_overlay.after_idle(self._on_extract_overlay_configure)

        # --- Tools overlay (occupies the PDF canvas area) ---
        self.tools_overlay = ctk.CTkFrame(self.root, fg_color="transparent", width=1, height=1)
        self._build_tools_cards(self.tools_overlay)
        self.tools_overlay.place_forget()

        # keep it responsive
        self.tools_overlay.bind("<Configure>", lambda e: self._on_tools_overlay_configure())

        # ======================= 📁 FILES TAB =======================

        # PDF Folder Drop Zone
        # --- Drop Folder card ---
        self.pdf_folder_card = ctk.CTkFrame(
            tab_files, fg_color="transparent",
            border_width=2, border_color="#C10206",
            corner_radius=8, width=240, height=120
        )
        self.pdf_folder_card.pack_propagate(False)
        self.pdf_folder_card.pack(pady=(10, 5))

        # Centered inner wrapper
        self._drop_wrap = ctk.CTkFrame(self.pdf_folder_card, fg_color="transparent")
        self._drop_wrap.place(relx=0.5, rely=0.5, anchor="center")  # <- centers content

        # Big title (centered)
        self.pdf_folder_title = ctk.CTkLabel(
            self._drop_wrap, text="DRAG & DROP", text_color="#B5B5B5",
            font=("Arial Black", 18), justify="center"
        )
        self.pdf_folder_title.pack(anchor="center")

        # Small subtitle (centered)
        self.pdf_folder_sub = ctk.CTkLabel(
            self._drop_wrap, text="Drop Folder with PDFs\nor Click to Browse",
            font=(BUTTON_FONT, 10), text_color="#B5B5B5",
            justify="center", wraplength=200
        )
        self.pdf_folder_sub.pack(anchor="center", pady=(2, 0))

        # Click + hover
        for w in (self.pdf_folder_card, self._drop_wrap, self.pdf_folder_title, self.pdf_folder_sub):
            w.bind("<Button-1>", lambda e: self.browse_pdf_folder())
            w.bind("<Enter>", lambda e: self.pdf_folder_card.configure(fg_color="gray40",border_color="gray70"))
            w.bind("<Leave>", lambda e: self.pdf_folder_card.configure(fg_color="#212121",border_color="#C10206"))

        # DnD
        if DND_ENABLED:
            try:
                self.pdf_folder_card.drop_target_register(DND_ALL)
                self.pdf_folder_card.dnd_bind('<<Drop>>', self.drop_pdf_folder)
            except Exception as e:
                print("Could not enable DnD on pdf_folder_card:", e)

        self.pdf_folder_entry = ctk.CTkEntry(self.root, width=240, height=24, font=(BUTTON_FONT, 9),
                                             placeholder_text="Select Folder with PDFs", border_width=1,
                                             corner_radius=3)
        self.pdf_folder_entry.place(x=-200, y=-223) #hide haha
        # Files Tree Label
        self.files_tree_label = ctk.CTkLabel(tab_files, text="📄 Files in Folder", font=(BUTTON_FONT, 10))
        self.files_tree_label.pack(pady=(10, 0))

        # Create Frame to hold the tree
        self.files_tree_frame = ctk.CTkFrame(tab_files, height=140, width=240)
        self.files_tree_frame.pack(pady=(2, 5), fill="both", expand=True)

        # 🧱 Create scrollable frame ONCE
        self.files_tree_scrollframe = ctk.CTkScrollableFrame(
            self.files_tree_frame,
            orientation="horizontal",
            width=240,
            height=150,
            fg_color="transparent"
        )
        self.files_tree_scrollframe.pack(fill="both", expand=True)

        # Inner container for Treeview
        self.files_tree_container = ctk.CTkFrame(self.files_tree_scrollframe, fg_color="transparent")
        self.files_tree_container.pack(side="left", fill="both", expand=True)
        self.files_tree_scrollbar = None

        # Placeholder for Treeview widget
        self.files_tree_widget = None

        self.files_tree_widget = CheckboxTreeview(self.files_tree_container, show="tree", height=10)
        self.files_tree_widget.pack(side="left", fill="both", expand=True)
        self.files_tree_widget.column("#0", width=800, stretch=False)

        # Counter Label (created only once)
        self.pdf_counter_label = ctk.CTkLabel(self.files_tree_frame, text="Selected PDFs: 0", font=(BUTTON_FONT, 9))
        self.pdf_counter_label.pack(pady=(5, 0), anchor="center")

        # ======================= 🔲 RECTANGLES TAB =======================


        # 🔲 Frame for Mode Buttons (Area / Revision)
        mode_frame = ctk.CTkFrame(tab_rectangles, width=240, height=24, fg_color="transparent")
        mode_frame.pack_propagate(False)
        mode_frame.pack(pady=(5, 5))

        self.mode_area_btn = ctk.CTkButton(mode_frame, text="🟥 Area", command=self.set_mode_area,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.mode_area_btn.pack(side="left", padx=5)

        self.mode_revision_btn = ctk.CTkButton(mode_frame, text="🟩 Revision Table", command=self.set_mode_revision,
                                               font=(BUTTON_FONT, 9), width=115, height=24)
        self.mode_revision_btn.pack(side="left", padx=5)

        #revision pattern
        pattern_options = [f"{k} — {', '.join(v['examples'])}" for k, v in REVISION_PATTERNS.items()]
        self.revision_dropdown_map = {f"{k} — {', '.join(v['examples'])}": k for k, v in REVISION_PATTERNS.items()}
        self.revision_key_to_display = {v: k for k, v in self.revision_dropdown_map.items()}
        self.revision_pattern_var = StringVar(value=pattern_options[-1])

        revision_pattern_frame = ctk.CTkFrame(tab_rectangles, fg_color="transparent")
        revision_pattern_frame.pack(pady=(5, 5), fill="x")

        self.revision_pattern_menu = ctk.CTkOptionMenu(
            revision_pattern_frame,
            font=("Verdana", 9),
            values=pattern_options,
            variable=self.revision_pattern_var,
            width=160,
            height=24,
            fg_color="red4",
            button_color="red4",
            text_color="white"
        )
        self.revision_pattern_menu.pack(side="left", fill="x", expand=True)

        self.detect_revision_pattern_button = ctk.CTkButton(
            revision_pattern_frame,
            text="Detect",
            command=self.detect_revision_pattern,
            font=(BUTTON_FONT, 9),
            width=90,
            height=24,
            fg_color="gray35"
        )
        self.detect_revision_pattern_button.pack(side="left", padx=(5, 0))
        self.detect_revision_pattern_button.configure(state="disabled")

        # 🛈 How to Use Button (Revision Table)
        self.revision_help_button = ctk.CTkButton(
            tab_rectangles,
            text="How to Use?",
            command=self.show_revision_help,
            font=(BUTTON_FONT, 9),
            fg_color="gray20", hover_color="gray30",
            width=240,
            height=24
        )
        self.revision_help_button.pack(pady=(0, 5))


        # ⬇️ Frame for Import/Export Buttons
        action_frame = ctk.CTkFrame(tab_rectangles, width=240, height=24, fg_color="transparent")
        action_frame.pack_propagate(False)
        action_frame.pack(pady=(25, 5))

        self.import_button = ctk.CTkButton(action_frame, text="⬇️ Import Areas", command=self.import_rectangles,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.import_button.pack(side="left", padx=5)

        self.export_button = ctk.CTkButton(action_frame, text="⬆️ Export Areas", command=self.export_rectangles,
                                           font=(BUTTON_FONT, 9), width=115, height=24)
        self.export_button.pack(side="left", padx=5)

        # 🗑 Clear Button (Full width)
        # 🧹 Row of Clear Buttons
        clear_frame = ctk.CTkFrame(tab_rectangles, fg_color="transparent", width=240, height=24)
        clear_frame.pack(pady=(5, 15))

        # 🟥 Clear Extraction Areas
        self.clear_extraction_button = ctk.CTkButton(clear_frame, text="Clear Area",
                                                     command=self.clear_extraction_areas,
                                                     font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_extraction_button.pack(side="left", padx=5)

        # 🟩 Clear Revision Table
        self.clear_revision_button = ctk.CTkButton(clear_frame, text="Clear Rev",
                                                   command=self.clear_revision_area,
                                                   font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_revision_button.pack(side="left", padx=5)

        # 🗑 Clear Button (Full width)
        self.clear_areas_button = ctk.CTkButton(clear_frame, text="Clear All", command=self.clear_all_areas,
                                                font=(BUTTON_FONT, 9), width=75, height=24, fg_color="gray35")
        self.clear_areas_button.pack(side="left", padx=5)

        # Treeview inside a sub-frame
        self.areas_frame = ctk.CTkFrame(tab_rectangles, width=240)
        self.areas_frame.pack_propagate(False)  # Optional to maintain control over child sizes

        self.areas_frame.pack(pady=5, fill="both", expand=True)
        self.areas_tree = EditableTreeview(self, self.areas_frame,
                                           columns=("Title", "x0", "y0", "x1", "y1"),
                                           show="headings", height=4)
        self.areas_tree.heading("Title", text="Title")
        self.areas_tree.column("Title", width=60, anchor="center")
        for col in ("x0", "y0", "x1", "y1"):
            self.areas_tree.heading(col, text=col)
            self.areas_tree.column(col, width=40, anchor="center")
        self.areas_tree.pack(side="left", fill="both", expand=True)

        scrollbar = ctk.CTkScrollbar(self.areas_frame, orientation="vertical", command=self.areas_tree.yview)
        self.areas_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.areas_tree.bind("<<TreeviewSelect>>", self.on_treeview_select)

        # ======================= 🚀 EXTRACT TAB =======================

        # 📦 Frame for OCR & DPI Settings
        extract_frame = ctk.CTkFrame(tab_extract, width=240, fg_color="transparent")
        extract_frame.pack(pady=(25, 5))

        # ─────────────── OCR Setting ───────────────
        ocr_row = ctk.CTkFrame(extract_frame, fg_color="transparent")
        ocr_row.pack(pady=(0, 5), fill="x")
        ocr_row.pack_forget()  # hidden immediately

        ocr_label = ctk.CTkLabel(ocr_row, text="OCR Mode:", font=(BUTTON_FONT, 9), width=80, anchor="w")
        ocr_label.pack(side="left", padx=(0, 5))
        ocr_label.pack_forget()  # hidden

        self.ocr_menu_var = StringVar(value="Default")
        self.ocr_menu = ctk.CTkOptionMenu(ocr_row,
                                          values=[ "Default", "OCR-All", "Text1st+Image-beta"],
                                          command=self.ocr_menu_callback,
                                          font=("Verdana Bold", 9),
                                          variable=self.ocr_menu_var,
                                          width=140, height=24)
        self.ocr_menu.pack(side="left", padx=(0, 5))
        self.ocr_menu.pack_forget()  # hidden

        # Add help "?" button next to OCR dropdown
        self.ocr_help_button = ctk.CTkButton(
            ocr_row,
            text="?",
            width=24,
            height=24,
            fg_color="gray20", hover_color="gray30",
            font=(BUTTON_FONT, 10),
            command=self.show_ocr_help
        )
        self.ocr_help_button.pack(side="left", padx=(5, 0))

        extract_frame.pack_forget()
        # --- A single column container that uses GRID
        extract_column = ctk.CTkFrame(tab_extract, fg_color="transparent")
        extract_column.pack(fill="both", expand=True, pady=(0, 10))

        extract_column.grid_columnconfigure(0, weight=1)
        # Make the description row NOT grow, and let the controls row take the extra space
        extract_column.grid_rowconfigure(0, weight=0)  # was 1
        extract_column.grid_rowconfigure(1, weight=1)  # was 0

        # ---- TOP: description (fixed height, hugs the top) ----
        self.extract_description_box = ctk.CTkTextbox(
            extract_column,
            wrap="word",
            fg_color="#292929",
            font=(BUTTON_FONT, 9),
            height=480,
        )
        self.extract_description_box.insert("end",
                                            "EXTRACTION SUMMARY\n\n"
                                            "Output Excel File (.xlsx):\n"
                                            "  - One row per PDF page\n"
                                            "  - Columns from selected areas\n"
                                            "  - Hyperlinks to original PDF files\n\n"
                                            "IF REVISION TABLE IS USED:\n"
                                            "  ️‼️‼️EXTRACTION TIME WILL INCREASE BY x3 or x4‼️‼️\n\n"
                                            "  - Adds revision rows (Rev, Desc, Date)\n"
                                            "  - Saves NDJSON with structured revision info\n\n"
                                            "IF TEXT IS OCR-ed:\n"
                                            "  - Texts will be colored red\n"
                                            )
        self.extract_description_box.configure(state="disabled")
        # stick to top + fill width (no vertical stretch)
        self.extract_description_box.grid(row=0, column=0, sticky="new", padx=0, pady=(0, 10))

        # ======================= 🧰 TOOLS TAB =======================
        # Create a frame for all tools
        tool_frame = ctk.CTkFrame(tab_tools)
        tool_frame.pack(pady=10, fill="both", expand=True)


        # Load PNG from style folder
        logo_image = Image.open(resource_path("style/xtractor-logo.png"))
        logo_image = logo_image.resize((180, 180), Image.Resampling.LANCZOS)
        logo_photo = CTkImage(light_image=logo_image, dark_image=logo_image, size=(180, 180))

        # Create a label to hold the image
        self.logo_label = ctk.CTkLabel(tab_tools, image=logo_photo, text="", width=180, height=45)
        self.logo_label.image = logo_photo  # Prevent garbage collection
        self.logo_label.pack(pady=(20, 10), anchor="center")

        self.version_label = ctk.CTkLabel(
            tab_tools,
            text=VERSION_TEXT,
            fg_color="transparent",
            text_color="gray59",
            font=(BUTTON_FONT, 9)
        )
        self.version_label.pack(pady=(0, 15), anchor="center")  # ⬅️ anchor set to center
        self.version_label.bind("<Button-1>", self.display_version_info)

        # Start watching tab selection to toggle PDF viewer vs. Extract panel
        self._current_tab = self.tab_view.get()
        self.root.after(100, self._watch_tab_selection)

        # after building tabs/overlay and calling self.root.update_idletasks()
        self._set_window_minsize_for_cards(min_cols=3)  # or 3 if you always want 3 side-by-side


        # If the app opens with Extract selected, apply hiding immediately
        if self._current_tab == "Extract":
            self._on_tab_changed("Extract")

        self.root.after_idle(self.update_floating_controls)

    def _layout_tools_overlay(self):
        try:
            self.root.update_idletasks()
            canvas_mapped = bool(self.pdf_viewer.canvas.winfo_ismapped())
            if canvas_mapped:
                x = self.pdf_viewer.canvas.winfo_x()
                y = self.pdf_viewer.canvas.winfo_y()
            else:
                sidebar_right = self.tab_view.winfo_x() + self.tab_view.winfo_width()
                x = max(sidebar_right + 8, CANVAS_LEFT_MARGIN)
                y = CANVAS_TOP_MARGIN if "CANVAS_TOP_MARGIN" in globals() else 0

            win_w = self.root.winfo_width()
            win_h = self.root.winfo_height()
            w = max(50, win_w - x - 4)
            h = max(50, win_h - y - 4)

            self.tools_overlay.place_configure(x=x, y=y)
            self.tools_overlay.configure(width=w, height=h)
            self.tools_overlay.lift()
        except Exception as e:
            print(f"Tools overlay layout error: {e}")

    def _on_tools_overlay_configure(self):
        # optional: if you need to recompute wrapping etc. you can do it here.
        pass

    def _build_tools_cards(self, parent):
        # header
        header = ctk.CTkLabel(parent, text="TOOLS", font=(BUTTON_FONT, 28, "bold"))
        header.place(relx=0.5, rely=0.04, anchor="n")

        # grid wrapper below header
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.place(relx=0, rely=0.12, relwidth=1, relheight=0.80)  # 12% top room for header
        wrap.grid_propagate(False)

        # pick 3 columns (responsive optionality omitted for simplicity)
        for c in range(2):
            wrap.grid_columnconfigure(c, weight=1, uniform="tools")
        # rows will expand as needed
        row = col = 0

        # build a compact “card” per tool
        for label, tool in tool_definitions.items():
            card = ctk.CTkFrame(wrap, corner_radius=12, border_width=1,
                                border_color="gray35", fg_color="#2a2a2a")
            card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
            card.grid_rowconfigure(0, weight=0)
            card.grid_rowconfigure(1, weight=1)
            card.grid_rowconfigure(2, weight=0)
            card.grid_columnconfigure(0, weight=1)

            title = ctk.CTkLabel(card, text=label, font=(BUTTON_FONT, 14, "bold"))
            title.grid(row=0, column=0, padx=12, pady=(12, 6), sticky="w")

            # short preview of instructions (first 2–3 lines)

            blurb = tool.get("blurb")
            if not blurb:
                blurb = self._brief(tool.get("instructions") or "")

            desc = ctk.CTkLabel(card, text=blurb, font=(BUTTON_FONT, 10), justify="left", text_color="gray80",
                                wraplength=260)
            desc.grid(row=1, column=0, padx=12, pady=(0, 8), sticky="nwe")

            btn_row = ctk.CTkFrame(card, fg_color="transparent")
            btn_row.grid(row=2, column=0, padx=12, pady=(0, 12), sticky="we")
            btn_row.grid_columnconfigure(0, weight=1)
            btn_row.grid_columnconfigure(1, weight=1)

            open_btn = ctk.CTkButton(
                btn_row, text="Open", height=28,
                command=(lambda lbl=label, t=tool: self._launch_tool(lbl, t))
            )
            open_btn.grid(row=0, column=0, padx=(0, 6), sticky="we")

            help_btn = ctk.CTkButton(btn_row, text="How to use?", height=28,
                                     fg_color="gray20", hover_color="gray30",
                                     command=(lambda t=tool: self.show_tool_instructions(t["instructions"])))
            help_btn.grid(row=0, column=1, padx=(6, 0), sticky="we")

            # next cell
            col += 1
            if col >= 2:
                col = 0
                row += 1

    def _brief(self, text: str, limit: int = 120) -> str:
        # collapse whitespace
        s = " ".join(text.split())
        if not s:
            return "—"
        # try first sentence
        import re
        m = re.split(r'(?<=[.!?])\s+', s, maxsplit=1)
        first = m[0] if m else s
        # trim bullets/prefixes
        first = first.lstrip("•-–1234567890. ").strip()
        # limit length
        return (first[:limit - 1] + "…") if len(first) > limit else first

    def _ensure_min_geometry_for_cols(self, n: int = 3):
        """Grow the window (if needed) so n cards can sit side-by-side, and set minsize."""
        try:
            self.root.update_idletasks()
            sidebar_w = self.tab_view.winfo_width() + (SIDEBAR_PADDING * 2)

            need_overlay = (n * CARD_MIN_W) + ((n - 1) * CARD_GAP) + (WRAPPER_PAD * 2)
            need_w = max(400, sidebar_w + need_overlay)  # 400: safety floor

            cur_w = max(1, self.root.winfo_width())
            cur_h = max(1, self.root.winfo_height())

            # set minsize first
            self.root.minsize(need_w, MIN_APP_H)

            # if current is narrower, actively grow it
            if cur_w < need_w:
                self.root.geometry(f"{need_w}x{cur_h}")
        except Exception as e:
            print(f"ensure_min_geometry error: {e}")

    def _layout_extract_overlay(self):
        try:
            self.root.update_idletasks()
            canvas_mapped = bool(self.pdf_viewer.canvas.winfo_ismapped())
            if canvas_mapped:
                x = self.pdf_viewer.canvas.winfo_x()
                y = self.pdf_viewer.canvas.winfo_y()
            else:
                sidebar_right = self.tab_view.winfo_x() + self.tab_view.winfo_width()
                x = max(sidebar_right + 8, CANVAS_LEFT_MARGIN)
                y = CANVAS_TOP_MARGIN if "CANVAS_TOP_MARGIN" in globals() else 0

            win_w = self.root.winfo_width()
            win_h = self.root.winfo_height()
            w = max(50, win_w - x - 4)
            h = max(50, win_h - y - 4)

            self.extract_overlay.place_configure(x=x, y=y)
            self.extract_overlay.configure(width=w, height=h)
            self.extract_overlay.lift()

            # Ensure the very first draw wraps correctly
            self.root.after_idle(self._on_extract_overlay_configure)
        except Exception as e:
            print(f"Overlay layout error: {e}")

    def _set_window_minsize_for_cards(self, min_cols: int = 1):
        """Ensure the window can't be resized narrower than the sidebar + cards."""
        try:
            self.root.update_idletasks()

            # sidebar: actual rendered width + the padding you give it
            sidebar_w = self.tab_view.winfo_width() + (SIDEBAR_PADDING * 2)

            # minimum overlay width to fit `min_cols` cards
            overlay_min_w = (
                    (min_cols * CARD_MIN_W) +
                    ((min_cols - 1) * CARD_GAP) +
                    (WRAPPER_PAD * 2)
            )

            total_min_w = max(400, sidebar_w + overlay_min_w)  # 400 = absolute safety floor

            # lock the app's minimum size
            self.root.minsize(total_min_w, MIN_APP_H)
        except Exception as e:
            print(f"minsize calc error: {e}")

    def _apply_card_style(self, key: str, state: str = "default"):
        """state: 'default' | 'hover' | 'selected'."""
        card = self._card_widgets.get(key)
        if not card:
            return

        # if selected, always show selected style
        if getattr(self, "_selected_mode_key", None) == key:
            card.configure(border_color=CARD_BORDER_SELECTED, fg_color=CARD_BG_SELECTED)
            return

        if state == "hover":
            card.configure(border_color=CARD_BORDER_HOVER, fg_color=CARD_BG_HOVER)
        else:  # default
            card.configure(border_color=CARD_BORDER_DEFAULT, fg_color=CARD_BG_DEFAULT)

    def _on_card_enter(self, key: str):
        self._apply_card_style(key, "hover")
        # optional: cursor feedback on whole card
        try:
            self._card_widgets[key].configure(cursor="hand2")
        except Exception:
            pass

    def _on_card_leave(self, key: str):
        self._apply_card_style(key, "default")
        try:
            self._card_widgets[key].configure(cursor="")
        except Exception:
            pass

    def _build_extract_mode_cards(self, parent):
        """Create three selectable cards for OCR modes (truly responsive, no overflow)."""
        # Container that fills the overlay
        self.extract_cards = ctk.CTkFrame(parent, fg_color="transparent")
        self.extract_cards.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.extract_cards.pack_propagate(False)

        # ===== Header =====
        self._cards_header = ctk.CTkLabel(
            self.extract_cards,
            text="MODE",
            font=(BUTTON_FONT, 36, "bold"),
            text_color="white"
        )
        self._cards_header.place(relx=0.5, y=self._px(8), anchor="n")

        # Use the SAME space for header and footer so the gap between cards and footer
        # equals the header space (title area).
        HEADER_H_FRAC = 0.12
        FOOTER_H_FRAC = 0.12
        BOTTOM_MARGIN_FRAC = 0.04  # ← add this (adjust 0.02–0.08 to taste)

        self._HEADER_H_FRAC = HEADER_H_FRAC
        self._FOOTER_H_FRAC = FOOTER_H_FRAC
        self._BOTTOM_MARGIN_FRAC = BOTTOM_MARGIN_FRAC

        self._WRAPPER_FOOTER_GAP = self._px(70)

        row_h = self._px(36)
        self._FOOTER_MIN_PX = (row_h * 2) + self._px(16)

        # ===== Cards wrapper (grid) =====
        wrapper = ctk.CTkFrame(self.extract_cards, fg_color="transparent")
        wrapper.place(relx=0, rely=HEADER_H_FRAC, relwidth=1,
                      relheight=(1 - HEADER_H_FRAC - FOOTER_H_FRAC - BOTTOM_MARGIN_FRAC))
        wrapper.grid_propagate(False)
        self._cards_wrapper = wrapper

        for c in range(3):
            wrapper.grid_columnconfigure(c, weight=1, uniform="cards")
        wrapper.grid_rowconfigure(0, weight=1)

        cards = [
            ("+IMAGE",
             "✓  Extracts embedded text.\n\n✓  OCR if no text.\n\n✓  +image in Excel.",
             "Text1st+Image-beta"),
            ("NORMAL",
             "✓  Extracts embedded text.\n\n✓  OCR if no text.\n\n",
             "Default"),
            ("OCR",
             "✘ Extracts embedded text.\n\n✓  OCR will always run.",
             "OCR-All"),
        ]

        self._card_widgets = {}
        self._card_desc_labels = {}
        self._card_order = []

        for col, (title, desc, key) in enumerate(cards):
            card = ctk.CTkFrame(wrapper, corner_radius=12, border_width=2,
                                border_color="gray35", fg_color="gray20")
            card.grid(row=0, column=col, padx=8, pady=8, sticky="nsew")

            CARD_TITLE_FONTS = {
                "Text1st+Image-beta": (BUTTON_FONT, 24, "bold"),
                "Default": (BUTTON_FONT, 24, "bold"),
                "OCR-All": (BUTTON_FONT, 24, "bold"),
            }

            card.grid_rowconfigure(0, weight=0)  # title
            card.grid_rowconfigure(1, weight=0)  # description
            card.grid_rowconfigure(2, weight=1)  # filler (absorbs extra vertical space)
            card.grid_columnconfigure(0, weight=1)

            title_lbl = ctk.CTkLabel(card, text=title,
                                     font=CARD_TITLE_FONTS.get(key, (BUTTON_FONT, 12, "bold")))
            title_lbl.grid(row=0, column=0, padx=20, pady=(24, 24), sticky="n")

            desc_lbl = ctk.CTkLabel(card, text=desc, font=(BUTTON_FONT, 10),
                                    wraplength=220, justify="left")
            desc_lbl.grid(row=1, column=0, padx=12, sticky="n")

            for w in (card, title_lbl, desc_lbl):
                # click (already in your code)
                w.bind("<Button-1>", lambda e, _k=key: self._select_mode(_k))
                # hover
                w.bind("<Enter>", lambda e, _k=key: self._on_card_enter(_k))
                w.bind("<Leave>", lambda e, _k=key: self._on_card_leave(_k))

            self._card_widgets[key] = card
            self._card_desc_labels[key] = desc_lbl
            self._card_order.append((key, card))

        # quick keyboard toggles
        try:
            parent.bind_all("<Key-1>", lambda e: self._select_mode("Text1st+Image-beta"))
            parent.bind_all("<Key-2>", lambda e: self._select_mode("Default"))
            parent.bind_all("<Key-3>", lambda e: self._select_mode("OCR-All"))
            parent.bind_all("<Return>", lambda e: self.start_extraction())  # Enter to start
        except Exception:
            pass

        # ===== Footer row (Output path on left, DPI in middle, Start on right) =====
        self._cards_footer = ctk.CTkFrame(
            self.extract_cards,
            fg_color="transparent",
            height=self._FOOTER_MIN_PX  # start with the min; we'll resize in the configure handler
        )
        self._cards_footer.place(
            relx=0, rely=1, anchor="sw",  # stick to bottom-left
            y=-int(self._BOTTOM_MARGIN_FRAC * self.extract_cards.winfo_height()),  # temp; real y set later
            relwidth=1  # width follows parent
        )

        # Prevent children from auto-stretching with the footer's relheight
        self._cards_footer.grid_propagate(False)

        # Grid: 2 rows (entry, browse), 3 cols (left stack, DPI tall, Start tall)
        self._cards_footer.grid_columnconfigure(0, weight=1)  # left expands horizontally
        self._cards_footer.grid_columnconfigure(1, weight=0)  # DPI
        self._cards_footer.grid_columnconfigure(2, weight=0)  # Start


        # Footer rows: DO NOT stretch vertically
        self._cards_footer.grid_rowconfigure(0, weight=0, minsize=row_h)
        self._cards_footer.grid_rowconfigure(1, weight=0, minsize=row_h)

        # Left: output entry (row 0)
        self.output_path_entry = ctk.CTkEntry(
            self._cards_footer,
            font=(BUTTON_FONT, 10),
            placeholder_text="Select Excel Output Path",
            border_width=1,
            corner_radius=6
        )
        self.output_path_entry.grid(row=0, column=0, padx=(12, 10), pady=(6, 3), sticky="ew")

        # Left: browse button (row 1)
        self.output_path_button = ctk.CTkButton(
            self._cards_footer,
            text="📂 Browse Output Path",
            font=(BUTTON_FONT, 10),
            command=self.browse_output_path,
            height=row_h
        )
        self.output_path_button.grid(row=1, column=0, padx=(12, 10), pady=(3, 10), sticky="ew")

        # MIDDLE: DPI dropdown spanning both rows (same height as Start)
        dpi_full_height = row_h * 2
        self._dpi_width = self._px(50)

        # fixed-height wrapper; use grid_propagate(False) (we're using grid, not pack)
        self.dpi_wrap = ctk.CTkFrame(self._cards_footer, width=self._dpi_width, height=dpi_full_height)
        self.dpi_wrap.grid(row=0, column=1, rowspan=2, padx=(0, 4), pady=(6, 10), sticky="n")
        self.dpi_wrap.grid_propagate(False)

        # values with newlines
        dpi_values = [f"{v}\nDPI" for v in ["50", "75", "150", "300", "450", "600"]]
        self.dpi_var = ctk.StringVar(value="150\nDPI")

        self.dpi_menu = CTkOptionMenuNoArrow(
            self.dpi_wrap,
            values=dpi_values,
            variable=self.dpi_var,
            command=self.dpi_callback,
            font=(BUTTON_FONT, 10),
            width=self._dpi_width,
            height=self._px(44),  # tall enough for two lines
            item_height=self._px(36)  # taller menu rows
        )
        self.dpi_menu.pack(fill="both", expand=True)

        # Right: Start Extraction (spans both rows) — keep fixed height, don't stretch
        self.start_button_overlay = ctk.CTkButton(
            self._cards_footer,
            text="▶    START EXTRACTION",
            font=("Arial Black", 13),
            corner_radius=10,
            height=dpi_full_height,  # fixed height
            width=self._px(220),
            command=self.start_extraction
        )
        self.start_button_overlay.grid(row=0, column=2, rowspan=2,
                                       padx=(10, 12), pady=(6, 10),
                                       sticky="n")  # was "nsew" → no vertical fill

        # ✅ Ensure "NORMAL" (Default) is visually selected on startup
        self._select_mode("Default", initial=True)



    def _on_extract_overlay_configure(self, event=None):
        try:
            if not hasattr(self, "_cards_wrapper"):
                return

            w = self.extract_overlay.winfo_width()
            h = self.extract_overlay.winfo_height()
            if w <= 1 or h <= 1:
                return

            # Fractions and minimums saved in _build_extract_mode_cards
            header_frac = getattr(self, "_HEADER_H_FRAC", 0.12)
            footer_frac = getattr(self, "_FOOTER_H_FRAC", 0.12)
            bottom_margin_frac = getattr(self, "_BOTTOM_MARGIN_FRAC", 0.04)
            footer_min_px = getattr(self, "_FOOTER_MIN_PX", self._px(36) * 2 + self._px(16))

            header_px = int(h * header_frac)
            footer_px = max(int(h * footer_frac), footer_min_px)
            bottom_margin_px = int(h * bottom_margin_frac)

            # --- place footer with a guaranteed pixel height, lifted by bottom margin
            # CTk rule: set size via configure(); only position with place_configure()
            self._cards_footer.configure(height=footer_px)
            self._cards_footer.place_configure(
                relx=0, rely=1, anchor="sw",  # anchored to bottom-left
                y=-bottom_margin_px,  # lift it up by the bottom margin
                relwidth=1
            )
            footer_y = h - footer_px - bottom_margin_px  # recompute for wrapper sizing

            # --- size the cards wrapper to exactly the space above the footer
            gap_px = getattr(self, "_WRAPPER_FOOTER_GAP", self._px(6))
            middle_h = max(50, footer_y - header_px - gap_px)  # leave a sliver above footer
            self._cards_wrapper.configure(height=middle_h)
            self._cards_wrapper.place_configure(relx=0, rely=0, y=header_px)

            # ===== Responsive cards layout (unchanged except we keep it neat) =====
            available = max(0, w - (WRAPPER_PAD * 2))

            # how many columns can fit: floor((available + GAP) / (MIN + GAP))
            cols = int((available + CARD_GAP) // (CARD_MIN_W + CARD_GAP))
            cols = max(1, min(3, cols))
            cols = 3  # keep your "force 3" rule

            for c in range(3):
                self._cards_wrapper.grid_columnconfigure(c, weight=(1 if c < cols else 0))

            # re-grid cards into rows/cols (less gap above footer)
            for _, card in self._card_order:
                card.grid_forget()

            # --- re-grid cards (cards still stretch; spacer row provides the gap) ---
            for _, card in self._card_order:
                card.grid_forget()

            # also forget an old spacer if it exists
            if hasattr(self, "_wrapper_gap_spacer"):
                try:
                    self._wrapper_gap_spacer.grid_forget()
                except Exception:
                    pass

            rows_needed = (len(self._card_order) + cols - 1) // cols

            for idx, (key, card) in enumerate(self._card_order):
                r = idx // cols
                c = idx % cols
                card.grid(row=r, column=c, padx=8, pady=(8, 8), sticky="nsew")

            # make the rows with cards expand
            for row in range(rows_needed):
                self._cards_wrapper.grid_rowconfigure(row, weight=1)

            # ---- spacer row just below the cards ----
            gap_px = getattr(self, "_WRAPPER_FOOTER_GAP", self._px(40))
            if not hasattr(self, "_wrapper_gap_spacer"):
                self._wrapper_gap_spacer = ctk.CTkFrame(self._cards_wrapper, fg_color="transparent", height=gap_px)
            else:
                self._wrapper_gap_spacer.configure(height=gap_px)

            # one row spanning all columns; no weight (fixed height)
            self._wrapper_gap_spacer.grid(row=rows_needed, column=0, columnspan=cols, sticky="we")
            self._cards_wrapper.grid_rowconfigure(rows_needed, weight=0, minsize=gap_px)

            # compute actual column width and update wraplength
            total_gaps = (cols - 1) * CARD_GAP
            col_width = max(CARD_MIN_W, (available - total_gaps) // cols)

            for c in range(3):
                self._cards_wrapper.grid_columnconfigure(c, weight=1, uniform="cards", minsize=int(col_width))

            for key, _ in self._card_order:
                self._card_desc_labels[key].configure(wraplength=int(col_width - 32))

        except Exception as e:
            print(f"overlay configure error: {e}")

    def set_dpi_width(self, px: int):
        # Don’t go below the content’s minimum, or it won’t shrink
        self._dpi_width = max(px, self._px(120))  # adjust 120 to taste
        self.dpi_wrap.configure(width=self._dpi_width)
        self.dpi_menu.configure(width=self._dpi_width)
        # Optional: lock the grid column to this width too
        self._cards_footer.grid_columnconfigure(1, weight=0, minsize=self._dpi_width)

    def _select_mode(self, key: str, initial: bool = False):
        # remember selection
        self._selected_mode_key = key

        # restyle all cards in one go
        for k in self._card_widgets.keys():
            if k == key:
                # force selected style
                self._card_widgets[k].configure(border_color=CARD_BORDER_SELECTED, fg_color=CARD_BG_SELECTED)
            else:
                self._apply_card_style(k, "default")

        # keep your existing logic
        self.ocr_settings['enable_ocr'] = key
        if not initial:
            self.ocr_menu_callback(key)
        try:
            self.ocr_menu_var.set(key)
        except Exception:
            pass
        if not initial:
            print(f"[Mode] OCR mode selected: {key}")

    def _toggle_floating_controls(self, show: bool):
        """Show/hide zoom slider, recent↩, and close X."""
        widgets = [self.recent_pdf_button, self.close_pdf_button, self.zoom_frame]
        if show:
            self.update_floating_controls()
        else:
            for w in widgets:
                try:
                    w.place_forget()
                except Exception:
                    pass

    def _hide_viewer(self):
        try:
            # always hide empty-state while an overlay (Extract/Tools) is up
            if hasattr(self.pdf_viewer, "_set_empty_state_visible"):
                self.pdf_viewer._set_empty_state_visible(False)

            # keep the canvas dimmed; we’re hiding scrollbars below anyway
            self.pdf_viewer.canvas.configure(bg="#1e1e1e")
        except Exception:
            pass

        # hide scrollbars so they don't show under overlays
        for name in ("h_scrollbar", "v_scrollbar"):
            sb = getattr(self.pdf_viewer, name, None)
            if sb is not None:
                try:
                    sb.place_forget()
                except Exception:
                    try:
                        sb.pack_forget()
                    except Exception:
                        pass

    def _show_viewer(self):
        try:
            self.pdf_viewer.resize_canvas(
                self.root.winfo_width(),
                self.root.winfo_height(),
                x_offset=CANVAS_LEFT_MARGIN
            )
            self.pdf_viewer.update_rectangles()

            # overlays are hidden now; if no PDF, show empty-state again
            if hasattr(self.pdf_viewer, "_set_empty_state_visible"):
                self.pdf_viewer._set_empty_state_visible(not self.pdf_viewer.pdf_document)
        except Exception as e:
            print(f"Error showing viewer: {e}")

    def _on_tab_changed(self, tab_name: str):
        if tab_name in ("Extract", "Tools"):
            self._hide_viewer()
            self._toggle_floating_controls(False)

            # disable any sidebar extract button if you were doing that
            try:
                self.extract_button.configure(state="disabled")
            except Exception:
                pass

            self._ensure_min_geometry_for_cols(3)

            # show the correct overlay
            if tab_name == "Extract":
                self.tools_overlay.place_forget()
                self._layout_extract_overlay()
                self.extract_overlay.lift()
                self.extract_overlay.update_idletasks()
                self._on_extract_overlay_configure()
            else:  # Tools
                self.extract_overlay.place_forget()
                self._layout_tools_overlay()
                self.tools_overlay.lift()
                self.tools_overlay.update_idletasks()
                self._on_tools_overlay_configure()

        else:
            # any other tab => show viewer again, hide overlays
            self.extract_overlay.place_forget()
            self.tools_overlay.place_forget()
            self._show_viewer()
            self._toggle_floating_controls(True)
            try:
                self.extract_button.configure(state="normal")
            except Exception:
                pass

    def _watch_tab_selection(self):
        """Poll the current tab since CTkTabview lacks a built-in change event."""
        try:
            current = self.tab_view.get()
            if current != getattr(self, "_current_tab", None):
                self._current_tab = current
                self._on_tab_changed(current)
        except Exception:
            pass
        # Poll about 5 times/second — responsive without being heavy
        self.root.after(200, self._watch_tab_selection)

    def show_tool_instructions(self, text):
        window = ctk.CTkToplevel(self.root)
        window.title("Tool Instructions")
        window.geometry("500x400")
        text_box = ctk.CTkTextbox(window, wrap="word")
        text_box.insert("end", text)
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def show_ocr_help(self):
        window = ctk.CTkToplevel(self.root)
        window.title("OCR Mode Explanation")
        window.geometry("480x300")
        text_box = ctk.CTkTextbox(window, wrap="word", font=(BUTTON_FONT, 11))
        text_box.insert("end",
                        "🧠 OCR Modes:\n\n"
                        "• Default:\n"
                        "   Extracts text normally. If no text is extracted, OCR is used.\n\n"
                        "• OCR-All:\n"
                        "   Ignores normal text. OCR is always applied.\n\n"
                        "• Text1st+Image-beta:\n"
                        "   Default mode and\n"
                        "   also saves and embeds area image into Excel.\n"
                        )
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def on_treeview_select(self, event):
        # Clear previous selection (restore original red)
        if self.pdf_viewer.selected_rectangle_id is not None:
            self.pdf_viewer.canvas.itemconfig(self.pdf_viewer.selected_rectangle_id, outline="red")
            self.pdf_viewer.selected_rectangle_id = None

        # Get the selected Treeview item
        selected = self.areas_tree.selection()
        if not selected:
            return

        item_id = selected[0]
        rect_index = self.treeview_item_ids.get(item_id)

        for rect_id in self.pdf_viewer.rectangle_list:
            self.pdf_viewer.canvas.itemconfig(rect_id, outline="red", width=2)

        if rect_index is not None and rect_index < len(self.pdf_viewer.rectangle_list):
            rect_id = self.pdf_viewer.rectangle_list[rect_index]
            # Highlight rectangle in yellow
            self.pdf_viewer.canvas.itemconfig(rect_id, outline="yellow", width=3)
            self.pdf_viewer.selected_rectangle_id = rect_id

    def clear_extraction_areas(self):
        """Clears only extraction areas, leaving the revision area untouched."""
        self.pdf_viewer.set_gui_areas([])
        self.update_areas_treeview()  # optional; viewer already redraws
        print("Cleared only extraction areas.")

    def clear_revision_area(self):
        """Clears only the revision area, leaving extraction areas untouched."""
        self.pdf_viewer.set_gui_revision_area(None)
        self.revision_column_selection = {"rev": None, "desc": None, "date": None}
        print("Cleared only revision table area.")

    def show_revision_help(self):
        window = ctk.CTkToplevel(self.root)
        window.title("Revision Table Area and Pattern Explanation")
        window.geometry("480x300")
        text_box = ctk.CTkTextbox(window, wrap="word", font=(BUTTON_FONT, 11))
        text_box.insert("end",
            "🟩 Revision Table Help\n\n"
            "  ️Note: EXTRACTION TIME WILL INCREASE BY x3 or x4 if you use this feature\n\n"
            "• Use 'Revision Table' mode to select a SINGLE AREA containing revision history -revision, date, description.\n\n"
            "• Do not include in the selection area the -header-/-footer- (the one with Rev Date Description).\n\n"
            "• After selecting the area, use the dropdown to choose the expected revision format (e.g. A, B, C or A1, B2).\n\n"
            "• Make sure the selected table area includes 3 columns: Revision, Description, and Date.\n\n")
        text_box.configure(state="disabled")
        text_box.pack(padx=10, pady=10, fill="both", expand=True)
        window.grab_set()

    def update_floating_controls(self):
        """
        Pin Recent↩, Close X, and the Zoom slider to the same x-coordinate
        as the PDF canvas.  Because the canvas is already placed with
        resize_canvas(), this avoids guessing sidebar/border widths.
        """
        self.root.update_idletasks()
        canvas_x = self.pdf_viewer.canvas.winfo_x()
        win_h = self.root.winfo_height()
        gap = self._px(4)
        y_top = self._px(23)
        y_zoom = win_h - self._px(57)

        self.recent_pdf_button.place(x=canvas_x + gap, y=y_top-4)
        self.close_pdf_button.place(x=canvas_x + gap + self._px(30), y=y_top-4)
        self.zoom_frame.place(x=canvas_x + gap, y=y_zoom)

    def setup_bindings(self):
        self.pdf_folder_entry.bind("<KeyRelease>", self.update_pdf_folder)
        self.output_path_entry.bind("<KeyRelease>", self.update_output_path)
        self.root.bind("<Configure>", self.on_window_resize)

    def setup_tooltips(self):
        create_tooltip(self.ocr_menu, "OCR options - select an OCR mode for text extraction")
        create_tooltip(self.dpi_menu, "DPI resolution")
        create_tooltip(self.pdf_folder_entry, "Select the main folder containing PDF files")
        create_tooltip(self.output_path_entry, "Select folder for the Excel output")
        create_tooltip(self.import_button, "Import a saved template of selected areas")
        create_tooltip(self.export_button, "Export the selected areas as a template")
        create_tooltip(self.clear_areas_button, "Clear all selected areas")
        create_tooltip(self.revision_pattern_menu, "Choose the revision format pattern")
        create_tooltip(self.recent_pdf_button, "Open recent PDF")
        create_tooltip(self.close_pdf_button, "Close the opened PDF")

    def build_folder_tree(self):
        # destroy old tree
        if self.files_tree_widget:
            self.files_tree_widget.destroy()
            self.files_tree_widget = None
        # destroy old scrollbar
        if self.files_tree_scrollbar:
            try:
                self.files_tree_scrollbar.destroy()
            except Exception:
                pass
            self.files_tree_scrollbar = None

        # guard
        if not self.pdf_folder or not os.path.isdir(self.pdf_folder):
            return

        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 7))
        style.configure("Treeview.Heading", font=("Arial", 8, "bold"))

        # new tree + scrollbar
        self.files_tree_widget = CheckboxTreeview(self.files_tree_container, show="tree", height=10)
        self.files_tree_widget.pack(side="left", fill="both", expand=True)
        self.files_tree_widget.column("#0", width=800, stretch=False)

        self.files_tree_scrollbar = ttk.Scrollbar(
            self.files_tree_container, orient="vertical", command=self.files_tree_widget.yview
        )
        self.files_tree_widget.configure(yscrollcommand=self.files_tree_scrollbar.set)
        self.files_tree_scrollbar.pack(side="right", fill="y")

        # root node
        root_text = os.path.basename(self.pdf_folder.rstrip(os.sep)) or self.pdf_folder
        root_node = self.files_tree_widget.insert("", "end", text=root_text, tags=("checked",))

        # helper to insert children (folders with PDFs and .pdf files)
        def insert_children(parent, folder_path):
            try:
                entries = sorted(os.listdir(folder_path))
            except Exception as e:
                print(f"Error accessing {folder_path}: {e}")
                return
            for entry in entries:
                full_path = os.path.join(folder_path, entry)
                if os.path.isdir(full_path):
                    if self.has_pdf(full_path):  # only show folders that contain PDFs somewhere under them
                        node = self.files_tree_widget.insert(parent, "end", text=entry, tags=("checked",))
                        insert_children(node, full_path)
                elif entry.lower().endswith(".pdf"):
                    # if we dropped a subset, only those should be checked; otherwise default to checked
                    is_dropped_subset = bool(getattr(self, "dropped_pdf_set", set()))
                    tags = ("checked",) if (not is_dropped_subset or full_path in self.dropped_pdf_set) else (
                    "unchecked",)
                    self.files_tree_widget.insert(parent, "end", text=entry, values=[full_path], tags=tags)

        insert_children(root_node, self.pdf_folder)

        # update counter when user clicks/changes checks
        self.files_tree_widget.bind("<<TreeviewSelect>>", lambda e: self.update_pdf_counter())
        self.files_tree_widget.bind("<ButtonRelease-1>", lambda e: self.root.after(100, self.update_pdf_counter))

        self.update_pdf_counter()

    def recursive_set_check_state(self, item_id):
        children = self.files_tree_widget.get_children(item_id)
        if not children:
            # Leaf node (file)
            item = self.files_tree_widget.item(item_id)
            values = item.get("values", [])
            if values:
                full_path = os.path.abspath(values[0])
                tag = "checked" if full_path in self.dropped_pdf_set else "unchecked"
                self.files_tree_widget.item(item_id, tags=(tag,))
                return tag == "checked"
            else:
                self.files_tree_widget.item(item_id, tags=("unchecked",))
                return False
        else:
            # Folder node
            any_child_checked = False
            for child in children:
                if self.recursive_set_check_state(child):
                    any_child_checked = True

            tag = "checked" if any_child_checked else "unchecked"
            self.files_tree_widget.item(item_id, tags=(tag,))
            return any_child_checked

    def drop_pdf_folder(self, event):
        raw_data = event.data.strip()
        raw_items = re.findall(r'{(.*?)}', raw_data) or [raw_data.strip()]
        cleaned_paths = [os.path.abspath(p.strip('"')) for p in raw_items]

        dropped_pdfs = [p for p in cleaned_paths if os.path.isfile(p) and p.lower().endswith(".pdf")]
        dropped_folders = [p for p in cleaned_paths if os.path.isdir(p)]

        def collect_pdfs_from_folder(folder_path):
            collected = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith(".pdf"):
                        collected.append(os.path.join(root, file))
            return collected

        all_pdfs = dropped_pdfs.copy()
        for folder in dropped_folders:
            all_pdfs.extend(collect_pdfs_from_folder(folder))

        if not all_pdfs:
            messagebox.showerror("Invalid Drop", "No PDF files found in the dropped folders or files.")
            return

        # choose a common root if possible
        try:
            common_root = os.path.commonpath(all_pdfs)
        except ValueError:
            # different drives on Windows -> just use the first file’s folder
            common_root = os.path.dirname(all_pdfs[0])

        if not os.path.isdir(common_root):
            common_root = os.path.dirname(all_pdfs[0])

        self.pdf_folder = common_root
        self.pdf_folder_entry.delete(0, ctk.END)
        self.pdf_folder_entry.insert(0, self.pdf_folder)

        # mark only the dropped files as "checked"
        self.dropped_pdf_set = set(all_pdfs)

        self.build_folder_tree()
        # (optional) if you want to force recalculating tag state top-down:
        # for iid in self.files_tree_widget.get_children(""):
        #     self.recursive_set_check_state(iid)
        self.update_pdf_counter()

        if not self.pdf_viewer.pdf_document:
            self.pdf_viewer.display_pdf(all_pdfs[0])
            self.recent_pdf_path = all_pdfs[0]
            self.update_revision_pattern_controls()
            print(f"Loaded first PDF from dropped items: {all_pdfs[0]}")

    def drop_sample_pdf(self, event):
        path = event.data.strip().replace("{", "").replace("}", "")
        if os.path.isfile(path) and path.lower().endswith(".pdf"):
            self.pdf_viewer.display_pdf(path)
            self.recent_pdf_path = path
            self.update_revision_pattern_controls()
            print(f"Dropped Sample PDF: {path}")
        else:
            messagebox.showerror("Invalid Drop", "Please drop a valid PDF file.")

    def set_mode_area(self):
        self.pdf_viewer.selection_mode = "area"
        print("🟥 Switched to Area Mode")
        self.highlight_mode_button(self.mode_area_btn)
        self.reset_mode_button(self.mode_revision_btn)

    def set_mode_revision(self):
        self.pdf_viewer.selection_mode = "revision"
        print("🟩 Switched to Revision History Mode")
        self.highlight_mode_button(self.mode_revision_btn)
        self.reset_mode_button(self.mode_area_btn)

    def highlight_mode_button(self, button):
        """Visually highlight the active mode button."""
        button.configure(fg_color="#3949AB", text_color="white", border_color="white", border_width=2)

    def reset_mode_button(self, button):
        """Reset the visual style of inactive buttons."""
        button.configure(fg_color="gray25", text_color="white", border_width=0)

    def update_pdf_counter(self):
        if not self.files_tree_widget:
            return

        def is_pdf(iid):
            item = self.files_tree_widget.item(iid)
            text = item.get("text", "")
            return text.lower().endswith(".pdf")

        checked = [iid for iid in self.files_tree_widget.get_checked() if is_pdf(iid)]
        count = len(checked)
        self.pdf_counter_label.configure(text=f"Selected PDFs: {count}")

    def has_pdf(self, folder_path):
        """Recursively checks if the folder or its subfolders contain any .pdf files."""
        try:
            for entry in os.listdir(folder_path):
                full_path = os.path.join(folder_path, entry)
                if os.path.isdir(full_path):
                    if self.has_pdf(full_path):
                        return True
                elif entry.lower().endswith(".pdf"):
                    return True
        except Exception as e:
            print(f"Error scanning {folder_path}: {e}")
        return False

    def ocr_menu_callback(self, choice):
        print("OCR menu dropdown clicked:", choice)

        def enable_ocr_menu(enabled):
            color = "red4" if enabled else "gray29"
            # Both of these may not exist yet during early init — guard them
            if hasattr(self, "ocr_menu"):
                try:
                    self.ocr_menu.configure(fg_color=color, button_color=color)
                except Exception:
                    pass
            if hasattr(self, "dpi_menu"):
                try:
                    self.dpi_menu.configure(
                        state="normal" if enabled else "disabled",
                        fg_color=color,
                        button_color=color
                    )
                except Exception:
                    pass

        if choice in ("Default", "OCR-All", "Text1st+Image-beta"):
            found_tesseract_path = find_tessdata()
            if found_tesseract_path:
                self.ocr_settings['tessdata_folder'] = found_tesseract_path
                enable_ocr_menu(True)
                if choice == "Default":
                    print("OCR will start if no text is extracted.")
                elif choice == "OCR-All":
                    print("OCR will be enabled for every area.")
                elif choice == "Text1st+Image-beta":
                    print("OCR will start if no text is extracted and images will also be extracted.")
            else:
                enable_ocr_menu(False)
                print("Tessdata folder not found. OCR disabled.")

        self.ocr_settings['enable_ocr'] = choice
        print("OCR mode:", self.ocr_settings['enable_ocr'])

    def dpi_callback(self, dpi_value):
        # Accept both plain numbers and labels like "150 DPI"
        try:
            num = int(str(dpi_value).split()[0])
        except Exception:
            num = 150
        self.ocr_settings['dpi_value'] = num
        print(f"DPI set to: {num}")

    def browse_pdf_folder(self):
        self.pdf_folder = filedialog.askdirectory()
        if not self.pdf_folder:
            return
        self.pdf_folder_entry.delete(0, ctk.END)
        self.pdf_folder_entry.insert(0, self.pdf_folder)
        # we are browsing a whole folder -> no subset filter
        self.dropped_pdf_set = set()
        self.build_folder_tree()

    def browse_output_path(self):
        """Opens a dialog to specify the output Excel file path."""
        self.output_excel_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if self.output_excel_path:  # Only update the entry if a file was selected
            self.output_path_entry.delete(0, ctk.END)
            self.output_path_entry.insert(0, self.output_excel_path)

    def update_zoom_slider(self, zoom_level):
        """Updates the zoom slider to reflect the current zoom level in PDFViewer."""
        self.zoom_var.set(zoom_level)

    def update_pdf_folder(self, event):
        self.pdf_folder = self.pdf_folder_entry.get()

    def update_output_path(self, event):
        self.output_excel_path = self.output_path_entry.get()

    def update_zoom(self, value):
        """Adjusts the zoom level of the PDFViewer based on slider input."""
        zoom_level = float(value)
        self.pdf_viewer.set_zoom(zoom_level)  # Update zoom in PDFViewer

    def start_extraction(self):
        # --- guards (unchanged) ---
        if not self.pdf_viewer.areas:
            messagebox.showerror("Extraction Error", "No areas defined. Please select areas before extracting.")
            return
        self.pdf_folder = self.pdf_folder_entry.get()
        if not self.pdf_folder or not os.path.isdir(self.pdf_folder):
            messagebox.showerror("Invalid Folder", "Select a valid folder.")
            return
        self.output_excel_path = self.output_path_entry.get()
        output_dir = os.path.dirname(self.output_excel_path)
        if not self.output_excel_path or not os.path.isdir(output_dir):
            messagebox.showerror("Invalid Output Path", "Select a valid folder.")
            return


        # collect checked PDFs
        checked = self.files_tree_widget.get_checked()
        selected_paths = []
        for iid in checked:
            item = self.files_tree_widget.item(iid)
            if item and "values" in item and item["values"]:
                path = item["values"][0]
                if path.lower().endswith(".pdf"):
                    selected_paths.append(Path(path))
        if not selected_paths:
            messagebox.showerror("No Files Selected", "Please check at least one PDF to extract.")
            return

        # close viewer to release any file locks
        self.pdf_viewer.close_pdf()

        # ensure the overlay stays on top during extraction
        self._hide_viewer()
        self._layout_extract_overlay()
        self.extract_overlay.lift()
        self.root.update_idletasks()

        # progress UI
        self.start_time = time.time()
        self.progress_window = ctk.CTkToplevel(self.root)
        self.progress_window.title("Progress")
        self.progress_window.geometry("320x150")
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        self.progress_window.attributes('-topmost', True)

        self.progress_label = ctk.CTkLabel(
            self.progress_window,
            text="Preparing files… (waiting for remote PDFs if needed)",
        )
        self.progress_label.pack(pady=(10, 2))
        self.total_files_label = ctk.CTkLabel(
            self.progress_window,
            text="Initializing…",
        )
        self.total_files_label.pack(pady=2)

        self.progress_var = ctk.DoubleVar(value=0)
        self.progress_bar = ctk.CTkProgressBar(self.progress_window, variable=self.progress_var,
                                               orientation="horizontal", width=260)
        self.progress_bar.pack(pady=8)
        try:
            self.progress_bar.configure(mode="indeterminate")
            self.progress_bar.start()
        except Exception:
            pass
        self._progress_is_indeterminate = True
        cancel_btn = ctk.CTkButton(self.progress_window, text="Cancel", width=80, command=self.on_cancel_extraction)
        cancel_btn.pack(pady=(2, 8))
        self.progress_window.protocol("WM_DELETE_WINDOW", self.on_cancel_extraction)

        # Build areas list for the request (tolerate dict or AreaSpec)
        areas_spec = []
        for a in self.pdf_viewer.get_gui_areas():
            try:
                from app.domain.models import AreaSpec
                if isinstance(a, AreaSpec):
                    areas_spec.append(a)
                    continue
            except Exception:
                pass
            title, (x0, y0, x1, y1) = _area_get_title_and_coords(a)
            areas_spec.append(AreaSpec(title=title, rect=(x0, y0, x1, y1)))

        # Revision area
        rev_spec = None
        if self.pdf_viewer.revision_area:
            title, (x0, y0, x1, y1) = _area_get_title_and_coords(self.pdf_viewer.revision_area)
            rev_spec = AreaSpec(title=title, rect=(x0, y0, x1, y1))

        # build domain request
        selected_pattern_display = self.revision_pattern_var.get()
        selected_pattern_key = self.revision_dropdown_map[selected_pattern_display]
        selected_revision_regex = REVISION_PATTERNS[selected_pattern_key]["pattern"]

        req = ExtractionRequest(
            pdf_paths=selected_paths,
            output_excel=Path(self.output_excel_path),
            areas=areas_spec,
            revision_area=rev_spec,
            revision_regex=selected_revision_regex,
            ocr=OcrSettings(
                mode=self.ocr_settings['enable_ocr'],
                dpi=int(self.ocr_settings['dpi_value']),
                tessdata_dir=Path(self.ocr_settings['tessdata_folder']) if self.ocr_settings.get(
                    'tessdata_folder') else None
            ),
            revision_column_index=self.revision_column_selection.get("rev"),
            revision_description_index=self.revision_column_selection.get("desc"),
            revision_date_index=self.revision_column_selection.get("date"),
        )

        # start job via controller
        self._job = self.extractor.start(req)

        def _tick():
            polled = self.extractor.poll(self._job)
            if polled is not None:
                processed, total = polled
                if total > 0:
                    if getattr(self, "_progress_is_indeterminate", False):
                        try:
                            self.progress_bar.stop()
                            self.progress_bar.configure(mode="determinate")
                        except Exception:
                            pass
                        self._progress_is_indeterminate = False
                        try:
                            self.progress_label.configure(text="Processing PDFs…")
                        except Exception:
                            pass
                    self.total_files_label.configure(text=f"Processed pages: {processed}/{total}")
                    self.progress_var.set(processed / max(total, 1))
                else:
                    try:
                        self.total_files_label.configure(text="Preparing files…")
                    except Exception:
                        pass
                self.root.after(100, _tick)
            else:
                # finished (or cancelled / error)
                try:
                    if getattr(self, "_progress_is_indeterminate", False):
                        try:
                            self.progress_bar.stop()
                            self.progress_bar.configure(mode="determinate")
                        except Exception:
                            pass
                        self._progress_is_indeterminate = False
                    self.progress_var.set(1)
                except Exception:
                    pass
                try:
                    self.progress_window.destroy()
                except Exception:
                    pass

                out = self.extractor.finish(self._job)  # Path | None
                self._job = None

                end_time = time.time()
                elapsed = end_time - self.start_time
                formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed))

                if not out or not out.exists():
                    # treat as cancelled or failed (controller does not return a separate status)
                    messagebox.showinfo("Stopped", f"Extraction was cancelled or failed after {formatted}.")
                    return

                if messagebox.askyesno("Extraction Complete", f"Completed in {formatted}.\nOpen the Excel file?"):
                    try:
                        os.startfile(out)
                    except Exception as e:
                        messagebox.showerror("Error", f"Could not open the Excel file: {e}")

        self.root.after(100, _tick)

    def on_cancel_extraction(self):
        if hasattr(self, "_job") and self._job:
            if messagebox.askyesno("Cancel extraction", "Stop the extraction now?"):
                self.extractor.cancel(self._job)
                try:
                    if getattr(self, "_progress_is_indeterminate", False):
                        try:
                            self.progress_bar.stop()
                            self.progress_bar.configure(mode="determinate")
                        except Exception:
                            pass
                        self._progress_is_indeterminate = False
                    self.progress_label.configure(text="Cancelling…")
                except Exception:
                    pass

    def on_window_resize(self, event=None):
        new_width = self.root.winfo_width()
        new_height = self.root.winfo_height()

        if hasattr(self, "prev_width") and hasattr(self, "prev_height"):
            if new_width == self.prev_width and new_height == self.prev_height:
                return

        self.prev_width = new_width
        self.prev_height = new_height

        try:
            sidebar_width = self.tab_view.winfo_width() + 20
            current_tab = self.tab_view.get()

            if current_tab not in ("Extract", "Tools"):
                self.pdf_viewer.resize_canvas(
                    self.root.winfo_width(), self.root.winfo_height(),
                    x_offset=CANVAS_LEFT_MARGIN
                )
                self.pdf_viewer.update_rectangles()
                self.zoom_frame.place_configure(x=sidebar_width + 0, y=new_height - 57)
                self.update_floating_controls()
            else:
                self._toggle_floating_controls(False)
                if current_tab == "Extract":
                    self._layout_extract_overlay()
                else:  # Tools
                    self._layout_tools_overlay()

        except Exception as e:
            print(f"Error resizing widgets: {e}")

    def display_version_info(self, event):
        version_text = """
        Created by: Rei Raphael Reveral
        
        Links:
        https://github.com/r-Yayap/MultiplePDF-Areas2Excel
        https://www.linkedin.com/in/rei-raphael-reveral
        """
        window = ctk.CTkToplevel(self.root)
        window.title("Version Info")
        text_widget = ctk.CTkTextbox(window, wrap="word", width=400, height=247)
        text_widget.insert("end", version_text)
        text_widget.pack(padx=10, pady=10, side="left")
        window.grab_set()



