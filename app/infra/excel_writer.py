# app/infra/excel_writer.py
from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Reorder the base metadata columns in the final Excel by editing this list.
# It must contain exactly these six keys (any order):
# "Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No", "Page Size"
OUTPUT_BASE_ORDER = [
    "Size (Bytes)", "Date Last Modified", "Page No", "Page Size", "Folder", "Filename"
]

def _clean_cell_value(val):
    """Strip characters Excel refuses to accept."""
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val


def write_from_csv(
    combined_csv: Path,
    out_path: Path,
    temp_image_folder: Path,
    unique_headers_mapping: Dict[int, str],
    needs_images: bool,
    pdf_root: Path,
    max_revisions: int,
) -> Path:
    """
    Stream the combined CSV into a final Excel workbook with optional embedded area images.
    Column order for the base metadata block is controlled by OUTPUT_BASE_ORDER.
    """
    # Area + extra + dynamic revision headers
    area_headers = [unique_headers_mapping[i] for i in range(len(unique_headers_mapping))]
    extra_headers = ["Latest Revision", "Latest Description", "Latest Date"]
    revision_headers = [f"Rev{i+1}" for i in range(max_revisions)]

    # Incoming base order fixed by the pipeline/combiner
    IN_BASE_ORDER = ["Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No", "Page Size"]

    # Sanity check: OUTPUT_BASE_ORDER must be a permutation of IN_BASE_ORDER
    if set(OUTPUT_BASE_ORDER) != set(IN_BASE_ORDER):
        raise ValueError(
            "OUTPUT_BASE_ORDER must contain exactly these keys (any order): "
            + ", ".join(IN_BASE_ORDER)
        )

    # Build header row (final Excel order)
    headers = ["UNID"] + OUTPUT_BASE_ORDER + area_headers + extra_headers + revision_headers

    # Workbook: write_only mode is incompatible with embedding images
    wb = Workbook(write_only=not needs_images)
    ws = wb.create_sheet("Sheet1") if wb.write_only else wb.active
    if not wb.write_only:
        ws.title = "Sheet1"
    ws.append(headers)

    # Lookups for mapping incoming -> outgoing positions
    name_to_incoming_idx = {name: i for i, name in enumerate(IN_BASE_ORDER)}

    # Helpful lookups for where things are in the *final* Excel:
    # Guard: ensure columns required for links/images exist
    for required in ("Folder", "Filename", "Page No"):
        if required not in headers:
            raise ValueError(f"Required column missing in headers: {required}")

    filename_col_idx0 = headers.index("Filename")
    folder_col_idx0 = headers.index("Folder")
    page_no_idx0 = headers.index("Page No")
    area_col_idxs0 = [headers.index(h) for h in area_headers]
    ocr_font = Font(color="FF3300")

    with open(combined_csv, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip CSV header from combiner
        for row in reader:
            # row layout from combiner:
            # [UNID, Size, DateMod, Folder, Filename, PageNo, PageSize, <areas...>, LatestRev, LatestDesc, LatestDate, Rev1, ..., RevN]
            unid = row[0]
            incoming_base = row[1:7]  # fixed order IN_BASE_ORDER

            # Reorder base columns to match OUTPUT_BASE_ORDER
            base_by_name = {k: incoming_base[name_to_incoming_idx[k]] for k in IN_BASE_ORDER}
            reordered_base = [base_by_name[name] for name in OUTPUT_BASE_ORDER]

            # Areas slice
            num_areas = len(area_headers)
            areas = row[7 : 7 + num_areas]

            # Latest fields
            base_index = 7 + num_areas
            latest_rev = row[base_index] if len(row) > base_index else ""
            latest_desc = row[base_index + 1] if len(row) > base_index + 1 else ""
            latest_date = row[base_index + 2] if len(row) > base_index + 2 else ""

            # Revision columns follow the latest fields
            revision_start = base_index + 3
            if max_revisions:
                padded = row[revision_start : revision_start + max_revisions]
                if len(padded) < max_revisions:
                    padded = list(padded) + [""] * (max_revisions - len(padded))
                else:
                    padded = list(padded)
            else:
                padded = []

            # Build final row (same order as headers)
            row_values = [
                _clean_cell_value(v)
                for v in ([unid] + reordered_base + areas + [latest_rev, latest_desc, latest_date] + padded)
            ]

            if needs_images:
                # Normal mode: we can style cells and embed images
                ws.append(row_values)
                r = ws.max_row

                # Hyperlink on Filename
                folder_val = row_values[folder_col_idx0]
                filename_val = row_values[filename_col_idx0]
                if folder_val and filename_val:
                    abs_path = os.path.abspath(pdf_root / folder_val / filename_val)
                    cell = ws.cell(row=r, column=filename_col_idx0 + 1)
                    cell.hyperlink = abs_path
                    cell.font = Font(color="0000FF", underline="single")

                # OCR red + image anchoring for each area column
                page_no_val = row_values[page_no_idx0]
                for i, col_idx0 in enumerate(area_col_idxs0):
                    cell = ws.cell(row=r, column=col_idx0 + 1)
                    if isinstance(cell.value, str) and "_OCR_" in cell.value:
                        cell.font = ocr_font
                        cell.value = cell.value.replace("_OCR_", "").strip()
                    try:
                        if folder_val and filename_val and page_no_val:
                            image_filename = f"{filename_val}_page{page_no_val}_area{i}.png"
                            img_path = temp_image_folder / image_filename
                            if img_path.exists():
                                img = ExcelImage(str(img_path))
                                img.anchor = f"{get_column_letter(col_idx0 + 1)}{r}"
                                ws.add_image(img)
                    except Exception:
                        # Ignore image issues but keep data
                        pass

            else:
                # Write-only mode: use WriteOnlyCell; images not supported
                row_cells = []
                folder_val = row_values[folder_col_idx0]
                filename_val = row_values[filename_col_idx0]
                abs_path = (
                    os.path.abspath(pdf_root / folder_val / filename_val)
                    if folder_val and filename_val
                    else None
                )

                for idx0, val in enumerate(row_values):
                    clean_val = _clean_cell_value(val)
                    c = WriteOnlyCell(ws, value=clean_val)
                    # Hyperlink on Filename
                    if idx0 == filename_col_idx0 and abs_path:
                        c.font = Font(color="0000FF", underline="single")
                        c.hyperlink = abs_path
                    # OCR marking for area cells
                    if idx0 in area_col_idxs0 and isinstance(clean_val, str) and "_OCR_" in clean_val:
                        c.value = clean_val.replace("_OCR_", "").strip()
                        c.font = ocr_font
                    row_cells.append(c)
                ws.append(row_cells)

    # Versioned filename if the target already exists
    final_out = Path(out_path)
    if final_out.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = final_out.with_suffix("")  # drop suffix
        final_out = final_out.with_name(f"{stem.name}_{ts}{final_out.suffix}")

    wb.save(str(final_out))
    return final_out

