import os
import pandas as pd
import customtkinter as ctk
import tkinter as tk

from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from tkinterdnd2 import TkinterDnD, DND_ALL


class CTkDnDEntry(ctk.CTkEntry, TkinterDnD.DnDWrapper):
    def __init__(self, *args, **kwargs):
        ctk.CTkEntry.__init__(self, *args, **kwargs)
        TkinterDnD.DnDWrapper.__init__(self)
        self.drop_target_register(DND_ALL)


def list_files(folder_path):
    file_dict = {'PDF': {}, 'DWG': {}}
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            path = os.path.join(root, file)
            name, ext = os.path.splitext(file)
            ext = ext.lower()
            if ext == '.pdf':
                file_dict['PDF'][path] = {
                    'name': name,
                    'size': os.path.getsize(path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(path)),
                    'path': path
                }
            elif ext == '.dwg':
                file_dict['DWG'][path] = {
                    'name': name,
                    'size': os.path.getsize(path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(path)),
                    'path': path
                }
    return file_dict


def prepare_data_for_export(file_dict, selected_folder):
    combined_dict = {}

    def update_entry(file_name, pdf_data=None, dwg_data=None):
        if file_name not in combined_dict:
            combined_dict[file_name] = {
                'PDF': None, 'DWG': None, 'FolderPDF': [], 'FolderDWG': [],
                'PDFSize': None, 'DWGSize': None, 'PDFModified': None, 'DWGModified': None,
                'PDFDuplicateCount': None, 'DWGDuplicateCount': None
            }

        if pdf_data:
            combined_dict[file_name]['PDF'] = pdf_data['name']
            combined_dict[file_name]['PDFSize'] = pdf_data['size']
            combined_dict[file_name]['PDFModified'] = pdf_data['modified']
            combined_dict[file_name]['FolderPDF'].append(os.path.relpath(os.path.dirname(pdf_data['path']), selected_folder))
            if len(combined_dict[file_name]['FolderPDF']) > 1:
                combined_dict[file_name]['PDFDuplicateCount'] = len(combined_dict[file_name]['FolderPDF'])

        if dwg_data:
            combined_dict[file_name]['DWG'] = dwg_data['name']
            combined_dict[file_name]['DWGSize'] = dwg_data['size']
            combined_dict[file_name]['DWGModified'] = dwg_data['modified']
            combined_dict[file_name]['FolderDWG'].append(os.path.relpath(os.path.dirname(dwg_data['path']), selected_folder))
            if len(combined_dict[file_name]['FolderDWG']) > 1:
                combined_dict[file_name]['DWGDuplicateCount'] = len(combined_dict[file_name]['FolderDWG'])

    pdfs = file_dict['PDF']
    dwgs = file_dict['DWG']

    for path, data in pdfs.items():
        name = data['name']
        update_entry(name, pdf_data=data)

    for path, data in dwgs.items():
        name = data['name']
        update_entry(name, dwg_data=data)

    df = pd.DataFrame.from_dict(combined_dict, orient='index').reset_index(drop=True)
    df['FolderPDF'] = df['FolderPDF'].apply(lambda x: ', '.join(x) if x else None)
    df['FolderDWG'] = df['FolderDWG'].apply(lambda x: ', '.join(x) if x else None)
    return df


def save_to_excel(df, path):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in ['PDFDuplicateCount', 'DWGDuplicateCount']:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1
            col_letter = chr(64 + col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(df)+1}",
                CellIsRule(operator="greaterThan", formula=["1"], stopIfTrue=True, fill=red_fill)
            )

    wb.save(path)
    wb.close()


class PDFDWGCheckerApp:
    def __init__(self, master):
        self.root = ctk.CTkToplevel(master)
        self.root.title("PDF & DWG File Checker")
        self.root.geometry("620x380")
        self.root.resizable(False, False)

        # Schedule raising and focusing AFTER the window initializes
        self.root.after(200, self._raise_and_focus)

        self.same_folder = ctk.BooleanVar(value=True)
        self.build_ui()

    def _raise_and_focus(self):
        self.root.attributes('-topmost', True)  # Temporarily stay on top
        self.root.lift()                        # Raise window
        self.root.focus_force()                 # Force focus
        # Remove always-on-top after 300ms
        self.root.after(300, lambda: self.root.attributes('-topmost', False))

    def build_ui(self):
        pdf_frame = ctk.CTkFrame(self.root)
        pdf_frame.pack(padx=15, pady=(20, 5), fill="x")

        ctk.CTkLabel(pdf_frame, text="📁 PDF Folder:", width=100).pack(side="left", padx=(0, 10))
        self.pdf_entry = CTkDnDEntry(pdf_frame)
        self.pdf_entry.dnd_bind('<<Drop>>', self.on_pdf_drop)
        self.pdf_entry.pack(side="left", expand=True, fill="x", padx=(0, 10))
        ctk.CTkButton(pdf_frame, text="Browse", width=70, command=self.browse_pdf).pack(side="left")

        self.folder_checkbox = ctk.CTkCheckBox(self.root, text="DWG folder is same as PDF folder",
                                               variable=self.same_folder, command=self.toggle_dwg_folder)
        self.folder_checkbox.pack(anchor="w", padx=25, pady=(0, 10))

        dwg_frame = ctk.CTkFrame(self.root)
        dwg_frame.pack(padx=15, pady=5, fill="x")

        ctk.CTkLabel(dwg_frame, text="📁 DWG Folder:", width=100).pack(side="left", padx=(0, 10))
        self.dwg_entry = CTkDnDEntry(dwg_frame, state="disabled")
        self.dwg_entry.dnd_bind('<<Drop>>', self.on_dwg_drop)
        self.dwg_entry.pack(side="left", expand=True, fill="x", padx=(0, 10))
        self.dwg_entry.configure(fg_color="gray20", text_color="gray70")
        self.dwg_button = ctk.CTkButton(dwg_frame, text="Browse", width=70, command=self.browse_dwg, state="disabled")
        self.dwg_button.pack(side="left")
        self.dwg_button.configure(fg_color="gray25", text_color="gray70")

        # Textbox for summary display
        # self.summary_box = ctk.CTkTextbox(self.root, width=580, height=120, font=("Verdana", 12))
        # self.summary_box.configure(state="disabled")
        # self.summary_box.pack(padx=15, pady=(10, 10))
        self.summary_box = tk.Text(self.root, width=70, height=7, font=("Verdana", 12), wrap="word")
        self.summary_box.configure(state="disabled")
        self.summary_box.pack(padx=15, pady=(10, 10))

        self.export_button = ctk.CTkButton(self.root, text="✅ Generate Excel Report", command=self.run_check,
                                           width=200, height=40, font=("Arial", 14))
        self.export_button.pack(pady=(0, 15))

    def toggle_dwg_folder(self):
        if self.same_folder.get():
            self.dwg_entry.configure(state="disabled", fg_color="gray20", text_color="gray70")
            self.dwg_button.configure(state="disabled", fg_color="gray25", text_color="gray70")
        else:
            self.dwg_entry.configure(state="normal", fg_color="gray15", text_color="white")
            self.dwg_button.configure(state="normal", fg_color="#3A7EBF", text_color="white")

    def on_pdf_drop(self, event):
        path = event.data.strip().replace("{", "").replace("}", "")
        if os.path.isdir(path):
            self.pdf_entry.delete(0, ctk.END)
            self.pdf_entry.insert(0, path)
            self.update_summary()
        else:
            messagebox.showerror("Invalid Drop", "Please drop a valid folder.")

    def on_dwg_drop(self, event):
        path = event.data.strip().replace("{", "").replace("}", "")
        if os.path.isdir(path):
            self.dwg_entry.delete(0, ctk.END)
            self.dwg_entry.insert(0, path)
            self.update_summary()
        else:
            messagebox.showerror("Invalid Drop", "Please drop a valid folder.")

    def browse_pdf(self):
        path = filedialog.askdirectory(title="Select PDF Folder")
        if path:
            self.pdf_entry.delete(0, "end")
            self.pdf_entry.insert(0, path)
            self.update_summary()
            self.root.after(200, self._raise_and_focus)  # Raise and focus window

    def browse_dwg(self):
        path = filedialog.askdirectory(title="Select DWG Folder")
        if path:
            self.dwg_entry.delete(0, "end")
            self.dwg_entry.insert(0, path)
            self.update_summary()
            self.root.after(200, self._raise_and_focus)  # Raise and focus window

    def update_summary(self):
        pdf_folder = self.pdf_entry.get()
        dwg_folder = self.dwg_entry.get() if not self.same_folder.get() else pdf_folder

        if not os.path.isdir(pdf_folder) or not os.path.isdir(dwg_folder):
            self._set_summary_text("Please select valid PDF and DWG folders.")
            return

        pdf_files = list_files(pdf_folder)['PDF']
        dwg_files = list_files(dwg_folder)['DWG']

        pdf_names = {data['name'] for data in pdf_files.values()}
        dwg_names = {data['name'] for data in dwg_files.values()}

        matching = pdf_names.intersection(dwg_names)
        pdf_no_match = pdf_names - matching
        dwg_no_match = dwg_names - matching

        duplicates_pdf = sum(1 for d in pdf_files.values() if d['size'] == 0)
        duplicates_dwg = sum(1 for d in dwg_files.values() if d['size'] == 0)

        # Count duplicates: files with the same name appearing multiple times
        pdf_name_counts = {}
        for d in pdf_files.values():
            pdf_name_counts[d['name']] = pdf_name_counts.get(d['name'], 0) + 1
        dwg_name_counts = {}
        for d in dwg_files.values():
            dwg_name_counts[d['name']] = dwg_name_counts.get(d['name'], 0) + 1

        duplicates_count_pdf = sum(count - 1 for count in pdf_name_counts.values() if count > 1)
        duplicates_count_dwg = sum(count - 1 for count in dwg_name_counts.values() if count > 1)

        corrupted_pdf = sum(1 for d in pdf_files.values() if d['size'] == 0)
        corrupted_dwg = sum(1 for d in dwg_files.values() if d['size'] == 0)

        summary_text = (
            f"Matching files: {len(matching)}\n"
            f"No Match (PDF): {len(pdf_no_match)}\n"
            f"No Match (DWG): {len(dwg_no_match)}\n"
            f"Duplicates: PDF({duplicates_count_pdf}), DWG({duplicates_count_dwg})\n"
            f"Corrupted (0 bytes): PDF({corrupted_pdf}), DWG({corrupted_dwg})"
        )

        self._set_summary_text(summary_text)

    def _set_summary_text(self, text):
        self.summary_box.configure(state="normal")
        self.summary_box.delete("1.0", "end")
        self.summary_box.insert("1.0", text)

        # Configure tags
        self.summary_box.tag_configure("green", foreground="green", font=("Verdana", 12, "normal"))
        self.summary_box.tag_configure("red_bold", foreground="red", font=("Verdana", 12, "bold"))

        import re

        # Find all numbers in the text with their positions
        matches = list(re.finditer(r"\b(\d+)\b", text))
        if not matches:
            self.summary_box.configure(state="disabled")
            return

        # First number corresponds to Matching files -> green (normal)
        first_match = matches[0]
        start_index = f"1.0 + {first_match.start()} chars"
        end_index = f"1.0 + {first_match.end()} chars"
        self.summary_box.tag_add("green", start_index, end_index)

        # All other numbers > 0 get red bold
        for match in matches[1:]:
            num_val = int(match.group(1))
            if num_val > 0:
                start_index = f"1.0 + {match.start()} chars"
                end_index = f"1.0 + {match.end()} chars"
                self.summary_box.tag_add("red_bold", start_index, end_index)

        self.summary_box.configure(state="disabled")

    def run_check(self):
        pdf_folder = self.pdf_entry.get()
        dwg_folder = self.dwg_entry.get() if not self.same_folder.get() else pdf_folder

        if not os.path.isdir(pdf_folder) or not os.path.isdir(dwg_folder):
            messagebox.showerror("Invalid Input", "Please select valid folder paths.")
            return

        pdf_files = list_files(pdf_folder)['PDF']
        dwg_files = list_files(dwg_folder)['DWG']
        combined = {'PDF': pdf_files, 'DWG': dwg_files}

        df = prepare_data_for_export(combined, pdf_folder)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            save_to_excel(df, save_path)
            messagebox.showinfo("Done", f"Excel file saved to:\n{save_path}")
            os.startfile(save_path)

    def run(self):
        self.root.mainloop()


def launch_pdf_dwg_gui(master):
    PDFDWGCheckerApp(master)


if __name__ == "__main__":
    ctk.set_appearance_mode("dark")
    from tkinterdnd2 import TkinterDnD
    app = TkinterDnD.Tk()
    app.geometry("800x600")
    app.title("Main App Window")
    PDFDWGCheckerApp(app)
    app.mainloop()
