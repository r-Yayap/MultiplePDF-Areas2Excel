#!/usr/bin/env python3
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog


def list_files_in_directory(selected_folder):
    file_list = []
    for root, dirs, files in os.walk(selected_folder):
        for file in files:
            full_path = os.path.join(root, file)
            relative_path = os.path.relpath(full_path, selected_folder)
            folder_name = os.path.dirname(relative_path)
            filename, file_format = os.path.splitext(file)

            # Stat + timestamp conversion all in one try/except
            try:
                st = os.stat(full_path)
                size = st.st_size
                mtime = st.st_mtime
                last_modified = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            except OSError as e:
                print(f"Skipping {full_path!r}: {e}")
                continue

            file_list.append({
                'Folder': folder_name,
                'Filename_xtn': filename + file_format,
                'Filename': filename,
                'Full Path': full_path,
                'Format': file_format.lstrip('.'),
                'Size (Bytes)': size,
                'Last Modified': last_modified
            })
    return file_list


def create_excel_file(file_list, output_excel_path):
    # Build DataFrame and drop full path column
    df = pd.DataFrame(file_list)
    df_without_path = df.drop(columns=['Full Path'])

    # Save to Excel and add hyperlinks
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_without_path.to_excel(writer, index=False, sheet_name='Files')
        workbook = writer.book
        worksheet = writer.sheets['Files']

        # Hyperlink only on the 'Filename' column
        fname_col = df_without_path.columns.get_loc('Filename') + 1
        for row_idx, info in enumerate(file_list, start=2):
            cell = worksheet[f"{get_column_letter(fname_col)}{row_idx}"]
            cell.hyperlink = info['Full Path']
            cell.style = "Hyperlink"

    print(f"Excel file with hyperlinks created at {output_excel_path}")


from tkinter import filedialog  # already imported at top
def select_input_folder(parent=None):
    return filedialog.askdirectory(title="Select Input Folder", parent=parent)



def generate_file_list_and_excel(parent=None):
    input_folder = select_input_folder(parent=parent)
    if not input_folder:
        return

    output_file = filedialog.asksaveasfilename(
        title="Save As", defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        , parent=parent
    )
    if not output_file:
        return
    if not output_file.lower().endswith('.xlsx'):
        output_file += '.xlsx'

    files = list_files_in_directory(input_folder)
    create_excel_file(files, output_file)
    print(f"Excel file created at {output_file}")


if __name__ == '__main__':
    generate_file_list_and_excel()

