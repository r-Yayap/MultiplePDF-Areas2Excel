import os
import pandas as pd
from tkinter import Tk, filedialog
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def list_files_in_directory(selected_folder):
    file_list = []
    for root, dirs, files in os.walk(selected_folder):
        for file in files:
            full_path = os.path.join(root, file)
            relative_path = os.path.relpath(full_path, selected_folder)
            folder_name = os.path.dirname(relative_path)
            filename, file_format = os.path.splitext(file)

            # Get file size and last modified date
            file_size = os.path.getsize(full_path)  # Size in bytes
            last_modified_timestamp = os.path.getmtime(full_path)
            last_modified_date = datetime.fromtimestamp(last_modified_timestamp).strftime('%Y-%m-%d %H:%M:%S')

            file_list.append({
                'Folder': folder_name,
                'Filename': filename,
                'Full Path': full_path,  # Include the full path for hyperlink
                'Format': file_format[1:],
                'Size (Bytes)': file_size,
                'Last Modified': last_modified_date
            })

    return file_list


def create_excel_file(file_list, output_excel_path):
    # Create a DataFrame from the file list
    df = pd.DataFrame(file_list)

    # Remove 'Full Path' before saving, since we will handle it separately for hyperlinks
    df_without_path = df.drop(columns=['Full Path'])

    # Save the DataFrame to an Excel file
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_without_path.to_excel(writer, index=False, sheet_name='Files')
        workbook = writer.book
        worksheet = writer.sheets['Files']

        # Add hyperlinks to the 'Filename' column
        for row_idx, file_info in enumerate(file_list, start=2):  # Start from row 2 to skip the header
            # Get the filename and its full path
            filename = file_info['Filename']
            full_path = file_info['Full Path']

            # Create the hyperlink in the 'Filename' column (assumed to be column B)
            col_letter = get_column_letter(df_without_path.columns.get_loc('Filename') + 1)
            cell = worksheet[f"{col_letter}{row_idx}"]
            cell.hyperlink = full_path  # Set the hyperlink
            cell.style = "Hyperlink"  # Apply hyperlink style

    print(f"Excel file with hyperlinks created at {output_excel_path}")


def select_input_folder():
    root = Tk()
    root.withdraw()  # Hide the main window

    input_folder = filedialog.askdirectory(title="Select Input Folder")
    return input_folder


def generate_file_list_and_excel():
    input_folder = select_input_folder()

    if input_folder:  # Check if folder is selected
        files = list_files_in_directory(input_folder)

        # Prompt the user for the output file name
        output_file_name = filedialog.asksaveasfilename(
            title="Save As",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )

        if output_file_name:
            # Ensure the file has the correct extension
            if not output_file_name.lower().endswith('.xlsx'):
                output_file_name += '.xlsx'

            # Construct the output_excel_path using output_file_name directly
            output_excel_path = os.path.join(os.path.dirname(output_file_name), output_file_name)
            create_excel_file(files, output_excel_path)
            print(f"Excel file created at {output_excel_path}")
